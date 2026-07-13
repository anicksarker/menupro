#!/usr/bin/env python3
# =============================================================================
#  build_wl_model_31_32.py   (Plan 31/32 - Paid-Up Whole Life)
# -----------------------------------------------------------------------------
#  Rebuilds the workbook from ONLY two CSV inputs:
#         1) Inputs.csv     (structured: scalars + dividend table + loan table)
#         2) cso_rates.csv  (mortality rate matrix)
#
#  Architecture notes for this version:
#    * 12 sheets: CSO Rates, Output, Inputs, MORTALITY, Commutation, CV_RATES,
#      PUA_Div, Projection_Monthly, RPU, ETI, APL, Loan.
#    * Output sheet = the summary (Policy No, Cash Value, PUA CV, ETI, RPU, APL,
#      Loan Outstanding) in columns A..G. The external "Audit Report" /
#      "Balancing Report" columns (which XLOOKUP into other workbooks) are
#      intentionally NOT generated - they need external audit files and are not
#      part of the input-driven model.
#    * Inputs sheet is fully CSV-driven and now carries THREE blocks, all pasted
#      positionally from Inputs.csv (blank separator rows preserved):
#         - scalars (rows 2..28)      - formula cells are overwritten
#         - dividend history table    - Div Date / Div Amount
#         - loan table                - Loan Date / Loan Amount
#      Column A can hold a label OR a date; both are parsed. Nothing is hardcoded.
#    * CSO Rates = pasted from cso_rates.csv.  Every other sheet = pure formulas.
#
#  Repeated (fill-down) formulas are written with a LOOP via Excel-style
#  translation (relative refs shift, absolute $refs stay, integer counters bump).
#
#  Usage:
#     pip install openpyxl pandas
#     python build_wl_model_31_32.py Inputs.csv cso_rates.csv out.xlsx
# =============================================================================
import sys, csv, re, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

INPUTS_CSV = sys.argv[1] if len(sys.argv) > 1 else "Inputs.csv"
CSO_CSV    = sys.argv[2] if len(sys.argv) > 2 else "cso_rates.csv"
OUT_XLSX   = sys.argv[3] if len(sys.argv) > 3 else "WL_Model_Paidup_31_32.xlsx"

SHEET_ORDER = ['CSO Rates','Output','Inputs','MORTALITY','Commutation','CV_RATES','PUA_Div',
               'Projection_Monthly','RPU','ETI','APL','Loan']
INPUTS_HEADER_ROW = 1                 # Inputs.csv line 0 -> Excel row 2
FORCE_TEXT_LABELS = {"Policy Number"} # identifiers kept as text even if all-digits

# =============================================================================
#  Fill / helper engine
# =============================================================================
_INT = re.compile(r'(?<![A-Za-z$0-9_.])\d+')

def _bump_ints(s, positions, delta):
    out, last = [], 0
    for i, m in enumerate(_INT.finditer(s)):
        if i in positions:
            out.append(s[last:m.start()]); out.append(str(int(m.group()) + delta)); last = m.end()
    out.append(s[last:]); return ''.join(out)

def _translate(formula, origin, dest):
    return Translator(formula, origin=origin).translate_formula(dest)

def _dec(v):
    if isinstance(v, list) and len(v) == 2 and v[0] == "__dt__":
        return datetime.datetime.fromisoformat(v[1])
    return v

def _cell(x):
    """Type a raw CSV string for the CSO matrix."""
    if x is None: return None
    x = str(x).strip()
    if x == "" or x.lower() == "nan": return None
    try:
        i = int(x)
        if str(i) == x: return i
    except ValueError: pass
    try: return float(x)
    except ValueError: return x

_DATE_RE = re.compile(r'^\d{1,2}-\d{1,2}-\d{4}$')
def _infer(x):
    """Infer the python type of a raw Inputs.csv cell (label, date, %, number...)."""
    x = (x or "").strip()
    if x == "": return None
    if _DATE_RE.match(x):
        for fmt in ("%d-%m-%Y", "%m-%d-%Y"):
            try: return datetime.datetime.strptime(x, fmt)
            except ValueError: pass
    if x.upper() in ("TRUE", "FALSE"): return x.upper() == "TRUE"
    if x.endswith("%"):
        try: return float(x[:-1].replace(",", "")) / 100.0
        except ValueError: return x
    xn = x.replace(",", "")                       # allow thousands separators e.g. 6,657.39
    if re.match(r'^-?[1-9]\d*$', xn) or xn == "0":
        return int(xn)
    if re.match(r'^-?\d*\.\d+$', xn):
        return float(xn)
    return x                                      # plain string

# =============================================================================
#  CSV loaders
# =============================================================================
def load_inputs(path):
    """Ordered list of (colA, colB) rows - blank separator rows KEPT so the row
    positions line up 1:1 with the Inputs sheet (scalars + tables)."""
    rows = []
    with open(path, newline="") as f:
        for row in csv.reader(f):
            a = row[0].strip() if row else ""
            b = row[1] if (row and len(row) > 1) else ""
            rows.append((a, b))
    while rows and rows[-1][0] == "" and str(rows[-1][1]).strip() == "":
        rows.pop()                                # drop only trailing empties
    return rows

def load_cso_df(path):
    return pd.read_csv(path, header=None, dtype=str, keep_default_na=False)

# =============================================================================
#  Sheet writer
# =============================================================================
def write_sheet(ws, name, spec, inputs=None, cso_df=None):
    written = set()
    def put(coord, value):
        ws[coord] = value
        written.add(coord)

    if name == "CSO Rates":
        for _, row in cso_df.iterrows():
            ws.append([_cell(v) for v in row.tolist()])
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    written.add(f"{get_column_letter(c)}{r}")

    elif name == "Inputs" and inputs is not None:
        # paste col A (labels / dates) and col B (values) positionally
        for i, (cola, colb) in enumerate(inputs):
            r = INPUTS_HEADER_ROW + 1 + i
            a = _infer(cola)
            if a is not None:
                put(f"A{r}", a)
            if cola in FORCE_TEXT_LABELS:          # e.g. policy number stays text
                b = colb.strip() if colb and colb.strip() else None
            else:
                b = _infer(colb)
            if b is not None:
                put(f"B{r}", b)
        # formula-driven B cells get overwritten by the fills below

    else:
        for coord, val in spec["statics"].items():
            put(coord, _dec(val))

    # formula fills - LOOP each column's origin formula down its rows
    arrays = set(spec.get("arrays", []))
    for col, r0, r1, kind, f0, pos in spec["fills"]:
        origin = f"{col}{r0}"
        for r in range(r0, r1 + 1):
            coord = f"{col}{r}"
            if kind == "lit":
                formula = f0
            elif kind == "trans":
                formula = _translate(f0, origin, coord)
            else:
                formula = _bump_ints(_translate(f0, origin, coord), pos, r - r0)
            put(coord, ArrayFormula(coord, formula) if coord in arrays else formula)

    # number formats + column widths
    colfmt, exc = spec.get("colfmt", {}), spec.get("fmt_exc", {})
    for coord in written:
        col = re.match(r"[A-Z]+", coord).group()
        fmt = exc.get(coord, colfmt.get(col))
        if fmt:
            ws[coord].number_format = fmt
    for col, w in spec.get("widths", {}).items():
        ws.column_dimensions[col].width = w

# =============================================================================
def main():
    inputs = load_inputs(INPUTS_CSV)
    cso_df = load_cso_df(CSO_CSV)
    wb = Workbook(); wb.remove(wb.active)
    for name in SHEET_ORDER:
        print(f"  building sheet: {name}")
        write_sheet(wb.create_sheet(title=name), name, MODEL_SPEC[name], inputs=inputs, cso_df=cso_df)
    wb.save(OUT_XLSX)
    print(f"\nSaved: {OUT_XLSX}")
    print("Open in Excel to recalculate. Summary is on the Output sheet (A3:G3):")
    print("Policy No, Cash Value, PUA Cash Value, ETI, RPU, APL, Loan Outstanding.")


# =============================================================================
#  MODEL DEFINITION  (fully editable - every formula & constant is below)
# =============================================================================
MODEL_SPEC = {
  'CSO Rates': {
    "statics": {},
    "arrays": [],
    "fills": [
      ['B', 4, 4, 'trans', '=_xlfn.CONCAT(B1,B2,B3)', []],
      ['C', 4, 4, 'trans', '=_xlfn.CONCAT(C1,C2,C3)', []],
      ['D', 4, 4, 'trans', '=_xlfn.CONCAT(D1,D2,D3)', []],
      ['E', 4, 4, 'trans', '=_xlfn.CONCAT(E1,E2,E3)', []],
      ['F', 4, 4, 'trans', '=_xlfn.CONCAT(F1,F2,F3)', []],
      ['G', 4, 4, 'trans', '=_xlfn.CONCAT(G1,G2,G3)', []],
      ['H', 4, 4, 'trans', '=_xlfn.CONCAT(H1,H2,H3)', []],
      ['I', 4, 4, 'trans', '=_xlfn.CONCAT(I1,I2,I3)', []],
      ['J', 4, 4, 'trans', '=_xlfn.CONCAT(J1,J2,J3)', []],
      ['K', 4, 4, 'trans', '=_xlfn.CONCAT(K1,K2,K3)', []],
      ['L', 4, 4, 'trans', '=_xlfn.CONCAT(L1,L2,L3)', []],
      ['M', 4, 4, 'trans', '=_xlfn.CONCAT(M1,M2,M3)', []],
      ['N', 4, 4, 'trans', '=_xlfn.CONCAT(N1,N2,N3)', []],
      ['O', 4, 4, 'trans', '=_xlfn.CONCAT(O1,O2,O3)', []],
      ['P', 4, 4, 'trans', '=_xlfn.CONCAT(P1,P2,P3)', []],
    ],
    "colfmt": {},
    "fmt_exc": {},
    "widths": {'A': 8.73},
  },
  'Output': {
    "statics": {
        'B1': 'Calculated values',
        'A2': 'Policy Number',
        'B2': 'Cash Value',
        'C2': 'PUA Cash Value',
        'D2': 'ETI',
        'E2': 'RPU',
        'F2': 'APL',
        'G2': 'Loan Outstanding',
    },
    "arrays": ['B3'],
    "fills": [
      ['A', 3, 3, 'trans', '=Inputs!B5', []],
      ['B', 3, 3, 'lit', '=_xlfn.LET(\n    _xlpm.ValDate, Inputs!B12,\n    _xlpm.DateRange, Projection_Monthly!$B$2:$B$1050,\n    _xlpm.CVRange, Projection_Monthly!$L$2:$L$1050,\n    _xlpm.ExactMatch, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.CVRange, "", 0),\n    IF(_xlpm.ExactMatch<>"", _xlpm.ExactMatch,\n        _xlfn.LET(\n            _xlpm.DateBefore, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", -1),\n            _xlpm.DateAfter, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", 1),\n            _xlpm.CVBefore, _xlfn.XLOOKUP(_xlpm.DateBefore, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.CVAfter, _xlfn.XLOOKUP(_xlpm.DateAfter, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.DaysBetween, _xlpm.DateAfter - _xlpm.DateBefore,\n            _xlpm.DaysFromBefore, _xlpm.ValDate - _xlpm.DateBefore,\n            _xlpm.Fraction, _xlpm.DaysFromBefore / _xlpm.DaysBetween,\n            _xlpm.CVBefore + _xlpm.Fraction * (_xlpm.CVAfter - _xlpm.CVBefore)\n        )\n    )\n)', []],
      ['C', 3, 3, 'trans', '=IF(Inputs!$B$20="Buy PUA",PUA_Div!$E$14,0)', []],
      ['D', 3, 3, 'trans', '=ETI!E6', []],
      ['E', 3, 3, 'trans', '=RPU!B17', []],
      ['F', 3, 3, 'trans', '=APL!F10', []],
      ['G', 3, 3, 'trans', '=IF(Inputs!B39="D",Loan!$G$10,Loan!$G$11)', []],
    ],
    "colfmt": {},
    "fmt_exc": {
        'A3': '#,##0.00',
        'B3': '#,##0.00',
        'C3': '#,##0.00',
        'D3': 'mm-dd-yy',
        'E3': '#,##0.00',
        'F3': 'mm-dd-yy',
        'G3': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
    },
    "widths": {'A': 12.91, 'B': 16.54},
  },
  'Inputs': {
    "statics": {},
    "arrays": ['B25'],
    "fills": [
      ['B', 7, 7, 'trans', '=ROUND((B11-B14)/365.25,0)', []],
      ['B', 13, 13, 'trans', '=IF(DATE(YEAR($B$12),MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12),MONTH($B$11),1),0))))\n<=$B$12,\nDATE(YEAR($B$12),MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12),MONTH($B$11),1),0)))),\nDATE(YEAR($B$12)-1,MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12)-1,MONTH($B$11),1),0)))))', []],
      ['B', 15, 15, 'trans', '=ROUND(YEAR(B13)-YEAR(B14)+(MONTH(B13)-MONTH(B14))/12+(DAY(B13)-DAY(B14)-1)/360,0)', []],
      ['B', 16, 16, 'trans', '=INT((B12-B11)/365.25)', []],
      ['B', 17, 17, 'trans', '=B16+1', []],
      ['B', 18, 18, 'trans', '=B9*1000', []],
      ['B', 23, 23, 'trans', '=B22*B25', []],
      ['B', 25, 25, 'lit', '=_xlfn.SWITCH(B24,"A",1,"S",0.5,"Q",0.25,"M",1/12,1)', []],
      ['B', 27, 27, 'trans', '=MAX(0, MIN(B8 - B7, B16))', []],
      ['B', 28, 28, 'trans', '=B15 < B8', []],
    ],
    "colfmt": {},
    "fmt_exc": {
        'B11': 'mm-dd-yy',
        'B12': 'mm-dd-yy',
        'B13': 'mm-dd-yy',
        'B14': 'mm-dd-yy',
        'B19': '0.00%',
        'B21': '0.00%',
        'B22': '0',
        'B23': '0',
        'B24': '0',
        'B25': '0',
        'B27': '0',
        'B28': '0',
        'A32': 'mm-dd-yy',
        'A33': 'mm-dd-yy',
        'A34': 'mm-dd-yy',
        'A35': 'mm-dd-yy',
        'B40': 'mm-dd-yy',
        'A42': 'mm-dd-yy',
        'A43': 'mm-dd-yy',
        'B43': '#,##0.00',
    },
    "widths": {'A': 31.54, 'B': 22.82},
  },
  'MORTALITY': {
    "statics": {
        'A1': 'Age',
        'C1': 'v*qx',
        'D1': 'v*q(x+1)*px',
        'A2': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 3, 123, 'trans', '=IFERROR(IF(B2>=1,"",A2+1),"")', []],
      ['B', 1, 1, 'trans', '=_xlfn.CONCAT(Inputs!B2,Inputs!B6,Inputs!B3)', []],
      ['B', 2, 123, 'trans', '=IFERROR(INDEX(\'CSO Rates\'!$A$5:$P$126,MATCH(A2,\'CSO Rates\'!$A$5:$A$126,0),MATCH($B$1,\'CSO Rates\'!$A$4:$P$4,0)),"")', []],
      ['C', 2, 123, 'trans', '=IFERROR(ETI!$B$7*B2,"")', []],
      ['D', 2, 123, 'trans', '=IFERROR(IF(A2="","",C2*(1-B3)),0)', []],
    ],
    "colfmt": {
        'B': '0.00000_)',
        'C': '0.00000_)',
        'D': '0.00000_)',
    },
    "fmt_exc": {
        'B1': 'General',
        'C1': 'General',
        'D1': 'General',
    },
    "widths": {'B': 18.82, 'C': 7.27, 'D': 11.73, 'E': 29.45},
  },
  'Commutation': {
    "statics": {
        'A1': 'Age',
        'B1': 'qx',
        'C1': 'lx',
        'D1': 'dx',
        'E1': 'v^x',
        'F1': 'Dx',
        'G1': 'Cx',
        'H1': 'Nx',
        'I1': 'Mx',
        'J1': 'ax_due',
        'K1': 'Ax',
        'L1': 'NLP_per_1000',
        'M1': 'LPP_NLP_per_1000',
        'N1': 'Alpha',
        'A2': 0,
        'C2': 10000000,
    },
    "arrays": [],
    "fills": [
      ['A', 3, 129, 'trans', '=IF(MORTALITY!A3="","",A2+1)', []],
      ['B', 2, 129, 'trans', '=IF(A2="","",IFERROR(INDEX(MORTALITY!$B:$B,MATCH(A2,MORTALITY!$A:$A,0)),0))', []],
      ['C', 3, 129, 'trans', '=IF(A3="","",C2*(1-B2))', []],
      ['D', 2, 129, 'trans', '=IF(A2="","",C2*B2)', []],
      ['E', 2, 129, 'trans', '=IF(A2="","",(1/(1+Inputs!$B$21))^A2)', []],
      ['F', 2, 129, 'trans', '=IF(A2="","",C2*E2)', []],
      ['G', 2, 129, 'trans', '=IF(A2="","",D2*E2*(1/(1+Inputs!$B$21)))', []],
      ['H', 2, 129, 'trans', '=IF(A2="","",IF(B2=1, F2, F2+H3))', []],
      ['I', 2, 129, 'trans', '=IF(A2="","",IF(B2=1, G2, G2+I3))', []],
      ['J', 2, 129, 'trans', '=IFERROR(IF(A2="","",H2/F2),1)', []],
      ['K', 2, 129, 'trans', '=IFERROR(IF(A2="","",I2/F2),1)', []],
      ['L', 2, 129, 'trans', '=IF(A2="","",K2/J2*1000)', []],
      ['M', 2, 129, 'trans', '=IF(A2="","",IF(A2>=Inputs!$B$8,0,I2/(H2-INDEX($H:$H,MATCH(Inputs!$B$8,$A:$A,0)))*1000))', []],
      ['N', 2, 129, 'trans', '=IF(A2="","",K2+Inputs!$B$19*(1-K2))', []],
    ],
    "colfmt": {
        'B': '0.00000_)',
        'C': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'D': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'E': '_(* #,##0.00000000_);_(* \\(#,##0.00000000\\);_(* \\-??_);_(@_)',
        'F': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'G': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'H': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'I': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'J': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
        'K': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
        'L': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
        'M': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
        'N': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
    },
    "fmt_exc": {
        'B1': 'General',
        'C1': 'General',
        'D1': 'General',
        'E1': 'General',
        'F1': 'General',
        'G1': 'General',
        'H1': 'General',
        'I1': 'General',
        'J1': 'General',
        'K1': 'General',
        'L1': 'General',
        'M1': 'General',
        'N1': 'General',
    },
    "widths": {'A': 9.45, 'B': 8.0, 'C': 13.54, 'D': 11.82, 'F': 13.54, 'G': 11.82, 'H': 14.54, 'I': 12.54, 'J': 11.82, 'K': 9.54, 'L': 12.82, 'M': 17.0, 'N': 9.54},
  },
  'CV_RATES': {
    "statics": {
        'A1': 'Code',
        'B1': 'Issue Age',
        'C1': 'Duration',
        'D1': 'Attained Age',
        'E1': 'qx',
        'F1': 'px',
        'G1': 'Interest Rate',
        'H1': 'Beta (NLP/1000)',
        'I1': 'CV (BOY)',
        'J1': 'CV (EOY)',
        'K1': 'CV FACTOR',
        'L1': 'Alpha',
        'M1': 'Expense Adj Premium',
        'C2': 0,
        'I2': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 2, 101, 'trans', '=IF(D2>100,"",Inputs!$B$4)', []],
      ['B', 2, 101, 'trans', '=IF(D2>100,"",Inputs!$B$7)', []],
      ['C', 3, 101, 'trans', '=IFERROR(IF(Inputs!$B$7+C2+1>=100,"",C2+1),"")', []],
      ['D', 2, 101, 'trans', '=IFERROR(IF(Inputs!$B$7+C2>=100,"",Inputs!$B$7+C2),"")', []],
      ['E', 2, 101, 'trans', '=IF(D2="","",INDEX(MORTALITY!$B:$B,MATCH(D2,MORTALITY!$A:$A,0)))', []],
      ['F', 2, 101, 'trans', '=IF(D2="","",1-E2)', []],
      ['G', 2, 101, 'trans', '=IF(D2="","",Inputs!$B$21)', []],
      ['H', 2, 101, 'trans', '=IF(D2="","",INDEX(Commutation!$L:$L,MATCH(Inputs!$B$7,Commutation!$A:$A,0)))', []],
      ['I', 3, 101, 'trans', '=IF(D3="","",J2)', []],
      ['J', 2, 101, 'trans', '=IF(D2>99,"",IF(E2=1,1000,MIN(1000,MAX(0,(INDEX(Commutation!$K:$K,MATCH(D3,Commutation!$A:$A,0))-(INDEX(Commutation!$K:$K,MATCH(Inputs!$B$7,Commutation!$A:$A,0))/((INDEX(Commutation!$H:$H,MATCH(Inputs!$B$7,Commutation!$A:$A,0))-INDEX(Commutation!$H:$H,MATCH(Inputs!$B$8,Commutation!$A:$A,0)))/INDEX(Commutation!$F:$F,MATCH(Inputs!$B$7,Commutation!$A:$A,0))))*MAX(0,(INDEX(Commutation!$H:$H,MATCH(D3,Commutation!$A:$A,0))-INDEX(Commutation!$H:$H,MATCH(Inputs!$B$8,Commutation!$A:$A,0)))/INDEX(Commutation!$F:$F,MATCH(D3,Commutation!$A:$A,0))))*1000))))', []],
      ['K', 2, 101, 'trans', '=J2', []],
      ['L', 2, 101, 'trans', '=IF(D2="","",IFERROR(INDEX(Commutation!$N:$N,MATCH(D2,Commutation!$A:$A,0)),1))', []],
      ['M', 2, 101, 'trans', '=IF(C2<(Inputs!$B$8-Inputs!$B$7),H2+Inputs!$B$19*(1-INDEX(Commutation!$K:$K,MATCH(Inputs!$B$7,Commutation!$A:$A,0)))*1000/(Inputs!$B$8-Inputs!$B$7),H2)', []],
    ],
    "colfmt": {
        'D': '0',
        'E': '0.00000_)',
        'F': '0.00000_)',
        'G': '0.00%',
    },
    "fmt_exc": {
        'E1': 'General',
        'F1': 'General',
        'G1': 'General',
    },
    "widths": {'A': 10.45, 'C': 13.45, 'D': 12.45, 'H': 14.27, 'I': 11.82, 'M': 19.18},
  },
  'PUA_Div': {
    "statics": {
        'A2': 'PUA Units',
        'A3': 'PUA Face Amount',
        'A4': 'Date of Birth',
        'A5': 'Valuation Date',
        'A7': 'Div Date (Anniversary Date)',
        'B7': 'Div Amount',
        'C7': 'Attained Age',
        'D7': 'PUA Single Premium',
        'E7': 'PUA Face Purchased',
        'F7': 'Total PUA Face',
        'A13': 'Last Anniverasry',
        'E13': 'PUA Cash Value',
    },
    "arrays": [],
    "fills": [
      ['A', 8, 11, 'trans', '=Inputs!A32', []],
      ['A', 14, 14, 'trans', '=Inputs!$B$13', []],
      ['B', 2, 2, 'trans', '=Inputs!$B$10', []],
      ['B', 3, 3, 'trans', '=B2*1000', []],
      ['B', 4, 4, 'trans', '=Inputs!$B$14', []],
      ['B', 5, 5, 'trans', '=Inputs!$B$12', []],
      ['B', 8, 11, 'trans', '=Inputs!B32', []],
      ['C', 8, 11, 'trans', '=ROUND(YEAR(A8)-YEAR($B$4)+(MONTH(A8)-MONTH($B$4))/12+(DAY(A8)-DAY($B$4)-1)/360,0)', []],
      ['C', 14, 14, 'trans', '=ROUND(YEAR(A14)-YEAR($B$4)+(MONTH(A14)-MONTH($B$4))/12+(DAY(A14)-DAY($B$4)-1)/360,0)', []],
      ['D', 8, 11, 'trans', '=IF(C8="","",INDEX(Commutation!$K:$K,MATCH(C8,Commutation!$A:$A,0)))', []],
      ['D', 14, 14, 'trans', '=IF(C14="","",INDEX(Commutation!$K:$K,MATCH(C14,Commutation!$A:$A,0)))', []],
      ['E', 8, 11, 'trans', '=B8/D8', []],
      ['E', 14, 14, 'trans', '=F10*D14', []],
      ['F', 8, 11, 'trans', '=$B$3+SUM($E$8:E8)', []],
    ],
    "colfmt": {
        'B': '0.00',
        'E': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'F': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)',
    },
    "fmt_exc": {
        'B2': 'General',
        'B3': 'General',
        'B4': 'mm-dd-yy',
        'B5': 'mm-dd-yy',
        'B7': 'General',
        'E7': 'General',
        'F7': 'General',
        'A8': 'mm-dd-yy',
        'A9': 'mm-dd-yy',
        'A10': 'mm-dd-yy',
        'A11': 'mm-dd-yy',
        'E13': 'General',
        'A14': 'mm-dd-yy',
        'E14': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)',
    },
    "widths": {'A': 24.18, 'B': 10.54, 'C': 11.45, 'D': 17.73, 'F': 13.18},
  },
  'Projection_Monthly': {
    "statics": {
        'A1': 'Plan',
        'B1': 'Date',
        'C1': 'Month Index',
        'D1': 'Duration',
        'E1': 'Attained Age',
        'F1': 'Last Anniv',
        'G1': 'Next Anniv',
        'H1': 'Frac within Year',
        'I1': 'CV Factor_n',
        'J1': 'CV Factor_n+1',
        'K1': 'CV per 1000',
        'L1': 'CV Total (Base)',
    },
    "arrays": [],
    "fills": [
      ['A', 2, 2, 'trans', '=IF(Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4)', []],
      ['A', 3, 985, 'trans', '=IF(A2="","",IF(INT((C2+1)/12)+Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4))', []],
      ['B', 2, 2, 'trans', '=IF(A2="","",Inputs!$B$11)', []],
      ['B', 3, 985, 'trans', '=IF(A3="","",EOMONTH(B2,1))', []],
      ['C', 2, 2, 'trans', '=IF(A2="","",0)', []],
      ['C', 3, 985, 'trans', '=IF(A3="","",C2+1)', []],
      ['D', 2, 985, 'trans', '=IF(A2="","",INT(C2/12))', []],
      ['E', 2, 985, 'trans', '=IF(A2="","",D2+Inputs!$B$7)', []],
      ['F', 2, 985, 'trans', '=IF(A2="","",DATE(YEAR(B$2)+D2,MONTH(B$2),DAY(B$2)))', []],
      ['G', 2, 985, 'trans', '=IF(A2="","",DATE(YEAR(F2)+1,MONTH(F2),DAY(F2)))', []],
      ['H', 2, 985, 'trans', '=IF(A2="","",IF(G2>F2,(B2-F2)/(G2-F2),0))', []],
      ['I', 2, 985, 'trans', '=IF(A2="","",IFERROR(INDEX(CV_RATES!$J:$J,MATCH(D2-1,CV_RATES!$C:$C,0)),0))', []],
      ['J', 2, 985, 'trans', '=IF(A2="","",IFERROR(IF(E2=MAX(MORTALITY!$A:$A),1000,INDEX(CV_RATES!$J:$J,MATCH(D2,CV_RATES!$C:$C,0))),0))', []],
      ['K', 2, 985, 'trans', '=IF(A2="","",IFERROR(I2+H2*(J2-I2),0))', []],
      ['L', 2, 985, 'trans', '=IF(A2="","",IFERROR(K2*Inputs!$B$9,0))', []],
    ],
    "colfmt": {
        'B': 'mm-dd-yy',
        'F': 'mm-dd-yy',
        'G': 'mm-dd-yy',
        'H': '0.00',
        'I': '0.00',
        'J': '0.00',
        'K': '0.00',
        'L': '0.00',
    },
    "fmt_exc": {
        'B1': 'General',
        'H1': 'General',
        'I1': 'General',
        'J1': 'General',
        'K1': 'General',
        'L1': 'General',
    },
    "widths": {'B': 11.0, 'C': 13.82, 'D': 10.45, 'E': 13.73, 'F': 11.54, 'G': 12.18, 'H': 16.45, 'I': 13.45, 'J': 15.0, 'K': 13.0, 'L': 15.54},
  },
  'RPU': {
    "statics": {
        'A1': 'Reduced Paid-Up (RPU) - Actuarial Calculation',
        'A3': 'Status',
        'A4': 'Attained Age',
        'A5': 'Original Sum Assured',
        'A7': '═══ Method 1: Cash Value / Net Single Premium ═══',
        'A8': 'Gross Cash Value (tVx)',
        'A9': 'Net Single Premium at Attained Age (Ax)',
        'A10': 'RPU SA (Uncapped) = CV / Ax',
        'A12': '═══ Method 2: Reserve Ratio (Actuarial Formula) ═══',
        'A13': 'Reserve per $1000 at Duration t',
        'A14': 'RPU SA = SA × (tVx/1000) / Ax',
        'A16': '═══════════════════════════════════════════════',
        'A17': 'FINAL RPU Sum Assured',
        'A19': 'Comparison:',
        'A20': 'Is Capping Applied?',
        'A22': 'Note: When CV/Ax > SA, capping prevents over-insurance.',
        'A23': 'Both methods should yield the same result when using',
        'A24': 'consistent actuarial assumptions.',
    },
    "arrays": [],
    "fills": [
      ['B', 3, 3, 'trans', '=IF(Inputs!$B$28=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")', []],
      ['B', 4, 4, 'trans', '=IF(Inputs!$B$28=FALSE,"",Inputs!$B$15)', []],
      ['B', 5, 5, 'trans', '=IF(Inputs!$B$28=FALSE,"",Inputs!$B$18)', []],
      ['B', 8, 8, 'trans', '=Output!$B$3', []],
      ['B', 9, 9, 'trans', '=IF(Inputs!$B$28=FALSE,"",INDEX(Commutation!$K:$K,MATCH(Inputs!$B$15,Commutation!$A:$A,0)))', []],
      ['B', 10, 10, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(B8/B9,""))', []],
      ['B', 13, 13, 'trans', '=IF(Inputs!$B$28=FALSE,"",INDEX(CV_RATES!$J:$J,MATCH(Inputs!$B$16,CV_RATES!$C:$C,0)))', []],
      ['B', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(B5*(B13/1000)/B9,""))', []],
      ['B', 17, 17, 'trans', '=IF(Inputs!$B$28=FALSE,"",MIN(B10,B5))', []],
      ['B', 20, 20, 'trans', '=IF(Inputs!$B$28=FALSE,"",IF(B10>B5,"YES - Calculated RPU exceeds Original SA","NO - Within limits"))', []],
    ],
    "colfmt": {},
    "fmt_exc": {
        'B8': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'B9': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'B10': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
    },
    "widths": {'A': 63.54, 'B': 27.27, 'D': 32.0, 'E': 22.0},
  },
  'ETI': {
    "statics": {
        'A1': 'Extended Term Insurance (ETI) - Calculations',
        'D2': 'Status',
        'A3': 'Discount Rate',
        'D3': 'ETI Years (full)',
        'F3': 'Formula updated',
        'A4': 'Face Amount',
        'D4': 'Partial Year (fraction)',
        'A5': 'Gross Cash Value',
        'D5': 'ETI Duration (years)',
        'A6': 'Attained Age at Val Date',
        'D6': 'ETI Expiry Date',
        'A7': 'v (1/(1+i))',
        'A8': 'Policy Year of ETI',
        'B8': 'Age',
        'C8': 'qx (mortality)',
        'D8': 'Cumulative Survival (t-1)px',
        'E8': 'Year t NSP Contribution',
        'F8': 'Cumulative n-year Term NSP',
        'G8': 'Cumulative cost of Term',
        'H8': 'Remaining CV',
        'I8': 'Indicator',
    },
    "arrays": ['E4'],
    "fills": [
      ['A', 9, 9, 'trans', '=IF(Inputs!$B$28=FALSE,"",IF($B$6>MAX(MORTALITY!$A:$A),"",1))', []],
      ['A', 10, 51, 'trans', '=IF(OR(A9="",B9>=MAX(MORTALITY!$A:$A)),"",A9+1)', []],
      ['B', 3, 3, 'trans', '=Inputs!$B$21', []],
      ['B', 4, 4, 'trans', '=IF(Inputs!$B$28=FALSE,"",Inputs!$B$18)', []],
      ['B', 5, 5, 'trans', '=IF(Inputs!$B$28=FALSE,"",APL!$B$9)', []],
      ['B', 6, 6, 'trans', '=IF(Inputs!$B$28=FALSE,"",Inputs!B15)', []],
      ['B', 7, 7, 'trans', '=1/(1+B3)', []],
      ['B', 9, 9, 'trans', '=B6', []],
      ['B', 10, 51, 'trans', '=IF(OR(B9="",B9>=MAX(MORTALITY!$A:$A)),"",B9+1)', []],
      ['C', 9, 51, 'trans', '=IF(A9="","",IFERROR(INDEX(MORTALITY!$B:$B,MATCH(B9,MORTALITY!$A:$A,0)),""))', []],
      ['D', 9, 9, 'trans', '=IF(A9="","",1)', []],
      ['D', 10, 51, 'trans', '=IF(A10="","",IFERROR(D9*(1-C9),""))', []],
      ['E', 2, 2, 'trans', '=IF(Inputs!$B$28=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")', []],
      ['E', 3, 3, 'trans', '=IFERROR(_xlfn.XLOOKUP(1,$I$9:$I$51,$A$9:$A$51),MAX(A9:A51))', []],
      ['E', 4, 4, 'lit', '=IFERROR(IF(H9="","",IF(INDEX(H9:H51,E3+1)<0,INDEX(H9:H51,E3)/(INDEX(H9:H51,E3)-INDEX(H9:H51,E3+1)),0)),"")', []],
      ['E', 5, 5, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(IF(OR(E3="",E4=""),"",E3+E4),""))', []],
      ['E', 6, 6, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(IF(E5="","",Inputs!B12+365*E5),""))', []],
      ['E', 9, 51, 'trans', '=IF(A9="","",IFERROR(POWER($B$7,A9)*D9*C9,""))', []],
      ['F', 9, 51, 'trans', '=IF(A9="","",IFERROR(SUM($E$9:E9),""))', []],
      ['G', 9, 51, 'trans', '=IF(A9="","",F9*$B$4)', []],
      ['H', 9, 51, 'trans', '=IF(A9="","",$B$5-G9)', []],
      ['I', 9, 51, 'trans', '=IF(A9="","",IF(AND(H10<0,H9>0),1,0))', []],
    ],
    "colfmt": {
        'D': '0.00',
        'E': '0.0000',
        'C': '0.000000',
        'G': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)',
        'H': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)',
    },
    "fmt_exc": {
        'D2': 'General',
        'E2': 'General',
        'B3': '0.00%',
        'D3': 'General',
        'E3': 'General',
        'F3': '_(* #,##0_);_(* \\(#,##0\\);_(* \\-??_);_(@_)',
        'B4': '_(* #,##0_);_(* \\(#,##0\\);_(* \\-??_);_(@_)',
        'D4': 'General',
        'E4': '0.00%',
        'B5': '_(* #,##0_);_(* \\(#,##0\\);_(* \\-??_);_(@_)',
        'D5': 'General',
        'E5': '0.00',
        'B6': '_(* #,##0_);_(* \\(#,##0\\);_(* \\-??_);_(@_)',
        'D6': 'General',
        'E6': 'mm-dd-yy',
        'B7': '_(* #,##0.0000000_);_(* \\(#,##0.0000000\\);_(* \\-??_);_(@_)',
        'C8': 'General',
        'D8': 'General',
        'E8': 'General',
        'G8': 'General',
        'H8': 'General',
    },
    "widths": {'A': 22.0, 'B': 12.27, 'C': 14.54, 'D': 18.82, 'E': 15.27, 'F': 16.82, 'G': 12.45, 'H': 11.45},
  },
  'APL': {
    "statics": {
        'A1': 'Automatic Premium Loan (APL)',
        'A3': 'APL Enabled (TRUE/FALSE)',
        'D3': 'Legend',
        'F3': 'Link with Data/Input',
        'A4': 'Policy Loan Interest Rate (p.a.)',
        'B4': 0.08,
        'F4': 'Formulae',
        'A5': 'Payments per Year (mode)',
        'F5': 'Dropdown Selection',
        'A6': 'Valuation Date',
        'A7': 'Annual Premium',
        'A8': 'Modal Premium',
        'E8': 'Status',
        'A9': 'CV at Valuation Date',
        'E9': 'Policy End Date',
        'A10': 'Loanable % of CV',
        'B10': 0.8,
        'E10': 'APL End Date',
        'A13': 'Period #',
        'B13': 'Due Date (approx)',
        'C13': 'Beginning Balance',
        'D13': 'Interest',
        'E13': 'APL Drawn',
        'F13': 'Payment Covered by APL',
        'G13': 'Ending Balance',
        'H13': 'Net CV (CV - Loan)',
        'I13': 'Policy Status',
        'J13': 'Indicator',
    },
    "arrays": [],
    "fills": [
      ['A', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",1)', []],
      ['A', 15, 199, 'trans', '=IF(OR(A14="",H14<=0),"",A14+1)', []],
      ['B', 3, 3, 'trans', '=TRUE()', []],
      ['B', 5, 5, 'trans', '=Inputs!$B$25', []],
      ['B', 6, 6, 'trans', '=Inputs!$B$12', []],
      ['B', 7, 7, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(INDEX(Inputs!B:B,MATCH("Annual Base Premium",Inputs!A:A,0)),0))', []],
      ['B', 8, 8, 'trans', '=IF(Inputs!$B$28=FALSE,"",IFERROR(INDEX(Inputs!B:B,MATCH("Modal Premium",Inputs!A:A,0)),IF(B7>0,B7/B5,0)))', []],
      ['B', 9, 9, 'trans', '=Output!$B$3', []],
      ['B', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",B6)', []],
      ['B', 15, 199, 'trans', '=IF(A15="","",EOMONTH(B14,12/$B$5))', []],
      ['C', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",0)', []],
      ['C', 15, 199, 'trans', '=IF(A15="","",G14)', []],
      ['D', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",C14*$B$4/$B$5)', []],
      ['D', 15, 199, 'trans', '=IF(A15="","",C15*$B$4/$B$5)', []],
      ['E', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",IF(AND(B14<=$F$9,$B$8+C14+D14<$B$9*$B$10),$B$8,0))', []],
      ['E', 15, 199, 'trans', '=IF(A15="","",IF(Inputs!$B$28=FALSE,"",IF(AND(B15<=$F$9,$B$8+C15+D15<$B$9*$B$10),$B$8,0)))', []],
      ['F', 8, 8, 'trans', '=IF(Inputs!$B$28=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")', []],
      ['F', 9, 9, 'trans', '=EDATE(Inputs!$B$11,Inputs!$B$8*12)', []],
      ['F', 10, 10, 'trans', '=IFERROR(_xlfn.XLOOKUP(1,J14:J199,B14:B199),"")', []],
      ['F', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",MIN($B$8,E14))', []],
      ['F', 15, 199, 'trans', '=IF(A15="","",MIN($B$8,E15))', []],
      ['G', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",C14+D14+E14)', []],
      ['G', 15, 199, 'trans', '=IF(A15="","",C15+D15+E15)', []],
      ['H', 14, 199, 'trans', '=IF(A14="","",IF(Inputs!$B$28=FALSE,"",$B$9-G14))', []],
      ['I', 14, 14, 'trans', '=IF(Inputs!$B$28=FALSE,"",IF(H14>0,"Active","LAPSED"))', []],
      ['I', 15, 199, 'trans', '=IF(A15="","",IF(H15>0,"Active","LAPSED"))', []],
      ['J', 14, 199, 'trans', '=IF(A14="","",IF(AND(H15<0,H14>0),1,0))', []],
    ],
    "colfmt": {
        'B': 'mm-dd-yy',
        'D': '0',
        'E': '0',
        'C': '0',
        'G': '0',
        'H': '0',
    },
    "fmt_exc": {
        'B3': 'General',
        'D3': 'General',
        'B4': '0%',
        'B5': 'General',
        'B7': '0',
        'B8': '0',
        'E8': 'General',
        'B9': '0',
        'E9': 'General',
        'F9': 'mm-dd-yy',
        'B10': '0%',
        'E10': 'General',
        'F10': 'mm-dd-yy',
        'B13': 'General',
        'C13': 'General',
        'D13': 'General',
        'E13': 'General',
        'G13': 'General',
        'H13': 'General',
    },
    "widths": {'A': 26.54, 'B': 16.45, 'C': 15.82, 'D': 7.45, 'E': 13.82, 'F': 21.82, 'G': 13.45, 'H': 16.18, 'I': 11.45, 'L': 9.45},
  },
  'Loan': {
    "statics": {
        'A1': 'LOAN MODULE - Multiple Loans & Repayments',
        'E1': 'NOTES:',
        'E2': '1. ADVANCE METHOD: Interest is charged at the beginning of each period (prepaid interest).',
        'A3': 'Input Parameters',
        'E3': '2. ARREAR METHOD: Interest is charged at the end of each period (accrued interest).',
        'A4': 'Loan Interest Rate',
        'E4': '3. Both methods use the formula: Interest = Principal × ((1 + Rate)^(Days/365) - 1)',
        'A5': 'Guaranteed Interest Rate',
        'E5': '4. Outstanding balance shown is the amount owed after each transaction/event.',
        'A6': 'Issue Age',
        'E6': '5. Loan transactions: Positive amounts = new loans, Negative amounts = repayments.',
        'A7': 'Issue Date',
        'E7': '6. Enter transactions in chronological order (sorted by date).',
        'A8': 'Valuation Date',
        'E8': '7. VALUATION row at row 114 calculates the final outstanding from last transaction to Valuation Date.',
        'A9': 'Face Amount (SA)',
        'A10': 'Premium Paying Term (PPT)',
        'E10': 'Advance Loan Outstanding',
        'E11': 'Arrear Loan Outstanding',
        'A13': 'LOAN TRANSACTIONS (Enter up to 100 loans; use negative amounts for repayments)',
        'E13': 'ADVANCE METHOD (Interest charged at beginning of period)',
        'N13': 'ARREAR METHOD (Interest charged at end of period)',
        'A14': 'Loan #',
        'B14': 'Loan Date',
        'C14': 'Loan Amount',
        'E14': 'Event #',
        'F14': 'Date',
        'G14': 'Type',
        'H14': 'Amount',
        'I14': 'Days from Prior',
        'J14': 'Interest Accrued',
        'K14': 'Outstanding After',
        'L14': 'Is Anniversary?',
        'N14': 'Event #',
        'O14': 'Date',
        'P14': 'Type',
        'Q14': 'Amount',
        'R14': 'Days from Prior',
        'S14': 'Interest Accrued',
        'T14': 'Outstanding After',
        'U14': 'Is Anniversary?',
        'A15': 1,
        'A16': 2,
        'B16': ['__dt__', '2023-03-24T00:00:00'],
        'C16': -28.53,
        'A17': 3,
        'B17': ['__dt__', '2024-03-24T00:00:00'],
        'C17': -36.51,
        'A18': 4,
        'B18': ['__dt__', '2025-03-24T00:00:00'],
        'C18': -56.97,
        'A19': 5,
        'A20': 6,
        'A21': 7,
        'A22': 8,
        'A23': 9,
        'A24': 10,
        'A25': 11,
        'A26': 12,
        'A27': 13,
        'A28': 14,
        'A29': 15,
        'A30': 16,
        'A31': 17,
        'A32': 18,
        'A33': 19,
        'A34': 20,
        'A35': 21,
        'A36': 22,
        'A37': 23,
        'A38': 24,
        'A39': 25,
        'A40': 26,
        'A41': 27,
        'A42': 28,
        'A43': 29,
        'A44': 30,
        'A45': 31,
        'A46': 32,
        'A47': 33,
        'A48': 34,
        'A49': 35,
        'A50': 36,
        'A51': 37,
        'A52': 38,
        'A53': 39,
        'A54': 40,
        'A55': 41,
        'A56': 42,
        'A57': 43,
        'A58': 44,
        'A59': 45,
        'A60': 46,
        'A61': 47,
        'A62': 48,
        'A63': 49,
        'A64': 50,
        'A65': 51,
        'A66': 52,
        'A67': 53,
        'A68': 54,
        'A69': 55,
        'A70': 56,
        'A71': 57,
        'A72': 58,
        'A73': 59,
        'A74': 60,
        'A75': 61,
        'A76': 62,
        'A77': 63,
        'A78': 64,
        'A79': 65,
        'A80': 66,
        'A81': 67,
        'A82': 68,
        'A83': 69,
        'A84': 70,
        'A85': 71,
        'A86': 72,
        'A87': 73,
        'A88': 74,
        'A89': 75,
        'A90': 76,
        'A91': 77,
        'A92': 78,
        'A93': 79,
        'A94': 80,
        'A95': 81,
        'A96': 82,
        'A97': 83,
        'A98': 84,
        'A99': 85,
        'A100': 86,
        'A101': 87,
        'A102': 88,
        'A103': 89,
        'A104': 90,
        'A105': 91,
        'A106': 92,
        'A107': 93,
        'A108': 94,
        'A109': 95,
        'A110': 96,
        'A111': 97,
        'A112': 98,
        'A113': 99,
        'A114': 100,
    },
    "arrays": ['B15', 'C15'],
    "fills": [
      ['B', 4, 4, 'trans', '=Inputs!$B$39', []],
      ['B', 5, 5, 'trans', '=Inputs!$B$21', []],
      ['B', 6, 6, 'trans', '=Inputs!$B$7', []],
      ['B', 7, 7, 'trans', '=Inputs!$B$11', []],
      ['B', 8, 8, 'trans', '=Inputs!$B$12', []],
      ['B', 9, 9, 'trans', '=Inputs!$B$18', []],
      ['B', 10, 10, 'trans', '=Inputs!$B$8', []],
      ['B', 15, 15, 'lit', '=_xlfn.LET(\n    _xlpm.loan_dates, Inputs!A43:A142,\n    _xlpm.loan_amts, Inputs!B43:B142,\n    _xlpm.div_dates, Inputs!A32:A36,\n    _xlpm.div_amts, Inputs!B32:B36,\n    _xlpm.div_option, Inputs!B20,\n    _xlpm.div_sign, IF(_xlpm.div_option="Reduce Loan", -1, 0),\n\n    _xlpm.loan_data, _xlfn.HSTACK(_xlpm.loan_dates, _xlpm.loan_amts),\n    _xlpm.div_data, _xlfn.HSTACK(_xlpm.div_dates, _xlpm.div_amts * _xlpm.div_sign),\n\n    _xlpm.all_data, _xlfn.VSTACK(_xlpm.loan_data, _xlpm.div_data),\n    _xlpm.dates_col, INDEX(_xlpm.all_data,,1),\n    _xlpm.amts_col, INDEX(_xlpm.all_data,,2),\n\n    _xlpm.valid, (_xlpm.dates_col<>"") * (_xlpm.amts_col<>"") * (_xlpm.amts_col<>0),\n    _xlpm.filtered, _xlfn._xlws.FILTER(_xlpm.all_data, _xlpm.valid, ""),\n    _xlpm.sorted, _xlfn._xlws.SORT(_xlpm.filtered, 1, 1),\n\n    INDEX(_xlpm.sorted,,1)\n)', []],
      ['C', 15, 15, 'lit', '=_xlfn.LET(\n    _xlpm.loan_dates, Inputs!A43:A142,\n    _xlpm.loan_amts, Inputs!B43:B142,\n    _xlpm.div_dates, Inputs!A32:A36,\n    _xlpm.div_amts, Inputs!B32:B36,\n    _xlpm.div_option, Inputs!B20,\n    _xlpm.div_sign, IF(_xlpm.div_option="Reduce Loan", -1, 0),\n\n    _xlpm.loan_data, _xlfn.HSTACK(_xlpm.loan_dates, _xlpm.loan_amts),\n    _xlpm.div_data, _xlfn.HSTACK(_xlpm.div_dates, _xlpm.div_amts * _xlpm.div_sign),\n\n    _xlpm.all_data, _xlfn.VSTACK(_xlpm.loan_data, _xlpm.div_data),\n    _xlpm.dates_col, INDEX(_xlpm.all_data,,1),\n    _xlpm.amts_col, INDEX(_xlpm.all_data,,2),\n\n    _xlpm.valid, (_xlpm.dates_col<>"") * (_xlpm.amts_col<>"") * (_xlpm.amts_col<>0),\n    _xlpm.filtered, _xlfn._xlws.FILTER(_xlpm.all_data, _xlpm.valid, ""),\n    _xlpm.sorted, _xlfn._xlws.SORT(_xlpm.filtered, 1, 1),\n\n    INDEX(_xlpm.sorted,,2)\n)', []],
      ['E', 15, 15, 'trans', '=IF($B15="","",ROWS($E$15:E15))', []],
      ['E', 16, 114, 'trans', '=IF($B16<>"",ROWS($E$15:E16),IF(AND($B16="",$B15<>"",E15<>""),$E15+1,""))', []],
      ['F', 15, 15, 'trans', '=IF($B15="","",$B15)', []],
      ['F', 16, 114, 'trans', '=IF($B16<>"", $B16, IF(AND($B16="",$B15<>"",F15<>""), $B$8, ""))', []],
      ['G', 10, 10, 'trans', '=INDEX(K15:K114,MATCH("VALUATION",G15:G114,0))', []],
      ['G', 11, 11, 'trans', '=INDEX(T15:T114,MATCH("VALUATION",P15:P114,0))', []],
      ['G', 15, 15, 'trans', '=IF($B15="","","LOAN")', []],
      ['G', 16, 114, 'trans', '=IF($B16<>"","LOAN",IF(AND($B16="",$B15<>"",G15<>""),"VALUATION",""))', []],
      ['H', 15, 15, 'trans', '=IF($B15="","",$C15)', []],
      ['H', 16, 114, 'trans', '=IF($B16<>"",$C16,IF(AND($B16="",$B15<>"",H15<>""),0,""))', []],
      ['I', 15, 15, 'trans', '=IF($B15="","",0)', []],
      ['I', 16, 114, 'trans', '=IF($B16<>"",F16-F15,IF(AND($B16="",$B15<>"",I15<>""),F16-F15,""))', []],
      ['J', 15, 15, 'trans', '=IF($B15="","",0)', []],
      ['J', 16, 114, 'trans', '=IF($B16<>"",K16-K15-H16,IF(AND($B16="",$B15<>"",J15<>""),K16-K15,""))', []],
      ['K', 15, 15, 'trans', '=IF($B15="","",H15)', []],
      ['K', 16, 114, 'trans', '=IF($B16<>"",K15*(1+$B$4)^(I16/365)+IF(H16>0,H16*(1+$B$4)^(I16/365),H16),IF(AND($B16="",$B15<>"",K15<>""),K15*(1+$B$4)^(I16/365),""))', []],
      ['L', 15, 15, 'trans', '=IF($B15="","",IF(AND(MONTH(F15)=MONTH($B$7),DAY(F15)=DAY($B$7)),"Yes",""))', []],
      ['L', 16, 114, 'trans', '=IF(F16="","",IF(AND(MONTH(F16)=MONTH($B$7),DAY(F16)=DAY($B$7)),"Yes",""))', []],
      ['N', 15, 15, 'trans', '=IF($B15="","",ROWS($N$15:N15))', []],
      ['N', 16, 114, 'trans', '=IF($B16<>"",ROWS($N$15:N16),IF(AND($B16="",$B15<>"",N15<>""),$N15+1,""))', []],
      ['O', 15, 15, 'trans', '=IF($B15="","",$B15)', []],
      ['O', 16, 114, 'trans', '=IF($B16<>"", $B16, IF(AND($B16="",$B15<>"",O15<>""), $B$8, ""))', []],
      ['P', 15, 15, 'trans', '=IF($B15="","","LOAN")', []],
      ['P', 16, 114, 'trans', '=IF($B16<>"","LOAN",IF(AND($B16="",$B15<>"",P15<>""),"VALUATION",""))', []],
      ['Q', 15, 15, 'trans', '=IF($B15="","",$C15)', []],
      ['Q', 16, 114, 'trans', '=IF($B16<>"",$C16,IF(AND($B16="",$B15<>"",Q15<>""),0,""))', []],
      ['R', 15, 15, 'trans', '=IF($B15="","",0)', []],
      ['R', 16, 114, 'trans', '=IF($B16<>"",O16-O15,IF(AND($B16="",$B15<>"",R15<>""),O16-O15,""))', []],
      ['S', 15, 15, 'trans', '=IF($B15="","",0)', []],
      ['S', 16, 114, 'trans', '=IF($B16<>"",T15*((1+$B$4)^(MAX(R16-1,0)/365)-1),IF(AND($B16="",$B15<>"",S15<>""),T15*((1+$B$4)^(MAX(R16-1,0)/365)-1),""))', []],
      ['T', 15, 15, 'trans', '=IF($B15="","",Q15)', []],
      ['T', 16, 114, 'trans', '=IF($B16<>"",T15+S16+Q16,IF(AND($B16="",$B15<>"",T15<>""),T15+S16,""))', []],
      ['U', 15, 15, 'trans', '=IF($B15="","",IF(AND(MONTH(O15)=MONTH($B$7),DAY(O15)=DAY($B$7)),"Yes",""))', []],
      ['U', 16, 114, 'trans', '=IF(O16="","",IF(AND(MONTH(O16)=MONTH($B$7),DAY(O16)=DAY($B$7)),"Yes",""))', []],
      ['V', 22, 103, 'trans', '=IF(B33="","",B33+0.1)', []],
      ['V', 104, 153, 'trans', '=IF(S104="","",S104+0.15)', []],
      ['V', 154, 154, 'trans', '=$B$8+0.3', []],
      ['W', 3, 10, 'trans', '=IF(X3="","",ROW()-2)', []],
      ['W', 12, 253, 'trans', '=IF(X12="","",ROW()-2)', []],
    ],
    "colfmt": {
        'A': '0',
        'W': '0',
        'C': '#,##0.00',
        'F': 'mm/dd/yyyy',
        'J': '#,##0.00',
        'K': '#,##0.00',
        'O': 'mm/dd/yyyy',
        'S': '#,##0.00',
        'T': '#,##0.00',
        'U': '#,##0.00',
        'V': '0.0',
    },
    "fmt_exc": {
        'A1': 'General',
        'A3': 'General',
        'A4': 'General',
        'B4': '0.00%',
        'A5': 'General',
        'B5': '0.00%',
        'A6': 'General',
        'A7': 'General',
        'B7': 'mm/dd/yyyy',
        'A8': 'General',
        'B8': 'mm/dd/yyyy',
        'A9': 'General',
        'A10': 'General',
        'G10': '#,##0.00',
        'G11': '#,##0.00',
        'A13': 'General',
        'A14': 'General',
        'C14': 'General',
        'F14': 'General',
        'J14': 'General',
        'K14': 'General',
        'O14': 'General',
        'S14': 'mm/dd/yyyy',
        'T14': 'General',
        'B15': 'mm-dd-yy',
        'H15': '#,##0.00',
        'Q15': '#,##0.00',
        'B16': 'mm-dd-yy',
        'H16': '#,##0.00',
        'Q16': '#,##0.00',
        'B17': 'mm-dd-yy',
        'B18': 'mm-dd-yy',
        'V104': '0.00',
        'V105': '0.00',
        'V106': '0.00',
        'V107': '0.00',
        'V108': '0.00',
        'V109': '0.00',
        'V110': '0.00',
        'V111': '0.00',
        'V112': '0.00',
        'V113': '0.00',
        'V114': '0.00',
        'V115': '0.00',
        'V116': '0.00',
        'V117': '0.00',
        'V118': '0.00',
        'V119': '0.00',
        'V120': '0.00',
        'V121': '0.00',
        'V122': '0.00',
        'V123': '0.00',
        'V124': '0.00',
        'V125': '0.00',
        'V126': '0.00',
        'V127': '0.00',
        'V128': '0.00',
        'V129': '0.00',
        'V130': '0.00',
        'V131': '0.00',
        'V132': '0.00',
        'V133': '0.00',
        'V134': '0.00',
        'V135': '0.00',
        'V136': '0.00',
        'V137': '0.00',
        'V138': '0.00',
        'V139': '0.00',
        'V140': '0.00',
        'V141': '0.00',
        'V142': '0.00',
        'V143': '0.00',
        'V144': '0.00',
        'V145': '0.00',
        'V146': '0.00',
        'V147': '0.00',
        'V148': '0.00',
        'V149': '0.00',
        'V150': '0.00',
        'V151': '0.00',
        'V152': '0.00',
        'V153': '0.00',
        'V154': '0.00',
    },
    "widths": {'A': 24.27, 'B': 24.45, 'C': 22.0, 'D': 10.09, 'E': 9.09, 'F': 15.45, 'G': 12.73, 'H': 15.45, 'I': 17.27, 'J': 18.18, 'K': 20.91, 'L': 16.36, 'M': 6.91, 'N': 9.09, 'O': 15.45, 'P': 12.73, 'Q': 15.45, 'R': 17.27, 'S': 18.18, 'T': 20.91, 'U': 16.36, 'V': 9.63, 'W': 14.36, 'X': 10.54, 'Z': 9.0, 'AA': 10.54, 'AB': 8.09, 'AC': 9.18, 'AD': 10.54, 'AE': 9.09},
  },
}


if __name__ == "__main__":
    main()
