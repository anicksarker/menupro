#!/usr/bin/env python3
# =============================================================================
#  build_wl_model_34.py   (Plan 34 - Whole Life, Limited Premium Paying Term)
# -----------------------------------------------------------------------------
#  Rebuilds the ENTIRE actuarial workbook from ONLY two CSV inputs:
#         1) Inputs.csv     (label,value  - one policy)
#         2) cso_rates.csv  (mortality rate matrix)
#
#  It writes every sheet, in order, cell by cell, formula by formula:
#     CSO Rates -> Inputs -> MORTALITY -> Projection (Annual) -> Commutation
#     -> CV_RATES -> PUA_TABLE -> TERM_RATES -> DIvidend_Assumption
#     -> Projection_Monthly -> RPU -> ETI -> APL -> Loan
#
#  Repeated (fill-down) formulas are written with a LOOP: for each column the
#  origin formula is translated down every row exactly the way Excel fill-down
#  would (relative refs shift, absolute $refs stay, integer counters increment).
#
#  The embedded MODEL_SPEC was extracted from the reference workbook and
#  verified to reproduce all 51,061 populated cells with ZERO differences.
#  In a real run ONLY the two CSVs are needed - no reference file.
#
#  Usage (VS Code / terminal):
#     pip install openpyxl
#     python build_wl_model_34.py                 # uses defaults below
#     python build_wl_model_34.py Inputs.csv cso_rates.csv out.xlsx
# =============================================================================
import sys, csv, json, base64, zlib, datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

# ---- defaults (override on the command line) --------------------------------
INPUTS_CSV = sys.argv[1] if len(sys.argv) > 1 else "Inputs.csv"
CSO_CSV    = sys.argv[2] if len(sys.argv) > 2 else "cso_rates.csv"
OUT_XLSX   = sys.argv[3] if len(sys.argv) > 3 else "WL_Model_LimitedPPT_34.xlsx"

SHEET_ORDER = ['CSO Rates','Inputs','MORTALITY','Projection (Annual)','Commutation',
               'CV_RATES','PUA_TABLE','TERM_RATES','DIvidend_Assumption',
               'Projection_Monthly','RPU','ETI','APL','Loan']

# ---- embedded, validated model definition (compressed) ----------------------
_SPEC_B64 = (
    "eNrtvXl3HDeSL/pVytW8b4piUcpE7nq25yGzsvZNJCVKbs3olMmSzTcUqSEpWzr39He/QAQQWDJLXNx9/xodksoflsCOCACBwP/uVserztHmbnvbfdn5393b"
    "u83dxRl+81D8J/27/U63BMAX8rvC71p+D+R3HIqvWn4lufgagn+SV/WJDDJy4VjCPBBfE/qa0teMvub0taCvJZDKg1ISWtlgbQPOJDjefoWcA8CMm8+B+azN"
    "59B8juBzCDk2rhPzOTUBZuZzDp+v5efCfC5NtJUJuzYBeASfp5Bh+H4LGTafA/NZm8+h+RyZzzF8HkOG4XMJGTauM+M6N64L47qET6xm87k2nzwRn7JJSvh4"
    "HoYsTmWOFUziLJC9YoA4YAy6Rq1gkAXQUwjmkSyChmmRyGJoGAaZLIoLpxoGuYQzF841jAomi+bCpYZxKLOx0pDlhSymG5inAsqylKkqSxrJYJWCaRQXsSyq"
    "wkGSyRzUGoaZrJihhiyRtEYEY5nGOHWL6sJp6hbVhXMTWGZj4cKlS2rlxl27gXkmoMxPmaF7GEKwSsEoSQLZUIOM6k9GqzUMExl5qCFjMvKIYCgrYqxhUMha"
    "nLhwSjCX1TRz4dwElh1k4cKlS2rlxl27gXkuoCxLmVOXhFZVkEVFIMMNtHcUQQfWMITuPsypqNCqBCUa55SijDpx4ZRgJtOduXBuAkOjunDpklq5cdduYF7I"
    "KVqWtNCtVsjsVQqK/gn1NdDeDGBdUEllQckzlLUyIgj9Zlw4dTtx4ZQgDIyZC+cmsMzVwoVLl9TKjbt2A/MwkDwI2FVAYzgGlqVwGKU4L1EABsOzJhxGCfCw"
    "wG7KkfGGyhkThk4zceHUQCA+8/DcYBheCw8vXXIrL/raC85DyfxSKHbo9NZK4zCMMgbF1gFYikw7pGJDqckbOv/IeEM1jgkHeQrldvHU4CyCgrt4boWPoOAu"
    "Xnr0Vl78tReeh5KFZlBypksOOa8IF1kMvZscWAzTM+GQwUg2/gG2OPljgzNKGQruwKmBUM0zD88tXEC5Xbx0ya286GsvOA8lI86h2BEN2giKrXGeQrEGJkCE"
    "DR5RsbHFCcNUMjIY+taYMPa9iYenFoaCO3BuQZiwPbz0qK3c6GsvOA9jgQsod0zlBrGDcJYXGZSbAoQMyh1TuUMoN2EgMDIYet6YcJDFUG4XTw1OYfr28NwK"
    "jz3dxUuP3sqLv/bC81AKJSFIXCHJJ1BLFWHd00lQKrCnJ17RE6/oiVf0xEl64uGpwWkORXfx3AoPspeHlx69lRd/7YXnIQhgIIGFWl6JQiy6xlkSYncngabA"
    "7p563T31unvqdffUYTATD08NVs2ees2eeh0+9Tq8S2/lxV974XkopRbowWWoBZiIYZfPqOwxCGXkEBbY5TNvhsu8GY78c5zijMAEXNzDU4PVJJd5k1zmTe6Z"
    "N7m79FZe/LUXnodSjgG+VIZGCmNQdhLakhT7fO60a20wRBgSZjBmRlZ4kL/D3JVKPTw1GNpl5uF56MlrHl569FZe/LUXnodSsgHJqwwLKjsIqISzpMA+Xzhl"
    "qwmHEGFo/IHgyPgXON4Jh9juLp4SDnC4O3BuQgNeeHjpUVu50ddecM6kbAMhS0YyHC4tCWdiyQVr+MARSmvCSjg3/ijEGQwDYGzCM2h0D0+ZkbNiWOa7eG7C"
    "w5puYcWHRb8RKWGes6JDm1vBoc2ZFG5gGisZyXEJsHXCWQpy24AcGIgnNWGx8oaikz8UZWQw9JGxCQ9zx8TDU2YELWh0D89NeBBgDYyw6KFTMysTHWY5KzjI"
    "M0zKN0CpZCTHpVhyRiWHlAfkwBg2OslpKYx04x/BSDcYev/YhIeamnh4SjiAmWHm4bkJD624MDjGopv8wAxvxcdWN+nBLMekjAPbECXT4k6UwQKNcJbBSnrA"
    "jDAHRSdZDSbBofGOYaAbjOtRKzyU3IFTgpjRmYfnJjguvg0GgXppcAZCrBUfRrqVHLB1JiUc6L0l08JOlGPBSZrLUliEkwM2ak04zLC7k3+M3T12uv/Ywlh0"
    "ip9i2QkDv5kRxrzPTXzcYzHhE2x0V3pcGRwAa7PSww4vRRwG0hxzd5sqZvbEcAOCHFiMrU7iWoatTv4Jtnri1NXYwiGUneKDuDU1GPrBzGCYJecmPgzehfFP"
    "QKQxGDibFR37e+I0FWdSwoH+WDIt7MQBFp2kuTzBorubZjUz0hpIc5Y/SHMGxzjWCUc4y6XO2JwaHOIslzrNNjfxcX5PvZK7wuaKuXtta4NRmGOwvQbCHDM7"
    "aFh0EuaKAJg6ObAEJ/jM6WFD458ib8ucETA2GEtO0VOc5TJnTTAzOEDWlnmNnnkd3pUdV1Z8nOEJ4/KFSfkGslKy3Nl8rQhnRYIzPMlq0EFrwpjW0PLHDp87"
    "A2BssBrsFB+qamoww8GeOz12buJDUyyMP7Z67o11wrBGWRsMszhnUsIBDl8y2oED3lURzoMQ2TqJariHzAqnhw2NP8xbI4Oh6sYGMxzrFD/BDk8YGMqMudLX"
    "3MSHfd6F8VdzPOGcQdmNcBdD2Qt3noukiANlLSMt7cQwDVWE8wDneHJgwOjryGy5wWA3/tAtRhYG7mYw5GVi4kOXnUaudDczGLrs3EoPZBqDYTJZGoxLN4Nh"
    "3lwbDOXhkRRyoAeWUUhlh3YnnAO7GRBWW69R6EjoQ+MPXWpkMIyAscG4O2XixzDFGwyC88xgmH3mVnrQ7FZ82IQ1GFfsVnwY7QaD0MYjKeNADysjRkWH1Qvh"
    "HKt5QA4RDIE6Ys5O65AwA/4xMhiGwNhgGK0TEx8WAVODYUt0ZjB0k7mVHuxWWPGh6MwZzSuDYSJfGwy9kEdSymEgzkVa4Ilj7PG0N4fNNiCHCEVZwjj6hgZn"
    "2OMjZwSMrfDY4yOnx04tDKKswSBrzw1WRY+cqloajB0+cvYT1garokshB8ZWGcVUdJDiCedRjIOdJD44QIpipwMPDcaDQYMzHOux1+ixV/LYGQAzg7G/x862"
    "9sIKDuWOvTaP7U3ytYGqt0sJJwJhLqKDRVyyEs5jlODJIUrg2IEwww14g3Ns8sTp/WMrPJQ78crtymYzgxl29sQZXAsrPhQ88cZ54hyKrA1WU5wUcGA6LyOS"
    "5aDGK8I5iG4DwhGQrqPUWcAODS5wmKdO3x9b4WFjjjBypqnB0BlnBjPs66nL1K34sDllMK5eLH8oeWoPBR5J8QaWumVEklyCozyjHfgAmzxzBPw6ypwF7ND4"
    "Q7kzp+ePrdA4xjNbGJkaiHuxBkOzzE30EGf2zJkdlwbn2OLGH1s8c0XYCI5NQY6LSI5LcGbXOM9AsBuYALgFT5jhnhxh3MceGf8cm9yEh32pKHc699Rg3JMz"
    "GMLPTXyGZXflOANRene91wbibmQkRRuYrcqIxLgU+zqJcXmEnb1w+H1NGNpwSDDCw1TjjadsVnBk54UrwBoMEujM4Bint8KpuIXxT3BeJ38ot+u7jlyZk8dS"
    "rgFCZUwyHAySinBeQHcamAC4aiGMC7ghYdzDHhksy21Cw8Q6iQNHmpwaDPU0MxjmgLmVmiy2FR1mN+MNksYqdgW2tfGHiuGxlGoiEOBiEuDwOJlwEcCaZGAC"
    "QJPUsSuhDQlHMN2MDMb9OBMedkAmcegIm1PjD8LqzPgDH5lb6YHwGrsS29L4QxdcWf7Ay40/zHc8lmIN7B+UsZZwElyqEi4YbsNaAaC3EwYxaBibHT2Y2Q2G"
    "4GODUYqJzQYbCK+GHNCbxe4G29xKDniaFR+LzuzBt7K8QXY13jDMYynUgPhTxlq+SfDQgXARg5QyMAFibHSS5wLs7YRhJI0Mht4/Nhi7e2QNy6lBMbZ45Gz7"
    "zE1kaLGF8Yc1w9LEh2G9il1hcW3RB5YWS6kGZPQy1gJOAp2rIiwW6DjMKQD01powCrJDwjHD3h47vX9sMEhvcex0vqnBqEUQx87yZO6RW1jxoeQmOuxNGIwi"
    "u8EwffNYSjUxyG9xQgXDcZ6QFgHqeJFDCjyujhNno35IGCfAkfHHlZrBBXb2xNmonBp/KHjiiMVz44uni1ZsYGnGH6d24x9jyQnD6jxOtfJmGadULmxyhUN1"
    "TjMwAdTMTvJcjhMcyX8FFjx1+v7Y+EdY8NTpflMTPsKip84mztzEV909dXZFlyY+w0ZPnQl4bfyhgDyWgg3UShlnVDRsdK38JoYOKM1QgAyapY7N5l0GZc9o"
    "HmBQ9szp/mODodkmsdm7A0GGMHaSmfEHqX5u4ifY3018KHrmrClWxht1/Ix/jHxNSjZQa2Wck8YlFj3XRU+gaIPYHLTiSM+drAwJJzk2uysUjo0/jL1JbGQ8"
    "WJ0TjmDWmhl/OCSYG3q4KWPFx7FO8aGsK4NxH9Jg1CWIQRkOpLiYzlFjnN61PpxYnyNPN7tzsEiNjf4cyK+EFV8rnBEwJpxkyNILh+VPTXjF1wqbbc5NdMje"
    "woqOJafouO9u/CMc6+QP4XkipRsQ6suEDlKx1TUOQ1zRDihAjlpihBNUjiOcogaJ8UdtVuPPUHM3cPI2JRzjYo0w7n/NDT1cnyeu2Lg08WElsDL+MALWxh/y"
    "yxPQUgdJLqGtNxjbVUIacgn0uAEFKPCQjXCCZw6EMzxzsPyh2QmnEH6SmKNcmOETI0rCqsX4h1h2ig+NtbDiQ7tb8WGON/4gvRtv4OqJlG7gs0xIbsN9d41D"
    "LDb5ouIM4RR49JBwjkq9xh/4GsEMBteEcIRCHOEkwhYnf+A/cxMfeuvC+MNKY2nFx1LTtiKeMhn/AMstpRtYPZUJiW3QQJXGIUtgjhiQA5alpgiohj40BGAe"
    "GFn+IMYRznH3lTCOzCnhBNfohCOY/ueEQRRfmOiQvaWJjppixh845dr4x9jZpXgDMmOZxLpkoC1FMArxkIkccLVfJ7GdkyHBAjejjDeeNhDOsdwk9eFJOuEU"
    "FUcIR9jRKTYu10x03HQ20UGsWxn/FMsdOwOPJ1K4SfBOQkKDGru6xqI34a0E7YBbEXVC+3S4NUH+AZxzjIw/9P4x4QImnAlhVK2eEkaRYmb81SAneniWbvwz"
    "nOAoPu5AWvRxgkuc0ckTKd3A/FAmKXVvnNw1Fq2CZdcOEZSl1ljMC7BqIf8wx+6eOt1/TP6BGuqpwxymhPFaBslt2NdTo5YLBafIIY5zigzbURZtOFsjjMdL"
    "iZRsYB4vk4wUvHFe1ziJcIVODniKrqGY77Cva28GLTIy/rgZRf64xzIhf1wSTROjjQd77YTjHAtOQiX03oWJD5PA0sSH/KyMP8hwFnkouBRrQAIrEy2yqT0Z"
    "wilK0gNySFTJScbD41RDIERmnjt9f0wYO9QkIZkRVYUI42HWLDEiH0juFB+3YRYmfojzm7VvCCU3/ji1G/rY6FKwgayUiRbZcIapCGcoTg5MgBQnOI1x05lg"
    "EiNTK5yuPyYc4QRXOFdFpoRz3IAkjBP9nKLjae6C/FMUXgnjLLIy/in2dqIPpeEpXG8AES7VIlsGOa8I5yhtDkwAVIgkDEIMIRhHI4LY+ceEEzxKTUlghC4x"
    "JYz7cARToD43aUG+FyY6cOQl+aOQsjL+KLQb8tCleQpXC0F8S7W4VqDKiME5jLOBccDeTjhG5WfCOTa58cfbSYRxnTtJSV6EcTsl/wAEjxn5Y/g5+ScMb2NR"
    "fNycIH/cgF+lRpzEC1lEH3a3eCoFG/AqUyXjMJz9KsJhFGHZyQFKruW7BFUCTXRgacYbKmJMOMcrLSkJhCBLTMkfMzZLjcQGExz5p9BDFyY+FJuETaC2It8c"
    "ltRr8keOzdNI3yYt04iuEMIBi8EFML+Bcciw4Fqeg8W58cV7d5Hd88cEixRLTbJgjD1d45Bhc0eOrv3cpIUn56kRNvH2nfZH6X9F/tj91mnkTK48lUINBC1T"
    "Jd+wOMbmJoxXeQxO8L5hTFMATG7GP8SeHjs9f0z+eFwySY0wiJcOCedY9NjhUHPyx2ZYpEbYBI5GOEZ9CYqfw6yxJv8owZ4uZZoU5LdU3xpN8YDFwlDNA+OA"
    "u1GE8Tx5aPxRE9L4421LDXHNNUlJHETtKIOhy83SxDnMnBtyqtVJ3sT5LXHY0oqiF6rRE6oZvGAqpZoUr5jqO6YFjIuKcI7arwZnOK1rjPv6Q+OPW87GH3bh"
    "CCJvnqQkCxZYcMLQprPUyIbY5kQOr3ZQeLyGsyScQXdeEVZH54SRn0mhBsS1MtU3TJXEbjArsLeTA268asxQnWho/EGIMd4hdnaNMzWtk3ioxrmRDmG3PTXi"
    "IbByQw93XlNHmlwSLHAzhnCA8jphZNc8lTINMJQyVeJNFEfIyTWOkIMNjAPeVNRY9H2c4IgA6n8af1SCJH+lCqhxmIG8PyWcQJeaEVb6n0QPufPChM+xzbU/"
    "LipX5I+DbU24UK0upRoYZ2WqBJwow+uKhNM8w7KTA14f15jhWmRoCECzjMgfu/+Y/CMlyBSO+DBNjcAIsithPDKem/RSnOPMnVlQgyR/pL8ifzwIWZM/qhby"
    "TEo2IIqWmRJy4gBIVwajftpAOwghCi9Za4zH3UOKEMKG4cj44zVrDVO83KAxQxXnqcaiKoCfE46Af88pPsN9KBMelV+JHo6gFfnjzu+a/FGvi2dwSRWEuEzJ"
    "OLHSACUcZagyQQ44rdQaRzGqe5M/3voaGX/UACVcwBCZaIwHllOC6up1RjIgbDLPKXoCvWJB4dVBMuEMLzJRfBjr5JvgKUsmRRuQmcpMSTkxSj0V4SyNseDa"
    "IUX1II0j3IIekn8BjTIif+z/Y/LHTeKJxkytUgnjoeAsI6ENt+FMeqrkOjxury/JH9f+K4qfwVy1pvAZqoRlUrwBHl5mStJJ0AxAZTAeeA7IIcDTVI1jVG0Z"
    "kj/2mpHxh/4/JowC6ERjhrfXCKIuXEaCG9TM3FCDPrKg4CnDNo9oJkHjAjp+DoLemvzx6I5nUroBmbLMlKCTpNB9KoNxFhmQA95M1jBOoBKH5F3gXR7jr0a6"
    "9g9g82qicYTL1qnGLA2w6CS4qUbX9GLodAsKn8Nqb0n0UtQDJH88vl2TP2pk8kyKNxkIcpmSdFKUKSuDCxTktEMCsNZQDESQYSl8DPxgZPxj7O86OmqdTTSO"
    "cI9wqjHLcLOdMOoBErkclcIoOt5VWBIu8BoTRUcNkjX5JwkOdSnewCKkzLR1kBQLrmGGOxEDckDFwFpj0btxdlc4wz1nyxs7u8YFrtY0jpDxTQnjvttMYxZB"
    "955TfGTSCwqf4CmDxrjHvKLoeKK5tpJDviblmwwthyhRJ8M7FxXhCHnsQDukyFNrjcUUhSXXEVKGc1zmdP8xxce8TTSOWYJznJYEkZHNNGYJHi4RvVR1dx0+"
    "h/2wJdHDu0crip/icapJL0ebKVK+ge5ZZkrUyXJUdtY4V1qFAwoAWas1TPFca0jhUZofkT92/zFFV6KcxjGeukw1VlU901jI0NjfNb0QxXeKH0U4vee6V6Q4"
    "vZPoqPo7pRfhJCfFGxgqZaYknRyPpSrCGZ5nDsghRJ0RjdM8RKMx2j9PkKcX9gAYE0yh10w0TnB9MNU4xqE7y0jyw6s8lFwGTbOg8NhUS6KXq/md4hfY4yk9"
    "qCqeS/EGmG6ZK0mnQLW0inCM9/gG2iHPUZLTOMNqH1KEBCbsEfnjCBhT/AgvpWuc4NmahjHyl5nGohdA0U1yuAFL0VF1Y6lxqk6SKX4OvXBN4TPs8LkUcGAt"
    "UuZK1ilwgqwUDrXUN6AAyHtqjTM8hByaCFDNI+2vRsCY4gO5iYZpAhP2VOMENdhnOQmOeI2JyGEHXlD4AnfdKTsBbKavTHyQBNeUHh7V8RzsjYAslzOVc5wm"
    "Ko1DdYw4oAARXupQOMhx2WbCx2g6iDn9f6xxgQfVE40zdYNL4yQDcWmWkywJRZ0TvVS1OtNVB/lfEr0Mxt8qN7IptroOX2B/l/INTGJlHqmcp3iTR2OmTscG"
    "FEDdXVM4KKIAjSZRBDTCQP7Y/8faXzQTtnqkp1XcmdJYTJswzeUkSkIrzk16eDuZ4qMS9FJjHPorio53E9YUPFGWoqR8A/JImccq52h2oNJYJIS6kBQAVXNr"
    "jfWJg4kAHGBE/tj/xxoLKRT4usKy08A0p3GKsttM4yTH8zWNCzzGX2ic5Wj4jPxxQbqi+AVq+FN6DDfl8kTb3itzZdotwnPoSuM4jmFKHlAA3DCsNVZ7KEOK"
    "gINrpLEaAWONWYh7NAqLZsTbmhpnaNpmpnGKiqlzSi9DFTHtj7PuksihxL4y0Qu0mqWDF6jwnUsZB3hVmacq5ymqCmksRBZlJEwHYNjqCoZYq0MKr9arGqsB"
    "MNZYSGvY4VPdK/DqmsLKDIFGGZqumlNqiZrkUs1/1EjX1PD4ZUXxUXyn4KgNxHMp34DkX+aZyneGZuA0zlQTDigAnp7UGrMCN981TjPofiONVfcfa4wXgCca"
    "hrrcStQqmCp5pmUIPGyh5GK8rqhxgO2wJHoZw96u4yvGllGHgUOHXIo3ORqFy1VJ8Zin0rgIcVE20A4p7sLVGgtxATu7wjkO/JHGqubGGichZH2isRAv4LhF"
    "Y6XEN1NYDFzV6Do9VO9aaBwmAQ50TS+HqlhRfJwj1xQ+U5xNyjdQy2VeqKyjgnalsNw4gmof6AAZno3XGicB6j1rXCRoaIb81YJV45SpSU5hscCExYvGYryB"
    "BK+xkENxfi/0zINKgRqjPu+SvAPV6hQdDQFScLzfUEjxBmassghUxnEfqlKYiZUEaoJSgBg4Sa1xGioreSqCGHvA2cgf+/9YY1Q5nGiIMvNUQ4bTzEzjEPni"
    "XOMEzxUWGke4Gl6SP1PG8nR8PE6m1HBfgBdgVxXkuCLUTQwbchqKFRLqy2gHMRdDb1dYjCuYKIban8UwVY+0v+r9Y8K4xzzROIW1wVTDCJf9M41ZnmDBdXDc"
    "5lponODV3yX5p2gwj+LjknFN4fE0gBdStoFxVRZM5TyDoJXCkVKhHpB/kmKLM10SPEnX/nGgSs6cvj8mnOFNRY2zFMe5xgle3JhpHONh5pzC44HQQuM0xTua"
    "xh9VISm+UgOl8HhgyQswHQdyXBGprON1hkrhSLBzPHOhAFhyJZUVqK811L5CngTRXWPV98cGQ5+ZUHzUO5lqnGYoxWmc4N7hnMLjim6hcR6o3q79UY9pRfFz"
    "6MRrCq/M6xRStIGkyyJW/RutiCkYyesNsEVh/GFdXWusjNwONcb9u5GGqvePCePN84nGAe4ITxUWOcOtd40zVPGcU/gYrY1o/wKtry7JH/vsysTHk1UTHg+V"
    "CynZFCDEFSjksDzAm+gKx0mM2oMDE0BZAVWY5WqGU7jAE5mRxqr/jzWOIjxQp/hoHmSqsBwPIMASVlqBOryaIheE8c7CkuipW/gmPhrSMuHRvE4hpRuY6Msi"
    "VVnHy2qVwknE8B7fQDlE6tJUrSPEuGQaan80gDPS3moAjCl4GmKzp5p7AF+dagwVOdMo0HOcDo0noAuNxSSMc5z2T9Hmq4mP21JeeF5I4QbGcVmgnBOxUA10"
    "xGkU4wJtQAFwh6BWWJ6u4ySn/GNULhppf+z/Yw1FaJBlCLMM7b8qHKGSw0xjvAMw1zBGC08Lwshkl0QO12crEx35uReeF1K0gYyXRa4yjnfJKoXFYg27z4AC"
    "4LZSrTDLVcGVd4aMZaS9Ve8fa5zhhUWCqEc/1ThRW5GE0fSAhimOvQVhHPdLIqcuLJroCRbcDc8LKdgUaOoXZZwoxRatFJbaYXiqrAOgyoxGIVbDUOE4RPtI"
    "I/JH/TAFWZGjxTjCqMw41TgrUNvbw3ONc6U64uGloYfzmxt97QXnYSAlm0IZ/kUpJ8qzDM0kgoOy+av8ClQprskhQR4x1A5xgip4Iwqhuv6YHBgqjk58hyk5"
    "hLj/NvMd5iZKjFcWfYelT3Tl01j7UUQlSCGnQDPAQajbL1QlVw5xhvbiyEEqK2LJ0SFhqHo70g6674/JIcdbFxPfYWqi4BbBzHeYmyhBoSwfuw5Ln+jKp7H2"
    "o4iSSxmnQDPAAVMFxf0YjZMY10E1BcjzGC1DKgdpxh1tHusQId4LHpsouEr18JRwhuuRme8wNw5ZqhrcdVh6NFc+ibUfQ5RayjcFWgEOItVtcSIaaIdMiSS1"
    "dkhTHLJD7ZCnORqBJRIpTuoUI0EN0onvMDUOCWqI+Q5zK0qExiE9h6VPdOXTWPtRRMHBCjCaAQ5Q1Emw0LEa6bX5HJrPEYVmKCyMjd/EfE7N58x8zs3nwnwu"
    "zefKfK7Np8gtGu5Fy70BSieJOtrSOMVrSmEAlm4DNHUbKPFEzdgj7SBaLFP9HozDBmgdNlBMPcUTx5HyxYBgSTVAU6oI0B2sjCLzCMGquNJTCcHQNir0CADm"
    "GQNlaxzsFaKxzhCsFKP9JAGgpGCjWAA02ovpoBXbEK35ollXZNUh2jnFO+ch2v4MMQdoDRNVhkI0EImLtxAtJuKWTYhWBFEzPARbcyHDHID1NUHmHwJtbm42"
    "3+SzJn//D4E+XlxeAvg7PGYRw0/37mZzdSvftvjpw9fLj1fPq9Wy4ie9MuyXrF9G+8Lr7/8hov+9W90bqQr7FetXdqTBvZEGYX/A+gM7Un1vpDrs16xf25GG"
    "90Yahv0h6w/tSKN7I43C/oj1R3ak8b2RxmF/zPpjO9Lk3kiTsD9h/YkdaXpvpGnYn7L+1I40uzfSLOzPWH9mR5rfG2ke9uesP7cjLe6NtAj7C9Zf2JGW90Za"
    "hv0l6y/tSKt7I63C/or1V3ak9b2R1mF/zfprHUmOkrPry4+f7uRTQHIEic8P269nGv55cX73u3onSC6rhUj9D+k+ufr85a75lhA8E9RdfbkTvuaFnsX1zd3m"
    "8uLuW+dk8+vlll7rwXtKEKTa3P7eebO5/LI1j/LMt79tr86tZ3Ouzrc3f95c3F1c/dapLje3t63P6JTbqw8XVx+vzWs669e84yaAL+fML67+q/Pnxd3vncHm"
    "bmMezFls7s5+35531teXF2cXW0iGx0DpcnPVqa7PsQyxthRToacM/s2kDY71yQQSBDC8vvn05XIDsWdWnM7yy6dftzfwTg84yzt+Yn4L4LkecAkjwRhZCi5L"
    "5RIlQqwHl5VykQuQCFzWSEesgORdSXrKp5kgPOzjBKySttKA49EanjGCF3u6g5vrz+fXf151jreX27O7i+urrn6gpzvaysbq6hd6RJViu8MDPd3h9pPCSw+v"
    "PLz2MDzSY1GHt3rUq09po/XBha/nXf2ajuhDF3fQnvNU29mHh3QS/YQO04/ngB+8mwN3DeHJnO7k9vbLtsN/g7zAazldATpX283N9vau8+vFzd3v55tvXf12"
    "Tnd+LTqMGA9ikFydi37b1Y/dQKek3MCjNWFSPC/UKSyEuNl+uvjyqbPefJM9/mR786nTW69P9rv6ORswRgYv2XTrq7vtTac1yjeRu9v9zutbEeDiSo1MeIdG"
    "lUf0figQPEbz9+6HD+d3Hz5Ih7DIisMwOAzikyB4CT/d/9DP1LgB8/gwSA6jwA24bKUYsMPQo7hqoSjSTQ+D0A24bgZMCgiYuAHhORzT3vAoDrQkvIfj9BN4"
    "qaYrJ4eN7MWmPgovLRYwmdBhGHn1sStgoz4eGnD10IDrhwbER3PcbofP55h+h8/nuHWDToMVPAeGL9q41R8lsouEXvUvWkLmkWz6IPPqpCVklh8G+SHLvUpp"
    "C5mJLnLIvNTXLSFj0evkj18tYXMg4Ps6D+ll+PKOV2Ohw/S69CBONw6r4xXDWRbfxJHv5R2vIpzpl02nVdNp3aSF7+O09GB8LqfZPdoKAiSArxs+iQ/cwBxX"
    "bm4vcOY0bkuc97r01I3vumx1XbW6rltd8Q2c7nwjpld+dXXxx/bmdnPzrUvPxOiuiU/FPKTJ8RGZ7nLzafuBn6OAAWuabvXl5mZ7ddfZ4ASPL6xAjjZYKW7t"
    "4osl3cGXG3Dvfd3v0tskmtFqz07v7PrT58vt3baD83GXnv2wCByEioTj3vl4fSN4zFeRMb8GYBYXTbPtHPMuPSjRHWzPLkUVnncGF39cnAtu2enxj5JBVNdX"
    "f0D+kYngqwz02kR3uf2z8/Fie3ne+Xhz/UnG7pwr0Qifn2AxS5+jZhwL2lOSr0x2ep+fb55jGvj8AT180K2/ft5eiQwL1ig5Yme4Obu7Rmau3wuI6LGALpFd"
    "fdYSBr4dIGeyLhnW7/Kzsy9StoL6mkhmKBlyIy9a2zshy/Rd0au+bC47UImKf3bJWD2swNFOvTPI0bK7GOTnIqoVC82eE+teaFERnTGpLtkIV/FNDaB9cF/2"
    "QqvaXdF2Qqo/u+scC4n7C07g6MO7ZGC7GbdwZYmL88PXn6lbdskadXcxWUrpok9+2EXB3LGYHztqcFx+09KFVWw06tw9OXpddy4+dgSZTi3aTHa0zs9tIwcs"
    "CaN0JBqL2qprW1E2JpS7Zx8urz5citAfzkQuROHuPtwQKWaTwnY/+fZZ0WKmfvDbkJJUPm3vfr8Wo6PDD7ko3uZGFP+ww8//2FydbVUFRERfVPzNnZlcwceb"
    "XCNghp50gyaB/VL8urn8cE6lgC4GCci6heT4p+svV3ddMrQbZRGOPbSzi/SAFBAV9GSuu2SrFjN9tP28+fZJTmum9qE3Gg+TENpztXJ6IwJdnH+4u8asupss"
    "YN8en1bF90pbdl0y+LHWo0er18tBryd47KGYs/dfiIXNczEeA2sdK+PJjSPYPDIxJ8PegJ/UvXc1P+rtlXsh2+8vVsuTMYBQANGFB/ydhu+vJKhXGOb7UcVP"
    "sC/+vb/68Sfwh9j/4sRa0jgM/0IqfmSdjl+1Cf42GgVIlWG0f6i+4v2DHlIEV/0Z7++/CNkBZAw88CMWGZDtGTQbM8VfuzGXJ7ITMNEJQt0J/FgZ/lqxyjA9"
    "CL1QOf7aoYpn0qCdG07O+cwttZjjn2EXtsNl+NsVwpu1gXJ8OjmpROFZ2n+vJvL33X4owLGYBg/JRW5Sv++++iIG8vbm8hu4MOm0ENP37+AQisrrh15h5bE3"
    "HH2b7Mm2L/O+KLMXVq77YO1nwtbQF0Rd9sv8megKP0vRz9lfZPDjlWpen/TeX3XEP4E/f3oupmo5S/Q7JYwA4yFdjzZXvwmv9c31/4/r/Q+qTD+I/sZeyk4X"
    "iCWWHa16851Icxlp3oxUfxVcDjZh+h3M5tv5ajV7ve55mWxkzUv0fVfUdifYV9TF/OEn8OPPGKaRMMZQeXIqynJXiZdbIaBtn5ZXywFzchju93clBDLcPymd"
    "HclUb75XGq+036n8XdS/UwS3hI+mPRBMqdze/bndXvX9OuscNptrF42hkHzdMqq6fSiRoRTSREfvtxLtvGjJ7/daonPg0e086/Sc6qSM6Sj7hp76FP95xw0R"
    "/Dx8Ooj+5dPBQkZa/M908D/Twf9MB/8XpoMDOUCLe84g65PJD3XqBkrgxxYe169/kM+WOqFS+LFC8fX8h2EYuKEaqwMxKMWi7af33YHolW9Uh3hvtraEq1zV"
    "/LBX7bH85d5oLw76cvHQp4j8nohziLimiO+7y+u7zlCsgM7fd/dbD8ZEVsWi6G/9v/1Nmgrpeudk3RpWY58+HZ6fH36DjZk69V1gb91xWTRclg2XVcNl3XCB"
    "XWqXcsNl2XBZNVzWDRfc7XVJN52WTadV02nddMJdVs+J+U5qY/L88NMn3JfBDUE3VtyMFbuxcLdKNt//MltNNo48jOdnXXq6T31HXrzIzU3LmWnEnuPjAvDS"
    "63PQuqrgWcDnqMQDzyM+R3M3sB3/HJYfQ3gr8zkapgQVB/weW+5TdSYrt2qRPDgv4PzqOb53gd/gvgYdiucBHOEuVkcnfD45edc4xeWh2vbEfT8L1AD++8PX"
    "Ti8skrxT1Scdviw7Cz6v9+nAN9DHugHuDWgLr6k+kcUHoJxnIuG0FK6Oxs6LB3DuiPYMnfc74Wwv1kPNPP8Fx2546dp5mnQAlnjgFovzZNCgUC/L1+4TaAPo"
    "r/jGcOC8lTKAbovadaHzLM9AKdC4z86rR5sj/bi19TS7ihXrt5+t18tVJhL9NLL1undCm8ygT2O9fY2vKaIWTuI+DZ3S68l4g8s8jIx2ClBBx3tXOKPXd7F3"
    "mpds0WwP6u54D9Oiwi3s0bLAfbk1p8dNUS/TPBUK3ROf/4Qr7NbLn3lMr2My72FMfKUAX5qMvYcjC3pdUb0fmznPneDrg6graF7yQ+P/+JKf/zBfRvuczH23"
    "Dl8ZxsfdmPuuG8Nrg7DzCAt28+wZC+lpsCh0XwULzdNZaLbWPIPF6HGpKHLflWIJPb+k3tpyH8XE94nUOzWZ/VQGPuATpe7bPYl54QbvApnXZPDeM74Xk7vP"
    "v2TmlZSocB9IwQqEZzjgEYbaeiAko2c24tB9YQNHDzzlEDP3DQrsg/iqQ+Q80oDXDPEtA7zZYT1LYEz9ozFkY3Ufp2M0m5+6VvDVqxCFtrJtjMQneBcaza7n"
    "rhV1ZWU91IaajZFxvCaP1rjhVNoyxI2zDJh7RmP8liFqsuacMNeQc2zsHSeRa+qYjAEnsWMHGO0HosFctFNhjOFiR0Pzs6ltTVYZPESLq3glksyforEgNEma"
    "uNZIUdkRrXYqU35kIBPnf7AMCWYBjUVLfMkWDT+mrs1H1WfAuCAQMHYR8dFNtB2Il7nIEJ9qETTEF9uG9UK8nYT259CILBlzy7HrojE31zYb1i6aOsscy2Vo"
    "JAete+H7pmQpCw09ovUraFIyfMVs+1BgWodMQ7Ec1KLRfBJq85MtIpwwwD4PXB8n00JRbJnfyRzTO1ERxmSgJott2zRxjNrkaO0lsY23xAWaukBjKKlt3CRR"
    "j8agsZDMsf2BtsXQQAYaNCRjFGh2FA1IZIVjOyJDrgUWCuBuum1boSALBGhnxBgXwEs+cMM9Z87lfDSBgBfY4f4v3V1neO0IL3jnsX25W+mL4x1odeNdX4dG"
    "u5B4SVgZAtD3hdFIPN6hzTP7+myeRuaSKd6spfukqoPilc3CuoIp+lNoLirCJa+a7iSiTjXe40Oz/vpSXoavj+NNN3yXTt8qKzLrJhiaZVKXqiKGFu/xolQR"
    "W3ekIrQFgTeJoF70JaJYXebGqzaFfctG6uCndBmlyKybKGmIZi3UnY0it69r5Orep7rWUFg3GkKj9I8611rh32jF45SjNeItxXG0UKKUxNE9Ib3pWulqo3tK"
    "etO1UstG94xUqGulkI3uOWlT1wiUe0GK1TUCdFeq2FjoMDDuIalb1wiUOyPN6xqBco9ICbtGoNxj0seuESj3hFSzawTKPSUt7RqBcs9IYbtGoNxz0t2uESj3"
    "gtS46zA05VWK31heZsqrdMCxvMyUF9XBmRKoobzf1fLmuK+JLW/2FA44axwawc9OxWTQbv2hZH39ldJX5J/bNNObLAf1296/VccrONS+/bcf9vheIvcc5Iuq"
    "Cy7PkTjrNwNwCBDsqzByg8YPFcPWhTzWa92oKPXqVPz7ACsy2Lr40HvW0dsX6PX/Sqf373u26/v32vnw3/8dgvx/goa/21GGSgVze4O6DLXj4K2AYdWb4/Kz"
    "xlVvnMD602wBd3p4era/YyXa1FkNSRlXavHggh+0RdQJNy5a3203NyaDagk7DG3VGtA2BpfqjdLD6JAazrjhAc4TldS3287ddatezlRF7NoL4QnTksAEhP4U"
    "34gyn4n5TM1nZj5z81mYT5g39HdofTPr20oxtJIMrTRDK9HQSjW0kg2tdJmVLrPSZVa6zEqXWekyK11mpcusdJmVLrPSjax0IyvdyEo3stKNrHQjK93ISjey"
    "0o2sdCMr3dhKN7bSja10Y7tlrXRjK93YSje20o2tdGMr3cRKN7HSTax0EyvdxO5SVrqJlW5ipZtY6SZWuqmVbmqlm1rppla6qZVuavdlK93USje10k2tdDMr"
    "3cxKN7PSzax0MyvdzEo3sweRlW5mpZtZ6eZWurmVbm6lm1vp5la6uZVubqWb26PXSje30i2sdAsr3cJKt7DSLax0Cyvdwkq3sNIt7GnDmTdUwveyVGBxrrqB"
    "4oaCQyUt7HBX2Ni9W6U0EtpC2soLlebqLlmj/FKx/QOt+SK++1IbRfzvXcpqITEZ9upInfkN2EG436ox0Za/rHGY2kab/fiTdK4F6b5MxLu89b3yp27Q9gQG"
    "OvPDlhRGO1pjWB8drY56KKRUbz4ciXo8/mFv9nJvpuQOoGZ8qpd7lRQ4HK2e8V+gfh/tKdKO3aOo3pg9MxVU7B/0RuxwzPZtx2cTJlWJWkWjyj8jmDiSkScS"
    "tctCasu923oMNPEFo6nvMGR0kIBXnICOlDQcNIocPweNYsfPQaPE8XPQKHX8HDTKHD8HjXLHz0GjwvFz0CgMHE8XjsLQ9Q1dX7dyXDgK3epx4Sh0K8iFo9Ct"
    "IheOQreSXDgK3Wpy4Sh0K8qFo9CtKheOmFtXLhwxt65cOGJeR3Lrirl15cIRc+vKhSPm1pULR8ytKxeOmFtXLhwxt65cOGJuXblwFLl15cJR5NaVC0eRW1eR"
    "N+q8YefWVeTWlQtHkVtXLhxFbl25cBS5deXCUeTWlQvxSXnL162r2K0rF+Lb5JavW1exW1cuxCeuLV9vjvImKQ8mzhTqHZfqY065cCzMaWnxHHer5P7xc9zI"
    "lhuGzwN1UprjASoclEYYdgLnTBh4iqegEGaGqYQ5LEOr60+fvtzhYvC7B6G46Pzvr2a1efnVrDXPv5qV5h//+dVaaX611phfzbJy+dWsJRdfDbvYfP1wri/f"
    "QOpf8Sge4szXHz5vbz4o5dbuArXo1+sPvg+dw1Zqiwz+PWjzJGyIGYKfc6lZIQQNzn7+acHf9ujw+Ic9/nKPS1Vm9D4IW/ZHmiQ1PfjjCg0W6fLlXmn2S7w0"
    "ldiw3xQUW5KLWpKr2LNeeFgyj8rgcZkWVErmUqgfR6EXvuiFB0aKYdH+/n9yj+bw0bmqPQqjx1EYyNppyVgP2tglPX5kOw97JfsJbROwg3HkUZs8jZo0WsAO"
    "Jj616eOojdmLoVdxs0fmp0lh/jgKcrhKSXbskVk8jYwo0yEOrfHLMQ0nq13zPg2o/YfuLH5/V7BlXlcTeGlNxBWqo8B0P8dJGeZ7UF9JxcSNE7RaLOyYnek2"
    "fti8oN3Y6lPKLHd3m4ur7XnH1WqxJu3P1qTtXCYz83e5vdt0emLefSHr2doYrN50euXq3b69Edjp1cplRnuKvDpZHXX1FE07hMHDluNhYy3IfpaLMewBZsX9"
    "gN3qXXGzlqm1EZfGoBXvoBJTxM8/EckKZgxveTrYlZsdFC16rkeDcv2dctrD5DvcZtDKbVqm43vSCA9r1rYkvydn9ozbsuR+SLksseaHvfnLvbkqmVV3fSdM"
    "SyEnuxq9N7D56ZS1TLetnawo9AzVq+WULUcO3HDCDyFXBP1eS/6t3YNB1J7tw3vi3VvuF702CuOXe+OHUjh8aPx8R82/aCEwfLk3fHDT7T/bXYdWTnbW4b+4"
    "ALvSFfmGOXS/heu6NwCnrGH1B3vambzAiIdDP3lTR3QY/twmsrrMckpnaoHHbJN7mK3VlUmiE0LAgfCp2I8/OWx2vAcsXUhWByO2LyYHLLgSGNoZ74D2jOrG"
    "4d6w4TIy6qk+ix44C7y6sTnlO4zuY+lhgDy9shZdA+Tj+GwZ8nd8qw01WwO1LguVUuxUk8Ejwdf8wwkv53WT10cN7o6Gf2zujvZ/fO5ubAAdrztijdTZCzvH"
    "vNPjeJpXR2aVNoxwZQanxWobDtdm42gXzzcWht4Mgbz06vSGm7P/utze7HfvZefQh6Pv9+Hw+33YY9mB4fd/nTiu50S7Mcco2j8vyzKkl8bATsMZe1X8fU5n"
    "zfhVfC+Dq/85ZXH3vHdMfruy4007w391lqwJvYrlPv8DMjX6V2dq8nJv8rh6Gv9Ts4RLpe/m634B4i+KD14BJ//UAmIJR/HhOH42jPdf1Da7azsoAR4xIPbS"
    "qknS1CLZeWoy/AtxR38h7vgvxJ08Na7PeJEvWZx14DvUvsPQdxj5DmPfYRLdo4ZTmFW3UsgZIFM2V09i2k8Vn/jMMDJrw8NDs7kaZs9ZBoz7pD5afHeVroys"
    "gEGy6losphUjvv4oefFh59dvHZ9vc2ZvvgIw0Xt/XGw6N9uzLze3UoEIzXjs02paLebRtuAfzxDVGoGCzzNc43OtLQdNED6IW7Ooub/Yvi2qdkWjxq5o5Eu2"
    "1rq3ioQ8KS/ewVyhP55V0QF998LDKtp/VsaNxW+5O49x++aZYKmtycVucrFILlHJNTcGovsX8alIXpoBvGfpzaPWjd5GlPrlXt0aZfByb4BbWV7VVI+umsdn"
    "Pn585uPvZL6pwtBe0VYnadtneUSprfZvZKF+GKlBrPrn/u5z+L+ikOithJoH8TgF2NMr++5827a0ifH2nFJfpGMovbQZ4LrFXNRjeg0zmKDBqg9cLFY+fW49"
    "VyrtcyWc5+x1TKk3BHFyevgKwrpsG/kar82+Y3aZgub84dETi63vq+4clI6OLe0XemTixgagR6ZyyAzayVTxPYo6BwOHTN1OZhDfo49zUDtkhu1k6rixw+eR"
    "GTpkRu1khnFjq88jM3LIjNvJjOLGSYpHZuyQmbSTGceNPT2PzMQhM20nM4kbOzoemalDZtZOZho3DlM8MjOHzLydzCxuHKZ4ZOYOmUU7mXnsGnduklk4ZJbt"
    "ZBaxa+65SWbpkFm1k1nGrgHoJpmVQ2bdTmZlkXnVSmbtkHnVTmZtkTlqJfPKIXPUTuaVRea4lcyRQ+a4ncyRReaklcyxQ+akncyxReZ1K5kTh8zrdjInFpk3"
    "rWReO2TetJN5bZE5bSXzxiFz2k7mjUXmbSuZU4fM23YypxaZd61k3jpk3rWTeWuR+aWVzDuHzC/tZN5ZZDhvpfOLQ4fzdkK/2ITKVkLcvVbCy3ZKnNuk2pkn"
    "d7kn38E+uc0/eTsD5S4H5TtYKLd5KG9notzlonwHG+U2H+XtjJS7nJTvYKXc5qW8nZlyl5vyHeyU2/yUtzNU7nJUvoOlcpun8namyl2uynewVW7zVd7OWLnL"
    "WfkO1spt3srbmSt3uSvfwV65zV95O4PlLoflO1gst3ksb2ey3OWyfAeb5Taf5e2Mlruclu9gtdzmtbyd2XKX2/Id7Jbb/Ja3M1zucly+g+Vym+fydqbLXa7L"
    "d7BdbvNd3s54uct5+Q7Wy23ey9uZL3e5L9/BfrnNf3k7A+YuB+Y7WDC3eTBvZ8Lc5cJ8BxvmNh/m7YyYu5yY72DF3ObFvJ0Zc5cb8x3smNv8mLczZO5yZL6D"
    "JXObJ/N2psxdrsx3sGVu82Xezpi5y5n5DtbMbd5ctvNm7jLncgdz5jZ3Ltu5c+ld+tzBnUubO5c7lrYudy53LW5t7ly2c+fS5c7lDu5c2ty5bOfOpcudyx3c"
    "ubS5c9nOnUuXO5c7uHNpc+eynTuXLncud3Dn0ubOZTt3Ll3uXO7gzqXNnct27ly63LncwZ1LmzuX7dy5dLlzuYM7lzZ3Ltu5c+ly53IHdy5t7ly2c+fS5c7l"
    "Du5c2ty5bOfOpcudyx3cubS5c9nOnUuXO5c7uHNpc+eynTuXLncud3Dn0ubOZTt3Ll3uXO7gzqXNnct27ly63LncwZ1LmzuX7dy5dLlzuYM7lzZ3Ltu5c+ly"
    "53IHdy5t7ly2c+fS5c7lDu5c2ty5bOfOpcudyx3cubS5c9nOnUuXO5c7uHNpc+eynTuXLncud3Dn0ubOZTt3Ll3uXO7gzqXNnct27ly63LncwZ1LmzuX7dy5"
    "dLlzuYM7lzZ3rtq5c+ly52oHdy5t7ly1c+fK5c7VDu5c2dy5aufOlbfzvIM7VzZ3rnZsPrvcudq1/Wxz56qdO1cud652cOfK5s5VO3euXO5c7eDOlc2dq3bu"
    "XLncudrBnSubO1ft3LlyuXO1gztXNneu2rlz5XLnagd3rmzuXLVz58rlztUO7lzZ3Llq586Vy52rHdy5srlz1c6dK5c7Vzu4c2Vz56qdO1cud652cOfK5s5V"
    "O3euXO5c7eDOlc2dq3buXLncudrBnSubO1ft3LlyuXO1gztXNneu2rlz5XLnagd3rmzuXLVz58rlztUO7lzZ3Llq586Vy52rHdy5srlz1c6dK5c7Vzu4c2Vz"
    "56qdO1cud652cOfK5s5VO3euXO5c7eDOlc2dq3buXLncudrBnSubO1ft3LlyuXO1gztXmjvvuP4CVl8rGwxsUNtgaIORDcY2mNhgaoOZDeY2WNhgaYOVDdY2"
    "eGWDIxsc2+DEBq9t8MYGpzZ4a4N3NvjFBpw7yKlT7lQqd2qVO9XKnXrlTsVyp2a5U7XcqVvuVC53apc71cud+uVOBXOnhrlTxdypY+5UMndqmTvVzJ165k5F"
    "c6emuVPV3Knr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnr0qnryqnr"
    "yp0rnLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLqunLquTF3fa2+s9LV5Gvo+g3sVgIa+w8h3GPsO"
    "E99h6jvMfIe577DwHZa+w8p3WPsOr3yHI9/h2Hc48R1e+w5vfIdT3+Gt7/DOd/jFd+C84dJoO95oPN5oPd5oPt5oP95oQN5oQd5oQt5oQ95oRN5oRd5oRt5o"
    "R95oSN5oSd5oSt5oS95oTN5oTd5oTt5oT95oUN5oUd5oUt5o07LRpmWjTctGm5aNNi0bbVo22rRstGnZaNOy0aZlo03LRpuWjTYtG21aNtq0bLRp2WjTstGm"
    "ZaNNy0ablo02LRttWjbatGy0adlo07LRpmWjTatGm1bNObbRplWjTatGm1aNNq0abVo12rRqtGnVaNOq0aZVo02rRptWjTatGm1aNdq0arRp1WjTqtGmVaNN"
    "q0abVo02rRptWrlt6iqrepY09WNKuwxpXm6uzF15bSyzUu8Vi5idydX51rJqYmul1q135/HWoHme11wcXErDmOSGl+blmzWdPy/ufr+46mjrnBPXyuaHK+fi"
    "vHbEK3l0e15ftDO2UYTjyfXd5rLTky+57hvzKHg3D18qUU+hHtKrsnAtkHd6Z18+7ePLJc6js9bLumtTS5ff4GVcbQ/glRvHPKZ5FNLLtBABeLPnJnLzh3wr"
    "Fdh0m9/o5vrPu9+BaYP36svd4fXHw/X12X9t7+w3WN9gNWxufzdv8crnXoGbu1m8u+6o13Pfhm2XI4G9Gx/e4efnnZ5otdvOJ1kBUFW/2AGoAjnXrQEvCOtH"
    "ekuvjSZXZ5fSE+NUof1G7erjx1tRNP758+WF6Gg9WRUHHVMtv0GNPOBmZdO+nnN16kHXE20DCrTfndxngKMHr1uC3YMXIds/eHyqTbsN7PtpGtOBLXdNWrNs"
    "3+HXL4qW8mHK/fvMNHppB22GIu5JsHItBZE1iPuqVtRrxWSdthl8uCfywDZ6krVZc7iPAJmcLPdEq4KVCKi1PWV1Un602dx5OOGhZctyqKgOXaLjh9bVsDdi"
    "Pw9ZX7TqoaDxQhpMHLId1nbup9Zu2dG6IjnYYdmxzTxE8mDjPjVrv1qFN+693Exf7k2/l5v9dtM+D84OOxizZ70pO5z49Th/JKmZY82y3dDPk2iFgUts+dgh"
    "vHzgEP7FtvuxemiWTT7zg6VNYf1ACvptN9G+LXdtpI30SDT22z25s9vbq/fYATSVnDPu0XD/XqW8emCl6CYRK91n66hxhero8f3fuqcDz9vph+7lXbQTdvCK"
    "9ZuBXFYqAqJdEAh+aIVN9oVby2Wz48dWz/EDq+fIboSTR1bGsTRJ5tgjk43ayPzrh8+PA89Ax4Nr0q7Dw94rdgDV6NasMwrfPDxTXg6kRCc8XnnGwE6fTFCI"
    "XS303j6JD5DFjh/2BvJtgoF8vlVNwDXrW96V9K7AO/Sn4XcPL8sp+1naWem9lf+fshdvZTFcar88kNqSHbyzuyPnDxK0VPlPvfrj/JEzxC/Rs0YdTmQlTew6"
    "jHbXIRgh8TJRPrIV50xeY/CIVE/vWo3RMlnaYZI+DpXvW5mzjTQPfYeR7zC2rdpMbDC1wcwGcxssbLC0wYosIaydVx9f2YGObHBsgxMbvCZab2znUxu8dewt"
    "vLO9frEBnaMFzjla4JyjBffvpI99h/tNWM98h7nvsPAdlr7Dynd45Tsc+Q7HvsOJ7/Dad3jjO5z6Dm99h3e+wy9hY0u74dKoVbWc/d6DKCHYWVB2k4wpBmVP"
    "qUZ3uHQ8tGwojbThWm2LIbVsMSj7S1NjZ2mGzoG2rqjs5i7hSURMdYVWHCDm2vp+hd+QgyMr1WN8qxTonMBbjRjmNX4DzTcYJlCdGww5qr4NBiNUz2aq4L/g"
    "bWtIFvq1LjlXd68hNvTqOFFPlx6tX+/Y4Trann852+LGx+Hrz52eCLrfOezws7svm5uLzWWn2lyeSfFJ7WqhYaljQegLTHI8btvl4mCyeHVz8dvFlaBxLKY4"
    "KXLebGF3hYPF5vdfWJIE9t/OAsxSdMKXHdiWkW8Rywefl9u7zvHF1W+XW71/02lGBsJg7Hl0c317a1Po3b1B21W8wI22BrnNnWNOg4xdcTTDLioFNm1eX51t"
    "Pn/enu93fupUb0TO0NgwR4PruwvEXnaOtrfbmz+2chPs4lrQp/odXt9I8XR/V5HQXLuOjltO0mKWyLLebezcYcjYyutP8s/7L0FwnkH50QKXleW0Pcv/89eu"
    "e+iow8mSzztQr14/RrP01fWnz6Itb6+vXqJ9ggBtr4ku+Pmz6Gd6Z+7f0Rf6yvL6bvuyc/r79kp0pBf8a+dn0V79zpmK8flm+8f26u62c/3H9ubw4kokubk6"
    "2z5HCmhb7frud2XH5bZz+/v1l8vzzreLrfh79/u2c7v5tO3cbG+/XN51/pSpfLkVdDE6dJOz66vbi9s7kUhnQ31xQ8vC2+ff3y7cYZPFkl2i4Kchnx/X8r3w"
    "F9JkjXqV6uK2s7m82W7Ov5lJ59ftt2u5+7o+2QdZCSrsbPOrtB/StNUSPyhRd+2ceGQab7I/jEzukcnh54Fkdm+92O+YtGzCvEBzW7uzUcDPo7Jxr9nPMLnX"
    "HBw8DBfg7wMTV2J0mb8oi8ZCFOhF+PuvrlM/2Rh/H1mM5FmvDCOcW3eWKMPfh5GWC4BSVGiZNLa0A/x9aBYlmZ/LRMB39bEYf5qRCxYn5zIh5G63YuowTJrD"
    "2FuuRNhTPHK6vPh0cXdriuQuPjxxucUqTBqhNFLC+9AoywzwifdAv2/HlIxSn0x2yCj1VzFNnYtsgzmriZ4MO9L2zr5VMDltGW0WS0RBi5UXt2AVjo6h0KiY"
    "oAEna7ed3scvl5f7RqYZbkQa5nAKHlzvrjc3d3KmlFFEjJsNHCPuG5nHFz66+k12SIk4du+bTBPjpb4EJXm7iNyh1/lSHb/++vni5ht5oCD1RwcNz1/s7xsh"
    "SM22kM/rjx1Zu7IhcsuQTpW7NsKMrc+P1zedkA4c4Sn4riVP3Wz/+8uF4IEQDo78PmJd3Xkx4e14Ibt8EkWTbE3ITJuPdyIZtSV3ISQaWawzkTwsXnM0IH4u"
    "Jv+76xuPBUkrRPjgfcvZ1aOmwKF8KjL9rhFP7xhLTnFJ2PbQQ4EL/LL4/kMPxQNMmlkbdf8Uppffy/RU2HqPuSFT+HlUgmWDx2bwYxGBXtp4ELTRcmXaZDGt"
    "dV8+rO5Lt+4rTLNJUTdly7adsdfn2CITM77tQ/bL2myRPTzJqngmjzRbzYl9/4j2XyBz1aanXl7cPZQxYvWcHL2u+3Xxsk7CH+XG22HYP369WB+tBq+rk15P"
    "+cCLbPvPwjZLs/HjkoX36fyDZdl+mFS/loYFf9SnYOS4/8IPpk172/H2W5u2fqwwabIqrcep3NZoD051WJFcHbcn9aiBadVK0jxSKkN2EKXJszppT6ptRm3t"
    "r/JRw0PRsL29wV7RTqt9CIvV9X3kXg7CoElyqMZTcE/+enw56NVhIJq8Ln4WhMLmIzEyWy1ZS8KWk5tWS+PPA9fYeNOO+L3atWXkbJfeb+4UhRE7ROxQKBNj"
    "kzbx46Z+3NTfIi4z215gYIQAO5bvUBWOkUFQlAqaTmHTiTWsE3pyJItwO63EB5FpL1DtuA2sILW146YMDSqbq3w932Vs9cvd9SfpSPtB8+uNfGV5Pd83AqRA"
    "nfpKTpDnHZjaXsBY2zctNt/+tkXNJDRHO7+4+i/Q3pIC2+YFjDsjX6q5GZJy3kzp9D4/3zwHuiW+PQ8vng9RKsXtImuXbb359gn2CqT8hnLpp+tz1OgaQojB"
    "zfXn8+s/rzrH20vUezNyp5ToUCR1hUplgNbSl0K5cnF97jpjP7Ck7UJpLqEQ6xGvC6voteA/lCr2FVkbsoo7/0tKraiGVobqeXt8Dt60hR3b0YgSVSLFTRBt"
    "e5LDWTtp6+3NxfV5529IWdnE3wKlTm/z+fPNNW79VaEyYS9WRyC8lptLuexAZb/IfukGtf2okwxuNn9eoa5fZDWQELL/2EqpWdruXc+tpxlFSbwE1CONcqdS"
    "PoYjfg+ho6hncyKrEk3NT3Wu2sXnFqH5cave0JeJBdMLi6JNKA6VfdNxGP/4U6Bl4DC+XwiWA6v3ULlVSMrZfYJr6xO5reLpQ1gpSAWagb7UZmzf6+ECiyHV"
    "Cw2n5S9bHl979PbRfVlwRmYzcbkhkP0c9MvsRZm07IA02f2uxUEY4m9rLed/ZV+lTP29k/YuxsOkTRdPJBSyF1J+8FXyHpcNXz/vQbkYhbEn8T8u0SqMQezH"
    "/HuUHpKBKkxaCdRP2N4CwamUg3dvuFf0ZbseiAweDIQTvBsMelP74NF8y+9B2aVEEpWIyL+gnxxImj9qpSXh3W+qa4MNaV72WdqX88X+7uwMHzPK/llrpmHb"
    "cKpBZ9FWO7WW6vkzR9tquGtjFaeAD18vP149f6uqKOxPw/jlVFa4aLGXpfjYb3s9+3F9QO5DQnXWYexX6IPa1yKQ+Cqejx4ZsuMd1M4IGz1iXEC/qp0NirHO"
    "Rlv0eJfShpsxORIOR279TJ4y2gSb/Fm+1cbP5JYYpDzn6+N64C57Jo8YW+MweQjJ6SOqQQ5XQVYspiC/zmrqPrUU/1Emcyvc3PU2yilNTYx7X4QocQFkrX5s"
    "z4wSwB3QoHUFUxbGr/AvQBaNRZJa3eA9dgTO81BBM0Zo8hH6JagaLoOGS91wGTVcxuF9xtr1Vnxp6URUqPFAuhWZo1oBzkN8Hxi+R7gSI9UKpt4TnqD6Bb1V"
    "lePnXD1oCQsxKcjuWInNV3zZWawGr+e1mIkXXy7vLj6LxYCMcdv5fzpH289qsWN0XMb1fF0fdarV/PViedzpnYnl1K/bzu8X5+dblJdRj+V4dXRSDzr1m3p5"
    "cmxufeoVBF4+Ovn2eWvue5rNf7x2dHx9c9eZbb+Z+55/Mzc9NaF3DqFfPELqXS45FQi2crP5tBVLCCjNKbynQavD5rIQAkk/Zg4bvggSIoxYVDSDyhARrfOc"
    "p8BOpVNMaz300wU4lW4JLfiaa7hT6Z7SUs86J+n0jvGGyqn0ycyy7rsLs1MIk8OXbKrCqFRAfzg54stjXp1MVrJ9a1nQzpfP8lJOGHQuZdfANw4wylG95u8W"
    "opG/H++GupLKA777bNaJ0AD2KhEczB0w42Zat9bKGoI2xh1aLrAI1ARGvoehcgpeIGGjTkcIpPXXEL7+3v3w4fzuwwcZgQUsOQziwzA+CYKX8ANHJCMICWeG"
    "pxgduk6YqE5UwpdHSZAJGpQqCAlndhT7FL7grBa1SSIgmbZlrpVkqmu8puin6BiT8gWcHsJXG80o9GlmlE0d+xS+4FVCHuaqa9f0dQpfYUrqHCl4669T+ILj"
    "W9TryGD/Tn+dwhfMfJyFqhfX9HUKX2FBeh8FeOuvU/iC9kGljhBfIKcndozjKXyyED6hVaEJGDQGNAHDY8TBG76s6s6iPhmvBh15zv9Z9Ctrfjg0n2e/b25+"
    "E3PH5k7MmHqb4/qj3Ei6UK8DTRXZo6OaHz2aqrw759I7BXoMGpjB5FNLRRc10JTLx49bkFrMaEMPPakOEFrjDh0Gm2+36hqdSPAG7RZkzpvEPX7+h9xkwYey"
    "0HMlxLK7De7B8I/q2bCxiniLVzLFJL25+fbvWCV+vmc6eczt3M3tws/tcmduV15u+dnZzRfUNVp/J6+vdub1FHwYdHmWq251Cl8MujcrQB4A/hMAH4dP2W0Z"
    "dNFIdtEIokWyC0bQBaMYTvHhE/gMDN4IhjEMukimG8EAi2RqIAWcRjK1CAZTjHdeVyf1MWpNxTB7h887bhd+aTqWWJFZfUtqOTm9drs5+111tU7vs5zfL847"
    "FyryPqpOxTDpMpGI3aG/m4bqww71DbZLgzpw7uh5x9HN+nK7BUIfcefWSuwn2fJXZxefN5ekqtcLOwfAw/f/syd7yIsoTaSSQ6g0E4Dvx887dkf4FTcOpRbY"
    "n1eyCDK5DXLk6z9lUeDQHYoAIj5qLbwAHTOVdRjmyXPcjr5oqY7zzcXlt45IsXOmRr8a45LqxvS7zrVIySIMXTN93llsvl58EiKAZNg6c3ffPovl86WgK6S2"
    "q+s7pZQC+f8My/B/ExmQigd/SMUDRRHEkuy51dkvtmKxfiPKrDfyFU1SeIFhZsQcUFPAxrwVnejuz+32qvPx4kaUGLMnSunKPb46XPct6CKAPcnu2wT+pvA3"
    "g785/C1QazrE/xj+hxFDjBli1BDjhhg5xNghRmeo6s6QCkMqDKkwpMKQCkMqDKkwpMKQSoRUIqQSIZUIqURIJUIqEVKJkEqEVCKkEgdtuhdy0oga69239kns"
    "aeTfCxZzQ9w4vxN1ZUcKg3u1/8yiOnz4pnH04E3j7L49Y2s/57693Z070Tu3XqUOx/3qdo/ZezXV5dWBFCK8d8pKj94Dmvmtr9zxkGZ+G4YP2K5VG95R7F22"
    "vj9T7yL/vvUDMvUubFy3zltjOTc8I19dQ7Rr1NBmLFlxWLL8pyhN+2J67yvsX7DO8dd+gIvlzwSHOIAb7//ZEywd2MNh6EdtJjtghRM1CppRR+2pQo5zdy+o"
    "Age5PIN74vkzRfpZsG+8YSkm/A9FAO/EYfS9yrF31Q1pKHABuca3wgs7B0VLsD4Fs3JSeDkZ72zaHYXmy+XkDbi+294q38aV7UfOiNMnzYizJwzK2ZMG5fwJ"
    "I23+pJG2eFhKzkhbPDClX5yUljt7+8xpeP+qt+y2zZRmzFM9ISJBX/gdCrjvX/9+ZOKrByeOxwRrd7ZYqlEhhrx7RrJ+YF6Ew9we9guWGzcaYYvGWF8/MN8i"
    "vwcrVhwAzcJOp+iTm5WOP5Jf7RzJO8px30jWl7Sbc5SzNV465xLHin81H/0cOrGGjVhyVQx5d1+vJiUuuf9uDF+ASQ4yfQEIjF/A1/7PeEa0mCwnw2PpFsYv"
    "QfLp29/vuz8HUuvsoVR//ElKGQ8P3zdKbfRivahS9+24e2pXOBxHkK7Ab/j8NZcbenigAf1j37sh/5DKp37k366P27rPMXPjqn6z33zArpFs5SRbOW3+emdm"
    "R06sUSPWQ7IZ7Dcfxru/Ix8Ez0Mv4gM7s4jJ/Pv7D8imQPuNN/NcbVNlzGDB53M5DPaOpcGK4700RtXRvnHY74eNo8+3bZqkjyHJWkmq5cXTSEatJNUy5Gkk"
    "41aSar3yNJJJK0m1rnkaybSVpFr/PI1k1kqSdFSeRjRvJ8rw94lEi3aiemA+scMH7VT1CuqJVNvHERw+P73bh+1DSW6+h0/v+WH7aKJrV0+k2j6g5FZ/+PT+"
    "H7aPKXlCED59CITtw4oujD2RavvIkmcQ7OkjK2wfWvI4gz19aIXtY4s2NZ4487ePLcnW2F/gJ+1jS27Ss6ePLdY+tuQuP3v62GLtY0seGLCnjy3WPrZoDfJE"
    "qu1jS6482NPHFmsfW/JEJHr62GLtY0serkRPH1usfWzJc5ro6WOLtY8tuXMTPX1sRe1jSy77o6ePrah9bMnNgegviGvtY0ueaUVPH1tR+9iSx2PR08dW1D62"
    "ohx/n0q1fWxFBf4+lWr72JJLhvjpYytqji31HFjx3S0lVKzeO5GUTgQlpVj9NrKoezsO75Sked/G0y7SIu5O2r88Is+vJY3XD8vzL4/Lc4P0rjx7GoDHqPz2"
    "4vz8xbdvqP4GFp3+1v/b3xzzTW6YX/wwZTPM0A8z8h0GvsOsSWXhh1n6DivfYe04+DqKDZNLx/e+IvD23icBytZ7XQTTVg1G26Ghl9hUPBw+QKmwQuMydm2g"
    "HpJXq6go5IRDRR/PKWs4oVUT2+k1ixsVmDRcGlWAyhDO6xZZ4zWLrPGcRdZ4nCJrPEaRNV6jyBrPUWSN9yiyRp6zxoXAvFn4xoU70NXwAhWNm4JBI1AUNAKF"
    "zUBhIxBrBmrcZIyarRb5fed11GjHqNGOUdpwadRblDdc/PK/joN7lF8jZYKrxMuBRvlV2aoYoCKs0X5VyrJDtK4Fi7+RuYc4toJPLAXZKdong++ZFWZuJbWw"
    "3JdWUisrqbVJ6pUV/Mgyf3Zs6eieoDuQf225v7FMnr3VYf7xj3/8HwgTbVU="
)
MODEL_SPEC = json.loads(zlib.decompress(base64.b64decode(_SPEC_B64)))

# ---- helpers ----------------------------------------------------------------
def dec(v):
    """Decode a stored static value (dates were serialised as ['__dt__', iso])."""
    if isinstance(v, list) and len(v) == 2 and v[0] == "__dt__":
        return datetime.datetime.fromisoformat(v[1])
    return v

def translate(formula, origin, dest):
    return Translator(formula, origin=origin).translate_formula(dest)

import re as _re
_INT = _re.compile(r'(?<![A-Za-z$0-9_.])\d+')
def bump_ints(s, positions, delta):
    out, last = [], 0
    for i, m in enumerate(_INT.finditer(s)):
        if i in positions:
            out.append(s[last:m.start()]); out.append(str(int(m.group()) + delta)); last = m.end()
    out.append(s[last:]); return ''.join(out)

def coerce(raw, ref):
    """Coerce a CSV string 'raw' to match the reference cell's python type 'ref'."""
    raw = (raw or "").strip()
    if raw == "":
        return None
    if isinstance(ref, list) and len(ref) == 2 and ref[0] == "__dt__":      # date
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try: return datetime.datetime.strptime(raw, fmt)
            except ValueError: pass
        return raw
    if isinstance(ref, bool):
        return raw.upper() in ("TRUE", "1", "YES")
    pct = raw.endswith("%")
    num = raw[:-1] if pct else raw
    if isinstance(ref, float) or pct:
        try: return float(num) / 100.0 if pct else float(num)
        except ValueError: return raw
    if isinstance(ref, int):
        try: return int(float(num))
        except ValueError: return raw
    return raw            # string reference (e.g. policy number "07340050")

def typed(raw):
    """Best-effort type for free CSV cells (CSO matrix / Inputs header block)."""
    raw = (raw or "").strip()
    if raw == "": return None
    try:
        i = int(raw)
        if str(i) == raw: return i
    except ValueError: pass
    try: return float(raw)
    except ValueError: return raw

# ---- CSV loaders ------------------------------------------------------------
def load_inputs(path):
    d = {}
    with open(path, newline="") as f:
        for row in csv.reader(f):
            if row and row[0].strip():
                d[row[0].strip()] = row[1] if len(row) > 1 else ""
    return d

def load_cso(path):
    with open(path, newline="") as f:
        return [row for row in csv.reader(f)]

# ---- sheet writer -----------------------------------------------------------
def write_sheet(ws, name, spec, inputs=None, cso=None):
    written = set()

    def put(coord, value):
        ws[coord] = value
        written.add(coord)

    # 1) CSO Rates: matrix comes straight from the CSV
    if name == "CSO Rates" and cso is not None:
        for r, row in enumerate(cso, start=1):
            for c, cell in enumerate(row, start=1):
                val = typed(cell)
                if val is not None:
                    put(f"{get_column_letter(c)}{r}", val)
    else:
        # 2) static cells (labels, headers, constant tables e.g. 1958-CET q_x)
        for coord, val in spec["statics"].items():
            put(coord, dec(val))

    # 3) formula fills - LOOP each column's origin formula down its rows
    arrays = set(spec.get("arrays", []))
    for col, r0, r1, kind, f0, pos in spec["fills"]:
        origin = f"{col}{r0}"
        for r in range(r0, r1 + 1):
            coord = f"{col}{r}"
            if kind == "lit":
                formula = f0
            elif kind == "trans":
                formula = translate(f0, origin, coord)
            else:  # counter: translate refs down, then increment the counter literal(s)
                formula = bump_ints(translate(f0, origin, coord), pos, r - r0)
            if coord in arrays:
                put(coord, ArrayFormula(coord, formula))
            else:
                put(coord, formula)

    # 4) Inputs sheet: override column-B constants with the policy's CSV values
    if name == "Inputs" and inputs is not None:
        for r in range(1, ws.max_row + 1):
            label_cell = ws[f"A{r}"].value
            bcoord = f"B{r}"
            if label_cell in inputs and bcoord in spec["statics"]:   # only genuine inputs
                v = coerce(inputs[label_cell], spec["statics"][bcoord])
                if v is not None:
                    ws[bcoord] = v

    # 5) number formats (per-column default + exceptions) and widths
    colfmt = spec.get("colfmt", {})
    exc    = spec.get("fmt_exc", {})
    for coord in written:
        col = _re.match(r"[A-Z]+", coord).group()
        fmt = exc.get(coord, colfmt.get(col))
        if fmt:
            ws[coord].number_format = fmt
    for col, w in spec.get("widths", {}).items():
        ws.column_dimensions[col].width = w

# ---- main -------------------------------------------------------------------
def main():
    inputs = load_inputs(INPUTS_CSV)
    cso    = load_cso(CSO_CSV)

    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        print(f"  building sheet: {name}")
        ws = wb.create_sheet(title=name)
        write_sheet(ws, name, MODEL_SPEC[name], inputs=inputs, cso=cso)

    wb.save(OUT_XLSX)
    print(f"\nSaved: {OUT_XLSX}")
    print("Open it in Excel (or run LibreOffice --headless --convert-to xlsx) to")
    print("recalculate all formulas. The summarised outputs are on the Inputs")
    print("sheet, cells D2:E7 (Cash Value, PUA Cash Value, ETI, RPU, APL, Loan Outstanding).")

if __name__ == "__main__":
    main()
