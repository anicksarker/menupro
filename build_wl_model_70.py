#!/usr/bin/env python3
# =============================================================================
#  build_wl_model_70.py   (Plan 70 - Whole Life)
# -----------------------------------------------------------------------------
#  Rebuilds the workbook from ONLY two CSV inputs:
#         1) Inputs.csv        (structured: scalars + dividend table + loan table)
#         2) cso_rates_v2.csv  (mortality rate matrix - UPDATED: 17 columns,
#                               now incl. the new female columns 41FX & 58FX)
#
#  Same engine/architecture as the Plan 34 builder. Differences for Plan 70:
#    * 13 sheets (adds DIvidend_Assumption): CSO Rates, Output, Inputs, MORTALITY,
#      Commutation, CV_RATES, PUA_Div, DIvidend_Assumption, Projection_Monthly,
#      RPU, ETI, APL, Loan.
#    * CSO Rates now has 17 data columns (A:R). Row-4 CONCAT(Table,Sex,UW) codes
#      are generated for every data column present, so the two extra female
#      columns (41FX, 58FX) are picked up automatically.
#    * MORTALITY's INDEX/MATCH lookup ranges are widened $A$5:$R$126 / $A$4:$R$4
#      so a female policy on table 41 or 58 resolves to the new columns.
#    * Inputs sheet keeps columns C..G exactly as the reference - including the
#      F/G "Legend" + "Vanish Premium Summary" block (statics + formulas).
#    * Output sheet = summary A..G only. The external "Audit"/"Balancing" columns
#      (which XLOOKUP into other workbooks) are NOT generated - they need external
#      audit files and are not part of the input-driven model.
#
#  Repeated (fill-down) formulas are written with a LOOP via Excel-style
#  translation (relative refs shift, absolute $refs stay, integer counters bump).
#
#  Usage:
#     pip install openpyxl pandas
#     python build_wl_model_70.py Inputs.csv cso_rates_v2.csv WL_Model_70.xlsx
# =============================================================================
import sys, csv, re, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula

INPUTS_CSV = sys.argv[1] if len(sys.argv) > 1 else "Inputs.csv"
CSO_CSV    = sys.argv[2] if len(sys.argv) > 2 else "cso_rates_v2.csv"
OUT_XLSX   = sys.argv[3] if len(sys.argv) > 3 else "WL_Model_70.xlsx"

SHEET_ORDER = ['Output','CSO Rates','Inputs','MORTALITY','Commutation','CV_RATES','PUA_Div',
               'DIvidend_Assumption','Projection_Monthly','RPU','ETI','APL','Loan']
INPUTS_HEADER_ROW = 1                 # Inputs.csv line 0 -> Excel row 2
FORCE_TEXT_LABELS = {"Policy Number"} # identifiers kept as text even if all-digits

# =============================================================================
#  Fill / helper engine
# =============================================================================
_INT = re.compile(r'(?<![A-Za-z$0-9_.])\d+')

def _bump_ints(s, positions, delta):
    out, last = [], 0
    for i, m in enumerate(_INT.finditer(s)):
        if i in positions:
            out.append(s[last:m.start()]); out.append(str(int(m.group()) + delta)); last = m.end()
    out.append(s[last:]); return ''.join(out)

def _translate(formula, origin, dest):
    return Translator(formula, origin=origin).translate_formula(dest)

def _dec(v):
    if isinstance(v, list) and len(v) == 2 and v[0] == "__dt__":
        return datetime.datetime.fromisoformat(v[1])
    return v

def _cell(x):
    """Type a raw CSV string for the CSO matrix."""
    if x is None: return None
    x = str(x).strip()
    if x == "" or x.lower() == "nan": return None
    try:
        i = int(x)
        if str(i) == x: return i
    except ValueError: pass
    try: return float(x)
    except ValueError: return x

_DATE_RE = re.compile(r'^\d{1,2}-\d{1,2}-\d{4}$')
def _infer(x):
    """Infer the python type of a raw Inputs.csv cell (label, date, %, number...)."""
    x = (x or "").strip()
    if x == "": return None
    if _DATE_RE.match(x):
        for fmt in ("%d-%m-%Y", "%m-%d-%Y"):
            try: return datetime.datetime.strptime(x, fmt)
            except ValueError: pass
    if x.upper() in ("TRUE", "FALSE"): return x.upper() == "TRUE"
    if x.endswith("%"):
        try: return float(x[:-1].replace(",", "")) / 100.0
        except ValueError: return x
    xn = x.replace(",", "")                       # allow thousands separators e.g. 6,657.39
    if re.match(r'^-?[1-9]\d*$', xn) or xn == "0":
        return int(xn)
    if re.match(r'^-?\d*\.\d+$', xn):
        return float(xn)
    return x                                      # plain string

# =============================================================================
#  CSV loaders
# =============================================================================
def load_inputs(path):
    """Ordered list of (colA, colB) rows - blank separator rows KEPT so the row
    positions line up 1:1 with the Inputs sheet (scalars + tables)."""
    rows = []
    with open(path, newline="") as f:
        for row in csv.reader(f):
            a = row[0].strip() if row else ""
            b = row[1] if (row and len(row) > 1) else ""
            rows.append((a, b))
    while rows and rows[-1][0] == "" and str(rows[-1][1]).strip() == "":
        rows.pop()                                # drop only trailing empties
    return rows

def load_cso_df(path):
    return pd.read_csv(path, header=None, dtype=str, keep_default_na=False)

def cso_concat_fills(cso_df):
    """Row-4 =CONCAT(colN1,colN2,colN3) code for EVERY data column present.
    Auto-adapts to the 17-column v2 table (or any width)."""
    ncols = cso_df.shape[1]                       # incl. the age column (A)
    fills = []
    for c in range(2, ncols + 1):                 # B, C, ... up to last data col (R for v2)
        L = get_column_letter(c)
        fills.append([L, 4, 4, 'trans', f'=_xlfn.CONCAT({L}1,{L}2,{L}3)', []])
    return fills

# =============================================================================
#  Sheet writer
# =============================================================================
def write_sheet(ws, name, spec, inputs=None, cso_df=None):
    written = set()
    def put(coord, value):
        ws[coord] = value
        written.add(coord)

    if name == "CSO Rates":
        for _, row in cso_df.iterrows():
            ws.append([_cell(v) for v in row.tolist()])
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if ws.cell(row=r, column=c).value is not None:
                    written.add(f"{get_column_letter(c)}{r}")

    elif name == "Inputs" and inputs is not None:
        # paste col A (labels / dates) and col B (values) positionally
        for i, (cola, colb) in enumerate(inputs):
            r = INPUTS_HEADER_ROW + 1 + i
            a = _infer(cola)
            if a is not None:
                put(f"A{r}", a)
            if cola in FORCE_TEXT_LABELS:          # e.g. policy number stays text
                b = colb.strip() if colb and colb.strip() else None
            else:
                b = _infer(colb)
            if b is not None:
                put(f"B{r}", b)
        # columns C..G (metadata legend + Vanish Premium Summary) come from statics
        for coord, val in spec["statics"].items():
            put(coord, _dec(val))
        # formula-driven cells (B overrides + G8..G13) get written by the fills below

    else:
        for coord, val in spec["statics"].items():
            put(coord, _dec(val))

    # formula fills - LOOP each column's origin formula down its rows
    arrays = set(spec.get("arrays", []))
    for col, r0, r1, kind, f0, pos in spec["fills"]:
        origin = f"{col}{r0}"
        for r in range(r0, r1 + 1):
            coord = f"{col}{r}"
            if kind == "lit":
                formula = f0
            elif kind == "trans":
                formula = _translate(f0, origin, coord)
            else:                                  # 'counter' - bump integer literal(s)
                formula = _bump_ints(_translate(f0, origin, coord), pos, r - r0)
            put(coord, ArrayFormula(coord, formula) if coord in arrays else formula)

    # number formats + column widths
    colfmt, exc = spec.get("colfmt", {}), spec.get("fmt_exc", {})
    for coord in written:
        col = re.match(r"[A-Z]+", coord).group()
        fmt = exc.get(coord, colfmt.get(col))
        if fmt:
            ws[coord].number_format = fmt
    for col, w in spec.get("widths", {}).items():
        ws.column_dimensions[col].width = w

# =============================================================================
def main():
    inputs = load_inputs(INPUTS_CSV)
    cso_df = load_cso_df(CSO_CSV)
    MODEL_SPEC['CSO Rates']['fills'] = cso_concat_fills(cso_df)   # 17-col aware
    wb = Workbook(); wb.remove(wb.active)
    for name in SHEET_ORDER:
        print(f"  building sheet: {name}")
        write_sheet(wb.create_sheet(title=name), name, MODEL_SPEC[name], inputs=inputs, cso_df=cso_df)
    wb.save(OUT_XLSX)
    print(f"\nSaved: {OUT_XLSX}")
    print("Open in Excel to recalculate. Summary is on the Output sheet (A3:G3):")
    print("Policy No, Cash Value, PUA Cash Value, ETI, RPU, APL, Loan Outstanding.")


# =============================================================================
#  MODEL DEFINITION  (fully editable - every formula & constant is below)
#  Extracted faithfully from WL_Model_79_latest.xlsx (the Plan-70 reference).
# =============================================================================
MODEL_SPEC = {
  'Output': {
    "statics": {
        'B1': 'Calculated values',
        'A2': 'Policy Number',
        'B2': 'Cash Value',
        'C2': 'PUA Cash Value',
        'D2': 'ETI',
        'E2': 'RPU',
        'F2': 'APL',
        'G2': 'Loan Outstanding',
    },
    "arrays": ['B3'],
    "fills": [
      ['A', 3, 3, 'lit', '=Inputs!B5', []],
      ['B', 3, 3, 'lit', '=_xlfn.LET(\n    _xlpm.ValDate, Inputs!B12,\n    _xlpm.DateRange, Projection_Monthly!$B$2:$B$1050,\n    _xlpm.CVRange, Projection_Monthly!$L$2:$L$1050,\n    _xlpm.ExactMatch, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.CVRange, "", 0),\n    IF(_xlpm.ExactMatch<>"", _xlpm.ExactMatch,\n        _xlfn.LET(\n            _xlpm.DateBefore, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", -1),\n            _xlpm.DateAfter, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", 1),\n            _xlpm.CVBefore, _xlfn.XLOOKUP(_xlpm.DateBefore, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.CVAfter, _xlfn.XLOOKUP(_xlpm.DateAfter, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.DaysBetween, _xlpm.DateAfter - _xlpm.DateBefore,\n            _xlpm.DaysFromBefore, _xlpm.ValDate - _xlpm.DateBefore,\n            _xlpm.Fraction, _xlpm.DaysFromBefore / _xlpm.DaysBetween,\n            _xlpm.CVBefore + _xlpm.Fraction * (_xlpm.CVAfter - _xlpm.CVBefore)\n        )\n    )\n)', []],
      ['C', 3, 3, 'lit', '=IF(Inputs!$B$20="Buy PUA",PUA_Div!$E$14,0)', []],
      ['D', 3, 3, 'lit', '=ETI!E6', []],
      ['E', 3, 3, 'lit', '=RPU!B17', []],
      ['F', 3, 3, 'lit', '=APL!F10', []],
      ['G', 3, 3, 'lit', '=IF(Inputs!B40="D",Loan!$G$10,Loan!$G$11)', []],
    ],
    "colfmt": {'G': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)', 'A': '0', 'E': '#,##0.00', 'C': '#,##0.00'},
    "fmt_exc": {'G2': 'General', 'B3': '#,##0.00', 'A2': 'General', 'E2': 'General', 'D3': 'mm-dd-yy', 'F3': 'mm-dd-yy', 'C2': 'General'},
    "widths": {'A': 16.18},
  },
  'CSO Rates': {
    "statics": {
    },
    "arrays": [],
    "fills": [
    ],
    "colfmt": {},
    "fmt_exc": {},
    "widths": {'A': 8.73},
  },
  'Inputs': {
    "statics": {
        'C1': 'Data File Name',
        'D1': 'Column Name in Data file',
        'C2': 'Selection',
        'F2': 'Legend',
        'C3': 'pconvert_benefit_info',
        'D3': 'c_bn_benf_undrwrt_class',
        'G3': 'Link with Data',
        'C4': 'pconvert_policy_info',
        'D4': 'c_pol_product_id',
        'G4': 'Formulae',
        'C5': 'pconvert_policy_info',
        'D5': 'c_pol_policy_number',
        'G5': 'Dropdown Selection',
        'C6': 'pconvert_benefit_info',
        'D6': 'c_bn_benf_sex_code',
        'C7': 'Formula',
        'F7': 'Vanish Premium Summary',
        'C8': 'Selection',
        'F8': 'Monthly Premium',
        'C9': 'pconvert_benefit_info',
        'D9': 'c_bn_original_units',
        'E9': 'For Base Policy',
        'F9': 'Vanish Point (Month #)',
        'C10': 'pconvert_benefit_info',
        'D10': 'c_bn_original_units',
        'E10': 'For PUA',
        'F10': 'Vanish Point (Date)',
        'C11': 'pconvert_benefit_info',
        'D11': 'c_bn_benf_issue_date',
        'F11': 'Vanish Point (Duration)',
        'C12': 'pconvert_policy_info',
        'D12': 'c_pol_paid_to_date',
        'F12': 'Total Accum Div at Valuation',
        'C13': 'Formula',
        'F13': 'Current Out-of-Pocket Premium',
        'C14': 'pconvert_name_addr_rela_info',
        'D14': 'c_na_policy_name_dob',
        'C15': 'Formula',
        'F15': 'Note: Vanish Premium Feature',
        'C16': 'Formula',
        'F16': "To use vanish premium, select 'Vanish Premium' from",
        'C17': 'Formula',
        'F17': 'the Dividend Option dropdown in cell B18.',
        'C18': 'Formula',
        'C19': 'User Selection',
        'C20': 'Basin Dividend History',
        'D20': 'DIVIDEND_DESCRIPTION',
        'C21': 'Based on Mortality rate',
        'C22': 'pconvert_benefit_info',
        'D22': 'c_bn_ba_prem_amt',
        'C23': 'Formula',
        'C24': 'pconvert_policy_info',
        'D24': 'c_pol_policy_mode',
        'C25': 'Formula',
        'C26': 'pconvert_policy_info',
        'D26': 'c_pol_policy_status',
        'C27': 'Formula',
        'C28': 'Formula',
        'C31': 'Basin Dividend History',
        'C32': 'Div Date',
        'D32': 'LAST_ANNIVERSARY_YEAR',
        'C33': 'Div Amount',
        'D33': 'ACTUAL_TOTAL_DIVIDEND_AMOUNT',
        'C40': 'pconvert_loan_info',
        'D40': 'c_ln_loan_curr_int_rate',
        'C41': 'pconvert_loan_info',
        'D41': 'c_ln_loan_int_method',
        'C43': 'Loan_History_LP',
        'C44': 'Loan Date',
        'D44': 'LOAN_DATE',
        'C45': 'Loan Amount',
        'D45': 'LOAN_AMOUNT',
    },
    "arrays": ['G10', 'G11'],
    "fills": [
      ['B', 7, 7, 'lit', '=ROUND((B11-B14)/365.25,0)', []],
      ['B', 13, 13, 'lit', '=IF(DATE(YEAR($B$12),MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12),MONTH($B$11),1),0))))\n<=$B$12,\nDATE(YEAR($B$12),MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12),MONTH($B$11),1),0)))),\nDATE(YEAR($B$12)-1,MONTH($B$11),MIN(DAY($B$11),\nDAY(EOMONTH(DATE(YEAR($B$12)-1,MONTH($B$11),1),0)))))', []],
      ['B', 15, 15, 'lit', '=ROUND(YEAR(B13)-YEAR(B14)+(MONTH(B13)-MONTH(B14))/12+(DAY(B13)-DAY(B14)-1)/360,0)', []],
      ['B', 16, 16, 'lit', '=B15-B7', []],
      ['B', 17, 17, 'lit', '=B16+1', []],
      ['B', 18, 18, 'lit', '=B9*1000', []],
      ['B', 23, 23, 'lit', '=B22*B25', []],
      ['B', 25, 25, 'lit', '=_xlfn.SWITCH(B24,"A",1,"S",0.52,"Q",0.265,"M",0.088,1)', []],
      ['G', 8, 8, 'lit', '=B23/12', []],
      ['G', 9, 9, 'lit', '=IF(B20="Vanish Premium",IFERROR(MATCH("Yes",Projection_Monthly!AH:AH,0)-1,"Not Yet"),"N/A")', []],
      ['G', 10, 10, 'lit', '=IF(B20="Vanish Premium",IF(ISNUMBER(G9),INDEX(Projection_Monthly!B:B,G9+1),"N/A"),"N/A")', []],
      ['G', 11, 11, 'lit', '=IF(B20="Vanish Premium",IF(ISNUMBER(G9),INDEX(Projection_Monthly!D:D,G9+1),"N/A"),"N/A")', []],
      ['G', 12, 12, 'lit', '=IF(B20="Vanish Premium",INDEX(Projection_Monthly!AF:AF,MATCH(B12,Projection_Monthly!B:B,1)),"N/A")', []],
      ['G', 13, 13, 'lit', '=IF(B20="Vanish Premium",INDEX(Projection_Monthly!AG:AG,MATCH(B12,Projection_Monthly!B:B,1)),"N/A")', []],
    ],
    "colfmt": {},
    "fmt_exc": {'B19': '0.00%', 'B29': '0.00%', 'B11': 'mm-dd-yy', 'A43': 'mm-dd-yy', 'B23': '0', 'B13': 'mm-dd-yy', 'B44': '#,##0.00', 'G12': '"$"#,##0.00', 'B21': '0.00%', 'A34': 'mm-dd-yy', 'A35': 'mm-dd-yy', 'B24': '0', 'A44': 'mm-dd-yy', 'G13': '"$"#,##0.00', 'G8': '"$"#,##0.00', 'B12': 'mm-dd-yy', 'G10': 'mm/dd/yyyy', 'A33': 'mm-dd-yy', 'B25': '0', 'B22': '0', 'A36': 'mm-dd-yy', 'B14': 'mm-dd-yy', 'B41': 'mm-dd-yy', 'B40': '0%'},
    "widths": {'A': 30.82, 'B': 17.54, 'C': 27.63, 'D': 32.27, 'E': 13.18, 'F': 28.0, 'G': 17.82, 'H': 8.82, 'I': 17.82},
  },
  'MORTALITY': {
    "statics": {
        'A1': 'Age',
        'C1': 'v*qx',
        'D1': 'v*q(x+1)*px',
        'A2': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 3, 123, 'trans', '=IFERROR(IF(B2>=1,"",A2+1),"")', []],
      ['B', 1, 1, 'lit', '=_xlfn.CONCAT(Inputs!B2,Inputs!B6,Inputs!B3)', []],
      ['B', 2, 123, 'trans', '=IFERROR(INDEX(\'CSO Rates\'!$A$5:$R$126,MATCH(A2,\'CSO Rates\'!$A$5:$A$126,0),MATCH($B$1,\'CSO Rates\'!$A$4:$R$4,0)),"")', []],
      ['C', 2, 123, 'trans', '=IFERROR(ETI!$B$7*B2,"")', []],
      ['D', 2, 123, 'trans', '=IFERROR(IF(A2="","",C2*(1-B3)),0)', []],
    ],
    "colfmt": {'D': '0.00000_)', 'B': '0.00000_)', 'C': '0.00000_)'},
    "fmt_exc": {'D1': 'General', 'C1': 'General', 'B1': 'General'},
    "widths": {'B': 11.09, 'C': 9.0, 'D': 12.63},
  },
  'Commutation': {
    "statics": {
        'A1': 'Age',
        'B1': 'qx',
        'C1': 'lx',
        'D1': 'dx',
        'E1': 'v^x',
        'F1': 'Dx',
        'G1': 'Cx',
        'H1': 'Nx',
        'I1': 'Mx',
        'J1': 'ax_due',
        'K1': 'Ax',
        'L1': 'NLP_per_1000',
        'A2': 0,
        'C2': 10000000,
    },
    "arrays": [],
    "fills": [
      ['A', 3, 102, 'trans', '=IF(OR(A2="",A2>=MAX(MORTALITY!$A:$A)),"",A2+1)', []],
      ['B', 2, 102, 'trans', '=IF(A2="","",IFERROR(INDEX(MORTALITY!$B:$B,MATCH(A2,MORTALITY!$A:$A,0)),0))', []],
      ['C', 3, 102, 'trans', '=IF(A3="","",IFERROR(C2*(1-B2),0))', []],
      ['D', 2, 102, 'trans', '=IF(A2="","",IFERROR(C2*B2,0))', []],
      ['E', 2, 102, 'trans', '=IF(A2="","",IFERROR((1/(1+Inputs!$B$21))^A2,0))', []],
      ['F', 2, 102, 'trans', '=IF(A2="","",IFERROR(C2*E2,0))', []],
      ['G', 2, 102, 'trans', '=IF(A2="","",IFERROR(D2*(1/(1+Inputs!$B$21))^(A2+1),0))', []],
      ['H', 2, 102, 'trans', '=IF(A2="","",IFERROR(IF(B2=1,F2,F2+H3),0))', []],
      ['I', 2, 102, 'trans', '=IF(A2="","",IFERROR(IF(B2=1,G2,G2+I3),0))', []],
      ['J', 2, 102, 'trans', '=IF(A2="","",IFERROR(H2/F2,0))', []],
      ['K', 2, 102, 'trans', '=IF(A2="","",IFERROR(I2/F2,0))', []],
      ['L', 2, 102, 'trans', '=IF(A2="","",IFERROR(1000*I2/H2,0))', []],
    ],
    "colfmt": {'B': '0.00000_)'},
    "fmt_exc": {'B102': 'General', 'B1': 'General'},
    "widths": {'A': 6.63, 'B': 10.73, 'C': 11.82, 'L': 12.73},
  },
  'CV_RATES': {
    "statics": {
        'A1': 'Code',
        'B1': 'Issue Age',
        'C1': 'Duration',
        'D1': 'Attained Age (EOY)',
        'E1': 'qx',
        'F1': 'px',
        'G1': 'Interest Rate',
        'H1': 'Beta (NLP/1000)',
        'I1': 'CV (BOY)',
        'J1': 'CV (EOY)',
        'K1': 'CV FACTOR',
        'C2': 0,
        'I2': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 2, 101, 'trans', '=IF(D2>100,"",Inputs!$B$4)', []],
      ['B', 2, 101, 'trans', '=IF(D2>100,"",Inputs!$B$7)', []],
      ['C', 3, 101, 'trans', '=IFERROR(IF(Inputs!$B$7+C2+1>=100,"",C2+1),"")', []],
      ['D', 2, 101, 'trans', '=IFERROR(IF(Inputs!$B$7+C2>=100,"",Inputs!$B$7+C2),"")', []],
      ['E', 2, 101, 'trans', '=IF(D2="","",INDEX(MORTALITY!$B:$B,MATCH(D2,MORTALITY!$A:$A,0)))', []],
      ['F', 2, 101, 'trans', '=IF(D2="","",1-E2)', []],
      ['G', 2, 101, 'trans', '=IF(D2="","",Inputs!$B$21)', []],
      ['H', 2, 101, 'trans', '=IF(D2="","",INDEX(Commutation!$L:$L,MATCH(Inputs!$B$7,Commutation!$A:$A,0)))', []],
      ['I', 3, 101, 'trans', '=IF(D3="","",J2)', []],
      ['J', 2, 101, 'trans', '=IF(D2>99,"",IF(Inputs!$B$7+3-1>MAX(MORTALITY!$A:$A),"",IF(E2=1,1000,IFERROR(((I2+H2)*(1+G2)-E2*1000)/F2,0))))', []],
      ['K', 2, 101, 'trans', '=IF(Inputs!$B$7+1-1>MAX(MORTALITY!$A:$A),"",IFERROR(J2,0))', []],
    ],
    "colfmt": {'J': '0.00000', 'G': '0.00%', 'F': '0.00000_)', 'D': '0', 'I': '0.00000', 'E': '0.00000_)', 'K': '0.00000'},
    "fmt_exc": {'I1': 'General', 'K41': '0', 'D1': '0.00', 'E1': 'General', 'F1': 'General', 'G1': 'General', 'K1': 'General', 'J1': 'General', 'J41': '0'},
    "widths": {'A': 10.45, 'C': 13.45, 'D': 12.45, 'I': 12.36, 'J': 13.45},
  },
  'PUA_Div': {
    "statics": {
        'A2': 'PUA Units',
        'A3': 'PUA Face Amount',
        'A4': 'Date of Birth',
        'A5': 'Valuation Date',
        'A7': 'Div Date (Anniversary Date)',
        'B7': 'Div Amount',
        'C7': 'Attained Age',
        'D7': 'PUA Single Premium',
        'E7': 'PUA Face Purchased',
        'F7': 'Total PUA Face',
        'A13': 'Last Anniverasry',
        'E13': 'PUA Cash Value',
    },
    "arrays": [],
    "fills": [
      ['A', 8, 11, 'trans', '=Inputs!A33', []],
      ['A', 14, 14, 'lit', '=Inputs!$B$13', []],
      ['B', 2, 2, 'lit', '=Inputs!$B$10', []],
      ['B', 3, 3, 'lit', '=B2*1000', []],
      ['B', 4, 4, 'lit', '=Inputs!$B$14', []],
      ['B', 5, 5, 'lit', '=Inputs!$B$12', []],
      ['B', 8, 11, 'trans', '=Inputs!B33', []],
      ['C', 8, 11, 'trans', '=ROUND(YEAR(A8)-YEAR($B$4)+(MONTH(A8)-MONTH($B$4))/12+(DAY(A8)-DAY($B$4)-1)/360,0)', []],
      ['C', 14, 14, 'lit', '=ROUND(YEAR(A14)-YEAR($B$4)+(MONTH(A14)-MONTH($B$4))/12+(DAY(A14)-DAY($B$4)-1)/360,0)', []],
      ['D', 8, 11, 'trans', '=IF(C8="","",INDEX(Commutation!$K:$K,MATCH(C8,Commutation!$A:$A,0)))', []],
      ['D', 14, 14, 'lit', '=IF(C14="","",INDEX(Commutation!$K:$K,MATCH(C14,Commutation!$A:$A,0)))', []],
      ['E', 8, 11, 'trans', '=B8/D8', []],
      ['E', 14, 14, 'lit', '=F10*D14', []],
      ['F', 8, 11, 'trans', '=$B$3+SUM($E$8:E8)', []],
    ],
    "colfmt": {'F': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)', 'E': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)', 'B': '0.00'},
    "fmt_exc": {'A8': 'mm-dd-yy', 'F7': 'General', 'A11': 'mm-dd-yy', 'E7': 'General', 'B2': 'General', 'A9': 'mm-dd-yy', 'B4': 'mm-dd-yy', 'A10': 'mm-dd-yy', 'A14': 'mm-dd-yy', 'E14': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)', 'B5': 'mm-dd-yy', 'E13': 'General', 'B7': 'General', 'B3': 'General'},
    "widths": {'A': 24.18, 'B': 10.54, 'C': 11.45, 'D': 17.73, 'F': 13.18},
  },
  'DIvidend_Assumption': {
    "statics": {
        'B1': 'Age',
        'A2': 'Duration',
        'B2': 0,
        'A3': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 4, 103, 'trans', '=A3+1', []],
      ['B', 3, 3, 'lit', '=Inputs!$B$29', []],
      ['B', 4, 103, 'trans', '=B3', []],
      ['C', 2, 2, 'lit', '=+B2+1', []],
      ['C', 3, 103, 'trans', '=B4', []],
      ['D', 2, 2, 'lit', '=+C2+1', []],
      ['D', 3, 103, 'trans', '=C4', []],
      ['E', 2, 2, 'lit', '=+D2+1', []],
      ['E', 3, 103, 'trans', '=D4', []],
      ['F', 2, 2, 'lit', '=+E2+1', []],
      ['F', 3, 103, 'trans', '=E4', []],
      ['G', 2, 2, 'lit', '=+F2+1', []],
      ['G', 3, 103, 'trans', '=F4', []],
      ['H', 2, 2, 'lit', '=+G2+1', []],
      ['H', 3, 103, 'trans', '=G4', []],
      ['I', 2, 2, 'lit', '=+H2+1', []],
      ['I', 3, 103, 'trans', '=H4', []],
      ['J', 2, 2, 'lit', '=+I2+1', []],
      ['J', 3, 103, 'trans', '=I4', []],
      ['K', 2, 2, 'lit', '=+J2+1', []],
      ['K', 3, 103, 'trans', '=J4', []],
      ['L', 2, 2, 'lit', '=+K2+1', []],
      ['L', 3, 103, 'trans', '=K4', []],
      ['M', 2, 2, 'lit', '=+L2+1', []],
      ['M', 3, 103, 'trans', '=L4', []],
      ['N', 2, 2, 'lit', '=+M2+1', []],
      ['N', 3, 103, 'trans', '=M4', []],
      ['O', 2, 2, 'lit', '=+N2+1', []],
      ['O', 3, 103, 'trans', '=N4', []],
      ['P', 2, 2, 'lit', '=+O2+1', []],
      ['P', 3, 103, 'trans', '=O4', []],
      ['Q', 2, 2, 'lit', '=+P2+1', []],
      ['Q', 3, 103, 'trans', '=P4', []],
      ['R', 2, 2, 'lit', '=+Q2+1', []],
      ['R', 3, 103, 'trans', '=Q4', []],
      ['S', 2, 2, 'lit', '=+R2+1', []],
      ['S', 3, 103, 'trans', '=R4', []],
      ['T', 2, 2, 'lit', '=+S2+1', []],
      ['T', 3, 103, 'trans', '=S4', []],
      ['U', 2, 2, 'lit', '=+T2+1', []],
      ['U', 3, 103, 'trans', '=T4', []],
      ['V', 2, 2, 'lit', '=+U2+1', []],
      ['V', 3, 103, 'trans', '=U4', []],
      ['W', 2, 2, 'lit', '=+V2+1', []],
      ['W', 3, 103, 'trans', '=V4', []],
      ['X', 2, 2, 'lit', '=+W2+1', []],
      ['X', 3, 103, 'trans', '=W4', []],
      ['Y', 2, 2, 'lit', '=+X2+1', []],
      ['Y', 3, 103, 'trans', '=X4', []],
      ['Z', 2, 2, 'lit', '=+Y2+1', []],
      ['Z', 3, 103, 'trans', '=Y4', []],
      ['AA', 2, 2, 'lit', '=+Z2+1', []],
      ['AA', 3, 103, 'trans', '=Z4', []],
      ['AB', 2, 2, 'lit', '=+AA2+1', []],
      ['AB', 3, 103, 'trans', '=AA4', []],
      ['AC', 2, 2, 'lit', '=+AB2+1', []],
      ['AC', 3, 103, 'trans', '=AB4', []],
      ['AD', 2, 2, 'lit', '=+AC2+1', []],
      ['AD', 3, 103, 'trans', '=AC4', []],
      ['AE', 2, 2, 'lit', '=+AD2+1', []],
      ['AE', 3, 103, 'trans', '=AD4', []],
      ['AF', 2, 2, 'lit', '=+AE2+1', []],
      ['AF', 3, 103, 'trans', '=AE4', []],
      ['AG', 2, 2, 'lit', '=+AF2+1', []],
      ['AG', 3, 103, 'trans', '=AF4', []],
      ['AH', 2, 2, 'lit', '=+AG2+1', []],
      ['AH', 3, 103, 'trans', '=AG4', []],
      ['AI', 2, 2, 'lit', '=+AH2+1', []],
      ['AI', 3, 103, 'trans', '=AH4', []],
      ['AJ', 2, 2, 'lit', '=+AI2+1', []],
      ['AJ', 3, 103, 'trans', '=AI4', []],
      ['AK', 2, 2, 'lit', '=+AJ2+1', []],
      ['AK', 3, 103, 'trans', '=AJ4', []],
      ['AL', 2, 2, 'lit', '=+AK2+1', []],
      ['AL', 3, 103, 'trans', '=AK4', []],
      ['AM', 2, 2, 'lit', '=+AL2+1', []],
      ['AM', 3, 103, 'trans', '=AL4', []],
      ['AN', 2, 2, 'lit', '=+AM2+1', []],
      ['AN', 3, 103, 'trans', '=AM4', []],
      ['AO', 2, 2, 'lit', '=+AN2+1', []],
      ['AO', 3, 103, 'trans', '=AN4', []],
      ['AP', 2, 2, 'lit', '=+AO2+1', []],
      ['AP', 3, 103, 'trans', '=AO4', []],
      ['AQ', 2, 2, 'lit', '=+AP2+1', []],
      ['AQ', 3, 103, 'trans', '=AP4', []],
      ['AR', 2, 2, 'lit', '=+AQ2+1', []],
      ['AR', 3, 103, 'trans', '=AQ4', []],
      ['AS', 2, 2, 'lit', '=+AR2+1', []],
      ['AS', 3, 103, 'trans', '=AR4', []],
      ['AT', 2, 2, 'lit', '=+AS2+1', []],
      ['AT', 3, 103, 'trans', '=AS4', []],
      ['AU', 2, 2, 'lit', '=+AT2+1', []],
      ['AU', 3, 103, 'trans', '=AT4', []],
      ['AV', 2, 2, 'lit', '=+AU2+1', []],
      ['AV', 3, 103, 'trans', '=AU4', []],
      ['AW', 2, 2, 'lit', '=+AV2+1', []],
      ['AW', 3, 103, 'trans', '=AV4', []],
      ['AX', 2, 2, 'lit', '=+AW2+1', []],
      ['AX', 3, 103, 'trans', '=AW4', []],
      ['AY', 2, 2, 'lit', '=+AX2+1', []],
      ['AY', 3, 103, 'trans', '=AX4', []],
      ['AZ', 2, 2, 'lit', '=+AY2+1', []],
      ['AZ', 3, 103, 'trans', '=AY4', []],
      ['BA', 2, 2, 'lit', '=+AZ2+1', []],
      ['BA', 3, 103, 'trans', '=AZ4', []],
      ['BB', 2, 2, 'lit', '=+BA2+1', []],
      ['BB', 3, 103, 'trans', '=BA4', []],
      ['BC', 2, 2, 'lit', '=+BB2+1', []],
      ['BC', 3, 103, 'trans', '=BB4', []],
      ['BD', 2, 2, 'lit', '=+BC2+1', []],
      ['BD', 3, 103, 'trans', '=BC4', []],
      ['BE', 2, 2, 'lit', '=+BD2+1', []],
      ['BE', 3, 103, 'trans', '=BD4', []],
      ['BF', 2, 2, 'lit', '=+BE2+1', []],
      ['BF', 3, 103, 'trans', '=BE4', []],
      ['BG', 2, 2, 'lit', '=+BF2+1', []],
      ['BG', 3, 103, 'trans', '=BF4', []],
      ['BH', 2, 2, 'lit', '=+BG2+1', []],
      ['BH', 3, 103, 'trans', '=BG4', []],
      ['BI', 2, 2, 'lit', '=+BH2+1', []],
      ['BI', 3, 103, 'trans', '=BH4', []],
      ['BJ', 2, 2, 'lit', '=+BI2+1', []],
      ['BJ', 3, 103, 'trans', '=BI4', []],
      ['BK', 2, 2, 'lit', '=+BJ2+1', []],
      ['BK', 3, 103, 'trans', '=BJ4', []],
      ['BL', 2, 2, 'lit', '=+BK2+1', []],
      ['BL', 3, 103, 'trans', '=BK4', []],
      ['BM', 2, 2, 'lit', '=+BL2+1', []],
      ['BM', 3, 103, 'trans', '=BL4', []],
      ['BN', 2, 2, 'lit', '=+BM2+1', []],
      ['BN', 3, 103, 'trans', '=BM4', []],
      ['BO', 2, 2, 'lit', '=+BN2+1', []],
      ['BO', 3, 103, 'trans', '=BN4', []],
      ['BP', 2, 2, 'lit', '=+BO2+1', []],
      ['BP', 3, 103, 'trans', '=BO4', []],
      ['BQ', 2, 2, 'lit', '=+BP2+1', []],
      ['BQ', 3, 103, 'trans', '=BP4', []],
      ['BR', 2, 2, 'lit', '=+BQ2+1', []],
      ['BR', 3, 103, 'trans', '=BQ4', []],
      ['BS', 2, 2, 'lit', '=+BR2+1', []],
      ['BS', 3, 103, 'trans', '=BR4', []],
      ['BT', 2, 2, 'lit', '=+BS2+1', []],
      ['BT', 3, 103, 'trans', '=BS4', []],
      ['BU', 2, 2, 'lit', '=+BT2+1', []],
      ['BU', 3, 103, 'trans', '=BT4', []],
      ['BV', 2, 2, 'lit', '=+BU2+1', []],
      ['BV', 3, 103, 'trans', '=BU4', []],
      ['BW', 2, 2, 'lit', '=+BV2+1', []],
      ['BW', 3, 103, 'trans', '=BV4', []],
      ['BX', 2, 2, 'lit', '=+BW2+1', []],
      ['BX', 3, 103, 'trans', '=BW4', []],
      ['BY', 2, 2, 'lit', '=+BX2+1', []],
      ['BY', 3, 103, 'trans', '=BX4', []],
      ['BZ', 2, 2, 'lit', '=+BY2+1', []],
      ['BZ', 3, 103, 'trans', '=BY4', []],
      ['CA', 2, 2, 'lit', '=+BZ2+1', []],
      ['CA', 3, 103, 'trans', '=BZ4', []],
      ['CB', 2, 2, 'lit', '=+CA2+1', []],
      ['CB', 3, 103, 'trans', '=CA4', []],
      ['CC', 2, 2, 'lit', '=+CB2+1', []],
      ['CC', 3, 103, 'trans', '=CB4', []],
      ['CD', 2, 2, 'lit', '=+CC2+1', []],
      ['CD', 3, 103, 'trans', '=CC4', []],
      ['CE', 2, 2, 'lit', '=+CD2+1', []],
      ['CE', 3, 103, 'trans', '=CD4', []],
      ['CF', 2, 2, 'lit', '=+CE2+1', []],
      ['CF', 3, 103, 'trans', '=CE4', []],
      ['CG', 2, 2, 'lit', '=+CF2+1', []],
      ['CG', 3, 103, 'trans', '=CF4', []],
      ['CH', 2, 2, 'lit', '=+CG2+1', []],
      ['CH', 3, 103, 'trans', '=CG4', []],
      ['CI', 2, 2, 'lit', '=+CH2+1', []],
      ['CI', 3, 103, 'trans', '=CH4', []],
      ['CJ', 2, 2, 'lit', '=+CI2+1', []],
      ['CJ', 3, 103, 'trans', '=CI4', []],
      ['CK', 2, 2, 'lit', '=+CJ2+1', []],
      ['CK', 3, 103, 'trans', '=CJ4', []],
      ['CL', 2, 2, 'lit', '=+CK2+1', []],
      ['CL', 3, 103, 'trans', '=CK4', []],
      ['CM', 2, 2, 'lit', '=+CL2+1', []],
      ['CM', 3, 103, 'trans', '=CL4', []],
      ['CN', 2, 2, 'lit', '=+CM2+1', []],
      ['CN', 3, 103, 'trans', '=CM4', []],
      ['CO', 2, 2, 'lit', '=+CN2+1', []],
      ['CO', 3, 103, 'trans', '=CN4', []],
      ['CP', 2, 2, 'lit', '=+CO2+1', []],
      ['CP', 3, 103, 'trans', '=CO4', []],
      ['CQ', 2, 2, 'lit', '=+CP2+1', []],
      ['CQ', 3, 103, 'trans', '=CP4', []],
      ['CR', 2, 2, 'lit', '=+CQ2+1', []],
      ['CR', 3, 103, 'trans', '=CQ4', []],
      ['CS', 2, 2, 'lit', '=+CR2+1', []],
      ['CS', 3, 103, 'trans', '=CR4', []],
      ['CT', 2, 2, 'lit', '=+CS2+1', []],
      ['CT', 3, 103, 'trans', '=CS4', []],
      ['CU', 2, 2, 'lit', '=+CT2+1', []],
      ['CU', 3, 103, 'trans', '=CT4', []],
      ['CV', 2, 2, 'lit', '=+CU2+1', []],
      ['CV', 3, 103, 'trans', '=CU4', []],
      ['CW', 2, 2, 'lit', '=+CV2+1', []],
      ['CW', 3, 103, 'trans', '=CV4', []],
      ['CX', 2, 2, 'lit', '=+CW2+1', []],
      ['CX', 3, 103, 'trans', '=CW4', []],
    ],
    "colfmt": {'AZ': '0.0%', 'BD': '0.0%', 'AB': '0.0%', 'BH': '0.0%', 'CS': '0.0%', 'AO': '0.0%', 'BA': '0.0%', 'Z': '0.0%', 'AP': '0.0%', 'X': '0.0%', 'BB': '0.0%', 'AK': '0.0%', 'CG': '0.0%', 'D': '0.0%', 'CQ': '0.0%', 'BF': '0.0%', 'CT': '0.0%', 'R': '0.0%', 'AG': '0.0%', 'AH': '0.0%', 'BW': '0.0%', 'BK': '0.0%', 'AM': '0.0%', 'BL': '0.0%', 'AX': '0.0%', 'CE': '0.0%', 'AL': '0.0%', 'U': '0.0%', 'L': '0.0%', 'AU': '0.0%', 'Y': '0.0%', 'BV': '0.0%', 'BM': '0.0%', 'I': '0.0%', 'O': '0.0%', 'F': '0.0%', 'CW': '0.0%', 'AN': '0.0%', 'K': '0.0%', 'BJ': '0.0%', 'BQ': '0.0%', 'C': '0.0%', 'BP': '0.0%', 'J': '0.0%', 'CA': '0.0%', 'Q': '0.0%', 'CN': '0.0%', 'CV': '0.0%', 'BX': '0.0%', 'AF': '0.0%', 'CI': '0.0%', 'CP': '0.0%', 'BO': '0.0%', 'T': '0.0%', 'AY': '0.0%', 'CR': '0.0%', 'P': '0.0%', 'CX': '0.0%', 'CO': '0.0%', 'V': '0.0%', 'B': '0.0%', 'CM': '0.0%', 'AR': '0.0%', 'BG': '0.0%', 'BU': '0.0%', 'AC': '0.0%', 'AQ': '0.0%', 'AV': '0.0%', 'BR': '0.0%', 'AJ': '0.0%', 'E': '0.0%', 'BN': '0.0%', 'AS': '0.0%', 'S': '0.0%', 'CD': '0.0%', 'H': '0.0%', 'BT': '0.0%', 'CL': '0.0%', 'CJ': '0.0%', 'BY': '0.0%', 'CH': '0.0%', 'CK': '0.0%', 'BC': '0.0%', 'AD': '0.0%', 'CC': '0.0%', 'AT': '0.0%', 'CF': '0.0%', 'AI': '0.0%', 'AW': '0.0%', 'W': '0.0%', 'CB': '0.0%', 'BS': '0.0%', 'BZ': '0.0%', 'AE': '0.0%', 'AA': '0.0%', 'BE': '0.0%', 'G': '0.0%', 'BI': '0.0%', 'M': '0.0%', 'N': '0.0%', 'CU': '0.0%'},
    "fmt_exc": {'U2': 'General', 'BF2': 'General', 'CF2': 'General', 'CK2': 'General', 'K2': 'General', 'CU2': 'General', 'CD2': 'General', 'CT2': 'General', 'AQ2': 'General', 'BU2': 'General', 'BN2': 'General', 'BV2': 'General', 'BM2': 'General', 'AF2': 'General', 'AD2': 'General', 'AC2': 'General', 'BX2': 'General', 'AZ2': 'General', 'E2': 'General', 'BI2': 'General', 'CP2': 'General', 'AN2': 'General', 'CN2': 'General', 'AO2': 'General', 'BT2': 'General', 'P2': 'General', 'CX2': 'General', 'AW2': 'General', 'D2': 'General', 'CE2': 'General', 'BA2': 'General', 'BQ2': 'General', 'AJ2': 'General', 'BO2': 'General', 'BE2': 'General', 'AH2': 'General', 'CA2': 'General', 'CM2': 'General', 'AP2': 'General', 'BR2': 'General', 'C2': 'General', 'BB2': 'General', 'R2': 'General', 'X2': 'General', 'CR2': 'General', 'CS2': 'General', 'CJ2': 'General', 'Q2': 'General', 'AX2': 'General', 'B2': 'General', 'CO2': 'General', 'AA2': 'General', 'AY2': 'General', 'V2': 'General', 'BZ2': 'General', 'AS2': 'General', 'BH2': 'General', 'H2': 'General', 'CB2': 'General', 'O2': 'General', 'G2': 'General', 'BW2': 'General', 'BS2': 'General', 'CG2': 'General', 'M2': 'General', 'S2': 'General', 'BD2': 'General', 'CI2': 'General', 'AG2': 'General', 'I2': 'General', 'AT2': 'General', 'N2': 'General', 'BL2': 'General', 'BC2': 'General', 'AR2': 'General', 'AU2': 'General', 'CV2': 'General', 'AK2': 'General', 'CL2': 'General', 'BY2': 'General', 'AB2': 'General', 'Z2': 'General', 'J2': 'General', 'BJ2': 'General', 'CQ2': 'General', 'B1': 'General', 'AV2': 'General', 'AL2': 'General', 'BP2': 'General', 'F2': 'General', 'CH2': 'General', 'AM2': 'General', 'CC2': 'General', 'L2': 'General', 'Y2': 'General', 'CW2': 'General', 'T2': 'General', 'BK2': 'General', 'BG2': 'General', 'W2': 'General', 'AI2': 'General', 'AE2': 'General'},
    "widths": {},
  },
  'Projection_Monthly': {
    "statics": {
        'A1': 'Plan',
        'AA1': 'CV from PUA',
        'AB1': 'CV Total (Incl PUA)',
        'AC1': 'Premium Offset Applied (Div + Accum Div growth)',
        'AD1': 'VP: Div Applied to Premium',
        'AE1': 'VP: Accum Div Used',
        'AF1': 'VP: Net Accum Div Balance',
        'AG1': 'VP: Out-of-Pocket Premium',
        'AH1': 'VP: Premium Vanished?',
        'B1': 'Date',
        'C1': 'Month Index',
        'D1': 'Duration',
        'E1': 'Attained Age',
        'F1': 'Last Anniv',
        'G1': 'Next Anniv',
        'H1': 'Frac within Year',
        'I1': 'CV Factor_n',
        'J1': 'CV Factor_n+1',
        'K1': 'CV per 1000',
        'L1': 'CV Total (Base)',
        'M1': 'PUA CV',
        'N1': 'Pre-Dividend PUA SA (cum)',
        'O1': 'Dividend Base SA',
        'P1': 'Monthly Div Rate',
        'Q1': 'Dividend Amount',
        'R1': 'Accum Div ',
        'S1': 'Accum Div Prev End',
        'T1': 'Accum Div Prev Growth',
        'U1': 'Out-of-Pocket Premium',
        'V1': 'Cash Dividend Paid',
        'W1': 'Dividend to PUA',
        'X1': 'PUA SP per $1 SA',
        'Y1': 'PUA SA Add (this month)',
        'Z1': 'PUA SA (cum)',
    },
    "arrays": [],
    "fills": [
      ['A', 2, 2, 'lit', '=IF(Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4)', []],
      ['A', 3, 985, 'trans', '=IF(A2="","",IF(INT((C2+1)/12)+Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4))', []],
      ['B', 2, 2, 'lit', '=IF(A2="","",Inputs!$B$11)', []],
      ['B', 3, 985, 'trans', '=IF(A3="","",EOMONTH(B2,1))', []],
      ['C', 2, 2, 'lit', '=IF(A2="","",0)', []],
      ['C', 3, 985, 'trans', '=IF(A3="","",C2+1)', []],
      ['D', 2, 985, 'trans', '=IF(A2="","",INT(C2/12))', []],
      ['E', 2, 985, 'trans', '=IF(A2="","",D2+Inputs!$B$7)', []],
      ['F', 2, 985, 'trans', '=IF(A2="","",DATE(YEAR(B$2)+D2,MONTH(B$2),DAY(B$2)))', []],
      ['G', 2, 985, 'trans', '=IF(A2="","",DATE(YEAR(F2)+1,MONTH(F2),DAY(F2)))', []],
      ['H', 2, 985, 'trans', '=IF(A2="","",IF(G2>F2,(B2-F2)/(G2-F2),0))', []],
      ['I', 2, 985, 'trans', '=IF(A2="","",IFERROR(INDEX(CV_RATES!$I:$I,MATCH(D2,CV_RATES!$C:$C,0)),0))', []],
      ['J', 2, 985, 'trans', '=IF(A2="","",IFERROR(IF(E2=MAX(MORTALITY!$A:$A),1000,INDEX(CV_RATES!$J:$J,MATCH(D2,CV_RATES!$C:$C,0))),0))', []],
      ['K', 2, 985, 'trans', '=IF(A2="","",IFERROR(I2+H2*(J2-I2),0))', []],
      ['L', 2, 985, 'trans', '=IF(A2="","",IFERROR(K2*Inputs!$B$9,0))', []],
      ['M', 2, 985, 'trans', '=IF(A2="","",IFERROR(K2*Inputs!$B$10,0))', []],
      ['N', 2, 2, 'lit', '=IF(A2="","",0)', []],
      ['N', 3, 985, 'trans', '=IF(A3="","",Z2)', []],
      ['O', 2, 985, 'trans', '=IF(A2="","",Inputs!$B$18+N2)', []],
      ['P', 2, 985, 'trans', '=IF(A2="","",VLOOKUP(D2,DIvidend_Assumption!$A$3:$CX$103,($E$2+2),0)/12)', []],
      ['Q', 2, 2, 'lit', '=IF(A2="","",0)', []],
      ['Q', 3, 985, 'trans', '=IF(A3="","",IFERROR(AB2*P3,""))', []],
      ['R', 2, 985, 'trans', '=IF(A2="","",IFERROR(IF(Inputs!$B$20="Accumulate",T2+Q2,IF(Inputs!$B$20="Premium Offset",MAX(0,T2+Q2-Inputs!$B$23),T2)),""))', []],
      ['S', 2, 2, 'lit', '=IF(A2="","",0)', []],
      ['S', 3, 985, 'trans', '=IF(A3="","",R2)', []],
      ['T', 2, 985, 'trans', '=IF(A2="","",IFERROR(S2*(1+Inputs!$B$21/12),""))', []],
      ['U', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Premium Offset",MAX(0,Inputs!$B$23-(Q2+T2)),IF(Inputs!$B$20="Vanish Premium",AG2*12,Inputs!$B$23)))', []],
      ['V', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Cash",Q2,0))', []],
      ['W', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="PUA",Q2,0))', []],
      ['X', 2, 985, 'trans', '=IF(A2="","",IFERROR(INDEX(#REF!,MATCH(E2,#REF!,1)),0))', []],
      ['Y', 2, 985, 'trans', '=IF(A2="","",IF(W2>0,IF(X2>0,W2/X2,0),0))', []],
      ['Z', 2, 985, 'trans', '=IF(A2="","",N2+Y2)', []],
      ['AA', 2, 2, 'lit', '=IF(A2="","",IFERROR(W2,0))', []],
      ['AA', 3, 985, 'trans', '=IF(A3="","",IFERROR(Z3*INDEX(#REF!,MATCH(E3,#REF!,1))/1000,0))', []],
      ['AB', 2, 985, 'trans', '=IF(A2="","",IFERROR(L2+AA2,0))', []],
      ['AC', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Premium Offset",MIN(Inputs!$B$23,Q2+T2),0))', []],
      ['AD', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Vanish Premium",MIN(Q2,Inputs!$B$23/12),0))', []],
      ['AE', 2, 2, 'lit', '=IF(A2="","",IF(Inputs!$B$20="Vanish Premium",MIN(MAX(0,Inputs!$B$23/12-AD2),0),0))', []],
      ['AE', 3, 985, 'trans', '=IF(A3="","",IF(Inputs!$B$20="Vanish Premium",MIN(MAX(0,Inputs!$B$23/12-AD3),AF2),0))', []],
      ['AF', 2, 2, 'lit', '=IF(A2="","",IF(Inputs!$B$20="Vanish Premium",MAX(0,Q2-AD2),R2))', []],
      ['AF', 3, 985, 'trans', '=IF(A3="","",IF(Inputs!$B$20="Vanish Premium",MAX(0,AF2+Q3-AD3-AE3)*(1+Inputs!$B$21/12),R3))', []],
      ['AG', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Vanish Premium",MAX(0,Inputs!$B$23/12-AD2-AE2),Inputs!$B$23/12))', []],
      ['AH', 2, 985, 'trans', '=IF(A2="","",IF(Inputs!$B$20="Vanish Premium",IF(AG2=0,"Yes","No"),"N/A"))', []],
    ],
    "colfmt": {'K': '0.00', 'R': '0.00', 'O': '0', 'W': '0.00', 'Z': '0.00', 'P': '0.00%', 'AA': '0.00', 'AC': '0.00', 'U': '0', 'M': '0.00', 'V': '0.00', 'J': '0.00', 'L': '0.00', 'AB': '0.00', 'S': '0.00', 'I': '0.00', 'N': '0.00', 'X': '0.000', 'B': 'mm-dd-yy', 'Q': '0.00', 'F': 'mm-dd-yy', 'T': '0.00', 'Y': '0.00', 'H': '0.00', 'G': 'mm-dd-yy'},
    "fmt_exc": {'H1': 'General', 'AB1': 'General', 'AA1': 'General', 'I1': 'General', 'B1': 'General', 'M1': 'General', 'O1': 'General', 'AC1': 'General', 'W1': 'General', 'R1': 'General', 'V1': 'General', 'J1': 'General', 'Q1': 'General', 'X1': 'General', 'T1': 'General', 'S1': 'General', 'N1': 'General', 'Z1': 'General', 'Y1': 'General', 'U1': 'General', 'L1': 'General', 'K1': 'General'},
    "widths": {'B': 10.45, 'C': 13.82, 'D': 10.45, 'E': 13.73, 'F': 11.54, 'G': 12.18, 'H': 16.45, 'I': 13.45, 'J': 15.0, 'K': 13.0, 'L': 15.54, 'N': 25.82, 'O': 17.45, 'P': 17.45, 'Q': 17.73, 'R': 12.18, 'S': 19.54, 'T': 22.73, 'U': 22.82, 'V': 19.0, 'W': 16.54, 'X': 17.27, 'Y': 23.82, 'Z': 14.45, 'AA': 13.73, 'AB': 18.54, 'AC': 45.0, 'AD': 21.82},
  },
  'RPU': {
    "statics": {
        'A1': 'Reduced Paid-Up (RPU) - Actuarial Calculation',
        'A3': 'Status',
        'B3': 'Applicable',
        'A4': 'Attained Age',
        'A5': 'Original Sum Assured',
        'A7': '═══ Method 1: Cash Value / Net Single Premium ═══',
        'A8': 'Gross Cash Value (tVx)',
        'A9': 'Net Single Premium at Attained Age (Ax)',
        'A10': 'RPU SA (Uncapped) = CV / Ax',
        'A12': '═══ Method 2: Reserve Ratio (Actuarial Formula) ═══',
        'A13': 'Reserve per $1000 at Duration t',
        'A14': 'RPU SA = SA × (tVx/1000) / Ax',
        'A16': '═══════════════════════════════════════════════',
        'A17': 'FINAL RPU Sum Assured',
        'A19': 'Comparison:',
        'A20': 'Is Capping Applied?',
        'A22': 'Note: When CV/Ax > SA, capping prevents over-insurance.',
        'A23': 'Both methods should yield the same result when using',
        'A24': 'consistent actuarial assumptions.',
    },
    "arrays": [],
    "fills": [
      ['B', 4, 4, 'lit', '=Inputs!$B$15', []],
      ['B', 5, 5, 'lit', '=Inputs!$B$18', []],
      ['B', 8, 8, 'lit', '=Output!$B$3', []],
      ['B', 9, 9, 'lit', '=INDEX(Commutation!$K:$K,MATCH(Inputs!$B$15,Commutation!$A:$A,0))', []],
      ['B', 10, 10, 'lit', '=IFERROR(B8/B9,"")', []],
      ['B', 13, 13, 'lit', '=INDEX(CV_RATES!$J:$J,MATCH(Inputs!$B$16,CV_RATES!$C:$C,0))', []],
      ['B', 14, 14, 'lit', '=IFERROR(B5*(B13/1000)/B9,"")', []],
      ['B', 17, 17, 'lit', '=MIN(B10,B5)', []],
      ['B', 20, 20, 'lit', '=IF(Inputs!$B$28=FALSE,"",IF(B10>B5,"YES - Calculated RPU exceeds Original SA","NO - Within limits"))', []],
    ],
    "colfmt": {'B': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)'},
    "fmt_exc": {'B13': '0.00', 'B20': 'General', 'B5': 'General', 'B14': '0.00', 'B3': 'General', 'B4': 'General'},
    "widths": {'A': 70.54, 'B': 22.0},
  },
  'ETI': {
    "statics": {
        'A1': 'Extended Term Insurance (ETI) - Calculations',
        'A3': 'Discount Rate',
        'D3': 'ETI Years (full)',
        'A4': 'Face Amount',
        'D4': 'Partial Year (fraction)',
        'A5': 'Gross Cash Value',
        'D5': 'ETI Duration (years)',
        'A6': 'Attained Age at Val Date',
        'D6': 'ETI Expiry Date',
        'A7': 'v (1/(1+i))',
        'A8': 'Policy Year of ETI',
        'B8': 'Age',
        'C8': 'Term Cost per $1 SA for 1 Year',
        'D8': 'Cash Value required for Base face Amt for 1 Year',
        'E8': 'Remaining CV after cumulative ETI cost',
        'F8': 'Indicator',
    },
    "arrays": ['E4'],
    "fills": [
      ['A', 9, 9, 'lit', '=IF($B$6>MAX(MORTALITY!$A:$A),"",1)', []],
      ['A', 10, 51, 'trans', '=IF(OR(A9="",B9>=MAX(MORTALITY!$A:$A)),"",A9+1)', []],
      ['B', 3, 3, 'lit', '=Inputs!$B$21', []],
      ['B', 4, 4, 'lit', '=Inputs!$B$18', []],
      ['B', 5, 5, 'lit', '=Output!$B$3', []],
      ['B', 6, 6, 'lit', '=Inputs!$B$15', []],
      ['B', 7, 7, 'lit', '=1/(1+B3)', []],
      ['B', 9, 9, 'lit', '=B6', []],
      ['B', 10, 51, 'trans', '=IF(OR(B9="",B9>=MAX(MORTALITY!$A:$A)),"",B9+1)', []],
      ['C', 9, 51, 'trans', '=IF(A9="","",IFERROR(INDEX(MORTALITY!$D:$D,MATCH(B9,MORTALITY!$A:$A,0)),""))', []],
      ['D', 9, 51, 'trans', '=IF(A9="","",IFERROR(C9*$B$4,""))', []],
      ['E', 3, 3, 'lit', '=IFERROR(_xlfn.XLOOKUP(1,$F$9:$F$51,$A$9:$A$51),MAX(A9:A51))', []],
      ['E', 4, 4, 'lit', '=IFERROR(IF(E3="","",IF(INDEX(E9:E51,E3+1)<0,INDEX(E9:E51,E3)/(INDEX(E9:E51,E3)-INDEX(E9:E51,E3+1)),0)),"")', []],
      ['E', 5, 5, 'lit', '=IFERROR(IF(OR(E3="",E4=""),"",E3+E4),"")', []],
      ['E', 6, 6, 'lit', '=IFERROR(IF(E5="","",Inputs!B12+365*E5),"")', []],
      ['E', 9, 9, 'lit', '=IF(A9="","",IFERROR($B$5-SUM($D$9),""))', []],
      ['E', 10, 51, 'trans', '=IF(A10="","",IFERROR($B$5-SUM($D$9:D10),""))', []],
      ['F', 9, 50, 'trans', '=IF(A9="","",IF(AND(E10<0,E9>0),1,0))', []],
      ['F', 51, 51, 'lit', '=IF(A51="","",0)', []],
    ],
    "colfmt": {'D': '0.00', 'C': '0.000000', 'E': '0.00'},
    "fmt_exc": {'D5': 'General', 'D3': 'General', 'B3': '0%', 'D6': 'General', 'C8': 'General', 'D4': 'General', 'D8': 'General', 'E6': 'mm-dd-yy', 'B5': '0', 'E8': 'General', 'E3': 'General', 'B7': '0.0000000', 'E4': '0.00%'},
    "widths": {'A': 23.18, 'B': 20.0, 'C': 17.73, 'D': 23.18, 'E': 19.54, 'F': 12.27},
  },
  'APL': {
    "statics": {
        'A1': 'Automatic Premium Loan (APL)',
        'A3': 'APL Enabled (TRUE/FALSE)',
        'D3': 'Legend',
        'F3': 'Link with Data/Input',
        'A4': 'Policy Loan Interest Rate (p.a.)',
        'F4': 'Formulae',
        'A5': 'Payments per Year (mode)',
        'F5': 'Dropdown Selection',
        'A6': 'Valuation Date',
        'A7': 'Annual Premium',
        'A8': 'Modal Premium',
        'E8': 'Status',
        'F8': 'Applicable',
        'A9': 'Gross Cash Value',
        'E9': 'Policy End Date',
        'A10': 'Loanable % of CV',
        'B10': 0.8,
        'E10': 'APL End Date',
        'A11': 'Max Loan Available at Val Date',
        'A13': 'Period #',
        'B13': 'Due Date (approx)',
        'C13': 'Beginning Balance',
        'D13': 'Interest',
        'E13': 'APL Drawn',
        'F13': 'Payment Covered by APL',
        'G13': 'Ending Balance',
        'H13': 'Remaining CV For Loan Eligibility',
        'I13': 'Indicator',
        'A14': 1,
        'C14': 0,
    },
    "arrays": [],
    "fills": [
      ['A', 15, 199, 'trans', '=IF(OR(A14="",H14<0),"",A14+1)', []],
      ['B', 3, 3, 'lit', '=TRUE()', []],
      ['B', 4, 4, 'lit', '=Inputs!$B$40', []],
      ['B', 5, 5, 'lit', '=Inputs!$B$25', []],
      ['B', 6, 6, 'lit', '=Inputs!$B$12', []],
      ['B', 7, 7, 'lit', '=IFERROR(INDEX(Inputs!B:B,MATCH("Annual Base Premium",Inputs!A:A,0)),0)', []],
      ['B', 8, 8, 'lit', '=IFERROR(INDEX(Inputs!B:B,MATCH("Modal Premium",Inputs!A:A,0)),IF(B7>0,B7/B5,0))', []],
      ['B', 9, 9, 'lit', '=Output!$B$3', []],
      ['B', 11, 11, 'lit', '=B9*B10', []],
      ['B', 14, 14, 'lit', '=B6', []],
      ['B', 15, 199, 'trans', '=IF(A15="","",EOMONTH(B14,12/$B$5))', []],
      ['C', 15, 199, 'trans', '=IF(A15="","",G14)', []],
      ['D', 14, 14, 'lit', '=C14*$B$4/$B$5', []],
      ['D', 15, 199, 'trans', '=IF(A15="","",C15*$B$4/$B$5)', []],
      ['E', 14, 14, 'lit', '=$B$8', []],
      ['E', 15, 199, 'trans', '=IF(A15="","",IF($B$8+D15>H14,0,E14))', []],
      ['F', 9, 9, 'lit', '=EDATE(Inputs!$B$11,100*12)', []],
      ['F', 10, 10, 'lit', '=IFERROR(_xlfn.XLOOKUP(1,I14:I199,B14:B199),"")', []],
      ['F', 14, 14, 'lit', '=MIN($B$8,E14)', []],
      ['F', 15, 199, 'trans', '=IF(A15="","",MIN($B$8,E15))', []],
      ['G', 14, 14, 'lit', '=C14+D14+E14', []],
      ['G', 15, 199, 'trans', '=IF(A15="","",C15+D15+E15)', []],
      ['H', 14, 14, 'lit', '=$B$11-G14', []],
      ['H', 15, 199, 'trans', '=IF(A15="","",$B$11-G15)', []],
      ['I', 14, 14, 'lit', '=IF(ISNUMBER(E15),IF(E15<=0,1,0),0)', []],
      ['I', 15, 198, 'trans', '=IF(A15="","",IF(ISNUMBER(E16),IF(E16<=0,1,0),0))', []],
      ['I', 199, 199, 'lit', '=IF(A199="","",0)', []],
    ],
    "colfmt": {'H': '0', 'G': '0', 'D': '0', 'C': '0', 'B': 'mm-dd-yy', 'E': '0'},
    "fmt_exc": {'D3': 'General', 'B3': 'General', 'B4': '0%', 'E8': 'General', 'B11': '0', 'B10': '0%', 'C13': 'General', 'F9': 'mm-dd-yy', 'B5': 'General', 'B13': 'General', 'B8': '0', 'E9': 'General', 'E10': 'General', 'F10': 'mm-dd-yy', 'E13': 'General', 'G13': 'General', 'D13': 'General', 'B7': '0', 'B9': '0', 'H13': 'General'},
    "widths": {'A': 27.27, 'B': 16.45, 'C': 15.82, 'D': 7.45, 'E': 13.82, 'F': 21.82, 'G': 20.18, 'H': 28.18, 'I': 10.45},
  },
  'Loan': {
    "statics": {
        'A1': 'LOAN MODULE - Multiple Loans & Repayments',
        'E1': 'NOTES:',
        'E2': '1. ADVANCE METHOD: Interest is charged at the beginning of each period (prepaid interest).',
        'A3': 'Input Parameters',
        'E3': '2. ARREAR METHOD: Interest is charged at the end of each period (accrued interest).',
        'A4': 'Loan Interest Rate',
        'E4': '3. Both methods use the formula: Interest = Principal × ((1 + Rate)^(Days/365) - 1)',
        'A5': 'Guaranteed Interest Rate',
        'E5': '4. Outstanding balance shown is the amount owed after each transaction/event.',
        'A6': 'Issue Age',
        'E6': '5. Loan transactions: Positive amounts = new loans, Negative amounts = repayments.',
        'A7': 'Issue Date',
        'E7': '6. Enter transactions in chronological order (sorted by date).',
        'A8': 'Valuation Date',
        'E8': '7. VALUATION row at row 114 calculates the final outstanding from last transaction to Valuation Date.',
        'A9': 'Face Amount (SA)',
        'A10': 'Premium Paying Term (PPT)',
        'E10': 'Advance Loan Outstanding',
        'E11': 'Arrear Loan Outstanding',
        'A13': 'LOAN TRANSACTIONS (Enter up to 100 loans; use negative amounts for repayments)',
        'E13': 'ADVANCE METHOD (Interest charged at beginning of period)',
        'N13': 'ARREAR METHOD (Interest charged at end of period)',
        'A14': 'Loan #',
        'B14': 'Loan Date',
        'C14': 'Loan Amount',
        'E14': 'Event #',
        'F14': 'Date',
        'G14': 'Type',
        'H14': 'Amount',
        'I14': 'Days from Prior',
        'J14': 'Interest Accrued',
        'K14': 'Outstanding After',
        'L14': 'Is Anniversary?',
        'N14': 'Event #',
        'O14': 'Date',
        'P14': 'Type',
        'Q14': 'Amount',
        'R14': 'Days from Prior',
        'S14': 'Interest Accrued',
        'T14': 'Outstanding After',
        'U14': 'Is Anniversary?',
        'A15': 1,
        'A16': 2,
        'A17': 3,
        'A18': 4,
        'A19': 5,
        'A20': 6,
        'A21': 7,
        'A22': 8,
        'A23': 9,
        'A24': 10,
        'A25': 11,
        'A26': 12,
        'A27': 13,
        'A28': 14,
        'A29': 15,
        'A30': 16,
        'A31': 17,
        'A32': 18,
        'A33': 19,
        'A34': 20,
        'A35': 21,
        'A36': 22,
        'A37': 23,
        'A38': 24,
        'A39': 25,
        'A40': 26,
        'A41': 27,
        'A42': 28,
        'A43': 29,
        'A44': 30,
        'A45': 31,
        'A46': 32,
        'A47': 33,
        'A48': 34,
        'A49': 35,
        'A50': 36,
        'A51': 37,
        'A52': 38,
        'A53': 39,
        'A54': 40,
        'A55': 41,
        'A56': 42,
        'A57': 43,
        'A58': 44,
        'A59': 45,
        'A60': 46,
        'A61': 47,
        'A62': 48,
        'A63': 49,
        'A64': 50,
        'A65': 51,
        'A66': 52,
        'A67': 53,
        'A68': 54,
        'A69': 55,
        'A70': 56,
        'A71': 57,
        'A72': 58,
        'A73': 59,
        'A74': 60,
        'A75': 61,
        'A76': 62,
        'A77': 63,
        'A78': 64,
        'A79': 65,
        'A80': 66,
        'A81': 67,
        'A82': 68,
        'A83': 69,
        'A84': 70,
        'A85': 71,
        'A86': 72,
        'A87': 73,
        'A88': 74,
        'A89': 75,
        'A90': 76,
        'A91': 77,
        'A92': 78,
        'A93': 79,
        'A94': 80,
        'A95': 81,
        'A96': 82,
        'A97': 83,
        'A98': 84,
        'A99': 85,
        'A100': 86,
        'A101': 87,
        'A102': 88,
        'A103': 89,
        'A104': 90,
        'A105': 91,
        'A106': 92,
        'A107': 93,
        'A108': 94,
        'A109': 95,
        'A110': 96,
        'A111': 97,
        'A112': 98,
        'A113': 99,
        'A114': 100,
    },
    "arrays": ['B15', 'C15'],
    "fills": [
      ['B', 4, 4, 'lit', '=Inputs!$B$40', []],
      ['B', 5, 5, 'lit', '=Inputs!$B$21', []],
      ['B', 6, 6, 'lit', '=Inputs!$B$7', []],
      ['B', 7, 7, 'lit', '=Inputs!$B$11', []],
      ['B', 8, 8, 'lit', '=Inputs!$B$12', []],
      ['B', 9, 9, 'lit', '=Inputs!$B$18', []],
      ['B', 10, 10, 'lit', '=Inputs!$B$8', []],
      ['B', 15, 15, 'lit', '=_xlfn.LET(\n    _xlpm.loan_dates, Inputs!A44:A143,\n    _xlpm.loan_amts, Inputs!B44:B143,\n    _xlpm.div_dates, Inputs!A33:A37,\n    _xlpm.div_amts, Inputs!B33:B37,\n    _xlpm.div_option, Inputs!B20,\n    _xlpm.div_sign, IF(_xlpm.div_option="Reduce Loan", -1, 0),\n\n    _xlpm.loan_data, _xlfn.HSTACK(_xlpm.loan_dates, _xlpm.loan_amts),\n    _xlpm.div_data, _xlfn.HSTACK(_xlpm.div_dates, _xlpm.div_amts * _xlpm.div_sign),\n\n    _xlpm.all_data, _xlfn.VSTACK(_xlpm.loan_data, _xlpm.div_data),\n    _xlpm.dates_col, INDEX(_xlpm.all_data,,1),\n    _xlpm.amts_col, INDEX(_xlpm.all_data,,2),\n\n    _xlpm.valid, (_xlpm.dates_col<>"") * (_xlpm.amts_col<>"") * (_xlpm.amts_col<>0),\n    _xlpm.filtered, _xlfn._xlws.FILTER(_xlpm.all_data, _xlpm.valid, ""),\n    _xlpm.sorted, _xlfn._xlws.SORT(_xlpm.filtered, 1, 1),\n\n    INDEX(_xlpm.sorted,,1)\n)', []],
      ['C', 15, 15, 'lit', '=_xlfn.LET(\n    _xlpm.loan_dates, Inputs!A44:A143,\n    _xlpm.loan_amts, Inputs!B44:B143,\n    _xlpm.div_dates, Inputs!A33:A37,\n    _xlpm.div_amts, Inputs!B33:B37,\n    _xlpm.div_option, Inputs!B20,\n    _xlpm.div_sign, IF(_xlpm.div_option="Reduce Loan", -1, 0),\n\n    _xlpm.loan_data, _xlfn.HSTACK(_xlpm.loan_dates, _xlpm.loan_amts),\n    _xlpm.div_data, _xlfn.HSTACK(_xlpm.div_dates, _xlpm.div_amts * _xlpm.div_sign),\n\n    _xlpm.all_data, _xlfn.VSTACK(_xlpm.loan_data, _xlpm.div_data),\n    _xlpm.dates_col, INDEX(_xlpm.all_data,,1),\n    _xlpm.amts_col, INDEX(_xlpm.all_data,,2),\n\n    _xlpm.valid, (_xlpm.dates_col<>"") * (_xlpm.amts_col<>"") * (_xlpm.amts_col<>0),\n    _xlpm.filtered, _xlfn._xlws.FILTER(_xlpm.all_data, _xlpm.valid, ""),\n    _xlpm.sorted, _xlfn._xlws.SORT(_xlpm.filtered, 1, 1),\n\n    INDEX(_xlpm.sorted,,2)\n)', []],
      ['E', 15, 15, 'lit', '=IF($B15="","",ROWS($E$15:E15))', []],
      ['E', 16, 114, 'trans', '=IF($B16<>"",ROWS($E$15:E16),IF(AND($B16="",$B15<>"",E15<>""),$E15+1,""))', []],
      ['F', 15, 15, 'lit', '=IF($B15="","",$B15)', []],
      ['F', 16, 114, 'trans', '=IF($B16<>"", $B16, IF(AND($B16="",$B15<>"",F15<>""), $B$8, ""))', []],
      ['G', 10, 10, 'lit', '=INDEX(K15:K114,MATCH("VALUATION",G15:G114,0))', []],
      ['G', 11, 11, 'lit', '=INDEX(T15:T114,MATCH("VALUATION",P15:P114,0))', []],
      ['G', 15, 15, 'lit', '=IF($B15="","","LOAN")', []],
      ['G', 16, 114, 'trans', '=IF($B16<>"","LOAN",IF(AND($B16="",$B15<>"",G15<>""),"VALUATION",""))', []],
      ['H', 15, 15, 'lit', '=IF($B15="","",$C15)', []],
      ['H', 16, 114, 'trans', '=IF($B16<>"",$C16,IF(AND($B16="",$B15<>"",H15<>""),0,""))', []],
      ['I', 15, 15, 'lit', '=IF($B15="","",0)', []],
      ['I', 16, 114, 'trans', '=IF($B16<>"",F16-F15,IF(AND($B16="",$B15<>"",I15<>""),F16-F15,""))', []],
      ['J', 15, 15, 'lit', '=IF($B15="","",0)', []],
      ['J', 16, 114, 'trans', '=IF($B16<>"",K16-K15-H16,IF(AND($B16="",$B15<>"",J15<>""),K16-K15,""))', []],
      ['K', 15, 15, 'lit', '=IF($B15="","",H15)', []],
      ['K', 16, 114, 'trans', '=IF($B16<>"",K15*(1+$B$4)^(I16/365)+IF(H16>0,H16*(1+$B$4)^(I16/365),H16),IF(AND($B16="",$B15<>"",K15<>""),K15*(1+$B$4)^(I16/365),""))', []],
      ['L', 15, 15, 'lit', '=IF($B15="","",IF(AND(MONTH(F15)=MONTH($B$7),DAY(F15)=DAY($B$7)),"Yes",""))', []],
      ['L', 16, 114, 'trans', '=IF(F16="","",IF(AND(MONTH(F16)=MONTH($B$7),DAY(F16)=DAY($B$7)),"Yes",""))', []],
      ['N', 15, 15, 'lit', '=IF($B15="","",ROWS($N$15:N15))', []],
      ['N', 16, 114, 'trans', '=IF($B16<>"",ROWS($N$15:N16),IF(AND($B16="",$B15<>"",N15<>""),$N15+1,""))', []],
      ['O', 15, 15, 'lit', '=IF($B15="","",$B15)', []],
      ['O', 16, 114, 'trans', '=IF($B16<>"", $B16, IF(AND($B16="",$B15<>"",O15<>""), $B$8, ""))', []],
      ['P', 15, 15, 'lit', '=IF($B15="","","LOAN")', []],
      ['P', 16, 114, 'trans', '=IF($B16<>"","LOAN",IF(AND($B16="",$B15<>"",P15<>""),"VALUATION",""))', []],
      ['Q', 15, 15, 'lit', '=IF($B15="","",$C15)', []],
      ['Q', 16, 114, 'trans', '=IF($B16<>"",$C16,IF(AND($B16="",$B15<>"",Q15<>""),0,""))', []],
      ['R', 15, 15, 'lit', '=IF($B15="","",0)', []],
      ['R', 16, 114, 'trans', '=IF($B16<>"",O16-O15,IF(AND($B16="",$B15<>"",R15<>""),O16-O15,""))', []],
      ['S', 15, 15, 'lit', '=IF($B15="","",0)', []],
      ['S', 16, 114, 'trans', '=IF($B16<>"",T15*((1+$B$4)^(MAX(R16-1,0)/365)-1),IF(AND($B16="",$B15<>"",S15<>""),T15*((1+$B$4)^(MAX(R16-1,0)/365)-1),""))', []],
      ['T', 15, 15, 'lit', '=IF($B15="","",Q15)', []],
      ['T', 16, 114, 'trans', '=IF($B16<>"",T15+S16+Q16,IF(AND($B16="",$B15<>"",T15<>""),T15+S16,""))', []],
      ['U', 15, 15, 'lit', '=IF($B15="","",IF(AND(MONTH(O15)=MONTH($B$7),DAY(O15)=DAY($B$7)),"Yes",""))', []],
      ['U', 16, 114, 'trans', '=IF(O16="","",IF(AND(MONTH(O16)=MONTH($B$7),DAY(O16)=DAY($B$7)),"Yes",""))', []],
      ['V', 22, 103, 'trans', '=IF(B33="","",B33+0.1)', []],
      ['V', 104, 153, 'trans', '=IF(S104="","",S104+0.15)', []],
      ['V', 154, 154, 'lit', '=$B$8+0.3', []],
      ['W', 3, 10, 'trans', '=IF(X3="","",ROW()-2)', []],
      ['W', 12, 253, 'trans', '=IF(X12="","",ROW()-2)', []],
    ],
    "colfmt": {'T': '#,##0.00', 'F': 'mm/dd/yyyy', 'O': 'mm/dd/yyyy', 'W': '0', 'S': '#,##0.00', 'V': '0.0', 'A': '0', 'J': '#,##0.00', 'K': '#,##0.00', 'U': '#,##0.00', 'C': '#,##0.00'},
    "fmt_exc": {'V128': '0.00', 'V149': '0.00', 'V107': '0.00', 'H16': '#,##0.00', 'A9': 'General', 'V135': '0.00', 'V115': '0.00', 'V141': '0.00', 'V108': '0.00', 'O14': 'General', 'V116': '0.00', 'A14': 'General', 'V120': '0.00', 'V143': '0.00', 'V114': '0.00', 'V104': '0.00', 'K14': 'General', 'V112': '0.00', 'V139': '0.00', 'V129': '0.00', 'V133': '0.00', 'J14': 'General', 'B8': 'mm/dd/yyyy', 'A10': 'General', 'V134': '0.00', 'B15': 'mm-dd-yy', 'V124': '0.00', 'V152': '0.00', 'B7': 'mm/dd/yyyy', 'V105': '0.00', 'T14': 'General', 'V117': '0.00', 'V121': '0.00', 'V130': '0.00', 'F14': 'General', 'H15': '#,##0.00', 'V111': '0.00', 'V132': '0.00', 'A7': 'General', 'V125': '0.00', 'A6': 'General', 'V136': '0.00', 'V154': '0.00', 'Q15': '#,##0.00', 'V137': '0.00', 'B4': '0.00%', 'V118': '0.00', 'V153': '0.00', 'V122': '0.00', 'V138': '0.00', 'A3': 'General', 'V146': '0.00', 'V113': '0.00', 'V119': '0.00', 'V148': '0.00', 'Q16': '#,##0.00', 'C14': 'General', 'V140': '0.00', 'B5': '0.00%', 'V144': '0.00', 'G11': '#,##0.00', 'A13': 'General', 'A4': 'General', 'V109': '0.00', 'V147': '0.00', 'V131': '0.00', 'A8': 'General', 'V145': '0.00', 'V151': '0.00', 'A1': 'General', 'V150': '0.00', 'V142': '0.00', 'V127': '0.00', 'G10': '#,##0.00', 'V123': '0.00', 'V126': '0.00', 'S14': 'mm/dd/yyyy', 'A5': 'General', 'V110': '0.00', 'V106': '0.00'},
    "widths": {'A': 24.27, 'B': 24.45, 'C': 22.0, 'D': 10.09, 'E': 9.09, 'F': 15.45, 'G': 12.73, 'H': 15.45, 'I': 17.27, 'J': 18.18, 'K': 20.91, 'L': 16.36, 'M': 6.91, 'N': 9.09, 'O': 15.45, 'P': 12.73, 'Q': 15.45, 'R': 17.27, 'S': 18.18, 'T': 20.91, 'U': 16.36, 'V': 9.63, 'W': 14.36, 'X': 10.54, 'Z': 9.0, 'AA': 10.54, 'AB': 8.09, 'AC': 9.18, 'AD': 10.54, 'AE': 9.09},
  },
}


if __name__ == "__main__":
    main()
