"""
align_to_module_34.py
========================
STAGE 3 of the "run v9, then align to 34" pipeline.

Run this AFTER calc_engine_v9.py has produced its output workbook. It
opens that file and overwrites it -- every calculation sheet, the
Inputs sheet, and the CSO Rates sheet -- with formulas verified against
your real module 34 reference workbook. Nothing v9 calculated survives;
v9's output is used here purely as a starting workbook shell (sheet
names / general xlsx container), not as a source of any values.

WHY REWRITE EVERYTHING INSTEAD OF PATCHING SELECTIVELY
------------------------------------------------------------
calc_engine_v9.py's formulas are a hand-reconstructed approximation of
a *different* product design (31/32's paid-up-at-age logic) built by
trial and error against its own comments ("Key fixes: ... correct
scaling"). Virtually none of its CV_RATES / Commutation /
Projection_Monthly / PUA_TABLE / TERM_RATES formulas match module 34's
real formulas cell-for-cell -- there isn't a small, identifiable set of
"differences" to patch. So rather than attempt a fragile line-by-line
diff-and-patch, this script does the same thing calc_engine_34.py does
when building from scratch: delete each target sheet and rebuild it
completely from the same verified formula data (~211 templates + 437
explicit formulas + ~618 static cells, extracted from your real module
34 reference workbook and checked formula-for-formula, 0 mismatches
across 47,817 cells -- see calc_engine_34.py's docstring for the
extraction method if you want the full detail).

The end result is functionally identical to what calc_engine_34.py
alone would produce -- this script exists only so v9 can remain the
common entry point across plan codes if that's how your pipeline is
organized, with this as the module-34-specific correction layer on
top. There is no accuracy benefit to routing through v9 first; it's
purely an architectural choice.

WHAT GETS REMOVED
----------------------
'Checked_Policies', 'Mortality1' (or a stray 'Mortality' sheet -- v9
creates a second, unused mortality sheet alongside the real
'MORTALITY' one, which Excel auto-renames to 'Mortality1' on save
since sheet names are case-insensitive), and any other sheet name not
part of module 34's real 14-sheet structure. None of these are read by
any formula in the real reference workbook.

WHAT THIS SCRIPT DOES NOT DO
---------------------------------
It does not try to recover or reuse anything from v9's Loan sheet
(loan transaction dates/amounts) -- module 34's own reference Loan
sheet contains manually-entered example data for one policy, same
caveat as calc_engine_34.py. If v9's Loan sheet had real per-policy
loan transactions wired in some other way, that data does not carry
forward here; flag it if that matters for your case.

USAGE
-----
    python align_to_module_34.py \\
        --source WL_model_34.xlsx \\
        --inputs Inputs_v2.csv \\
        --cso cso_rates.csv \\
        --output WL_model_34_aligned.xlsx

--inputs/--cso must be the REAL (unpadded) module 34 CSVs -- not the
v9-padded ones prepare_inputs_for_calc_v9.py wrote (those still have
the dummy Paid-Up Age / Expense Loading Factor rows, which have no
place in module 34's actual Inputs sheet).

Requires: openpyxl only (single self-contained file, same shape as
calc_engine_34.py -- no other local files needed).
"""


import argparse
import csv
import os
import re
import datetime as dt

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula
import datetime



# ========================================================================
# EMBEDDED MODEL DATA -- extracted from the reference workbook and verified
# (see module docstring above). Real, readable data -- no encoding, no blob.
# Identical to the data embedded in calc_engine_34.py.
# ========================================================================
SHEET_ORDER = ['Inputs', 'CSO Rates', 'MORTALITY', 'CV_RATES', 'PUA_TABLE', 'TERM_RATES', 'DIvidend_Assumption', 'Projection_Monthly', 'RPU', 'ETI', 'APL', 'Commutation', 'Loan']

# 'Inputs' sheet: every OTHER cell is populated straight from Inputs.csv at
# runtime (see build_inputs_sheet()). These are just the cells that are
# themselves formulas in the reference workbook and must never be
# overwritten with a CSV value (Issue Age, Current age, Modal Premium, etc.)
INPUTS_FORMULAS = {
    'B14': ('formula', '=INT((B12-B13)/365.25)'),
    'B15': ('formula', '=INT((B12-B11)/365.25)'),
    'B16': ('formula', '=B15+1'),
    'B17': ('formula', '=B9*1000'),
    'B22': ('formula', '=B21*B24'),
    'B24': ('array', '=_xlfn.SWITCH(B23,"Annual",1,"Semi-Annual",0.5,"Quarterly",0.25,"Monthly",1/12,1)', 'B24'),
    'B26': ('formula', '=MIN(B8,B15)'),
    'B27': ('formula', '=EDATE(B11,B8*12)>B12'),
    'B7': ('formula', '=ROUND((B11-B13)/365.25,0)'),
    'E2': ('array', '=_xlfn.LET(\n    _xlpm.ValDate, B12,\n    _xlpm.DateRange, Projection_Monthly!$B$2:$B$985,\n    _xlpm.CVRange, Projection_Monthly!$AA$2:$AA$985,\n    _xlpm.ExactMatch, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.CVRange, "", 0),\n    IF(_xlpm.ExactMatch<>"", _xlpm.ExactMatch,\n        _xlfn.LET(\n            _xlpm.DateBefore, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", -1),\n            _xlpm.DateAfter, _xlfn.XLOOKUP(_xlpm.ValDate, _xlpm.DateRange, _xlpm.DateRange, "", 1),\n            _xlpm.CVBefore, _xlfn.XLOOKUP(_xlpm.DateBefore, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.CVAfter, _xlfn.XLOOKUP(_xlpm.DateAfter, _xlpm.DateRange, _xlpm.CVRange),\n            _xlpm.DaysBetween, _xlpm.DateAfter - _xlpm.DateBefore,\n            _xlpm.DaysFromBefore, _xlpm.ValDate - _xlpm.DateBefore,\n            _xlpm.Fraction, _xlpm.DaysFromBefore / _xlpm.DaysBetween,\n            _xlpm.CVBefore + _xlpm.Fraction * (_xlpm.CVAfter - _xlpm.CVBefore)\n        )\n    )\n)', 'E2'),
    'E3': ('formula', '=(PUA_TABLE!I72*B10)'),
    'E4': ('formula', '=ETI!E6'),
    'E5': ('formula', '=RPU!B17'),
    'E6': ('formula', '=APL!H4'),
}

# Number formats for the date/percent input cells (so they display correctly)
INPUTS_FORMATS = {
    'B11': 'mm-dd-yy',
    'B12': 'mm-dd-yy',
    'B13': 'mm-dd-yy',
    'B18': '0.00%',
    'B20': '0.00%',
    'B21': '0',
    'C12': 'd-mmm',
    'C13': 'd-mmm',
    'O10': 'mm-dd-yy',
    'O8': 'mm-dd-yy',
    'O9': 'mm-dd-yy',
    'P10': 'mm-dd-yy',
    'P8': 'mm-dd-yy',
    'P9': 'mm-dd-yy',
    'Q10': 'mm-dd-yy',
    'Q8': 'mm-dd-yy',
    'Q9': 'mm-dd-yy',
    'R10': 'mm-dd-yy',
    'R8': 'mm-dd-yy',
    'R9': 'mm-dd-yy',
    'S10': 'mm-dd-yy',
    'S8': 'mm-dd-yy',
    'S9': 'mm-dd-yy',
}

# The 12 formula-driven calculation sheets. For each sheet:
#   'statics'       -- non-formula cells (headers, row labels) written as-is
#   'col_templates' -- {column_letter: [{'template': ..., 'ranges': [(r1,r2),...]}]}
#                      one template string reused (via row substitution) across
#                      one or more contiguous row ranges in that column
#   'explicit'      -- {'A2': '=...'} one-off formulas that don't fit a template
#   'col_format'    -- {column_letter: number_format} applied to that column's
#                      formula cells only (dates/percents/etc.)
FUNCTIONAL_SHEETS = {}

# ------------------------------------------------------------------------
# MORTALITY
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['MORTALITY'] = {
    'dims': (122, 5),
    'statics': {
        'A1': 'Age',
        'A2': 0,
        'D1': 'Age',
        'D10': 8,
        'D100': 98,
        'D101': 99,
        'D102': 100,
        'D103': 101,
        'D104': 102,
        'D105': 103,
        'D106': 104,
        'D107': 105,
        'D108': 106,
        'D109': 107,
        'D11': 9,
        'D110': 108,
        'D111': 109,
        'D112': 110,
        'D113': 111,
        'D114': 112,
        'D115': 113,
        'D116': 114,
        'D117': 115,
        'D118': 116,
        'D119': 117,
        'D12': 10,
        'D120': 118,
        'D121': 119,
        'D122': 120,
        'D13': 11,
        'D14': 12,
        'D15': 13,
        'D16': 14,
        'D17': 15,
        'D18': 16,
        'D19': 17,
        'D2': 0,
        'D20': 18,
        'D21': 19,
        'D22': 20,
        'D23': 21,
        'D24': 22,
        'D25': 23,
        'D26': 24,
        'D27': 25,
        'D28': 26,
        'D29': 27,
        'D3': 1,
        'D30': 28,
        'D31': 29,
        'D32': 30,
        'D33': 31,
        'D34': 32,
        'D35': 33,
        'D36': 34,
        'D37': 35,
        'D38': 36,
        'D39': 37,
        'D4': 2,
        'D40': 38,
        'D41': 39,
        'D42': 40,
        'D43': 41,
        'D44': 42,
        'D45': 43,
        'D46': 44,
        'D47': 45,
        'D48': 46,
        'D49': 47,
        'D5': 3,
        'D50': 48,
        'D51': 49,
        'D52': 50,
        'D53': 51,
        'D54': 52,
        'D55': 53,
        'D56': 54,
        'D57': 55,
        'D58': 56,
        'D59': 57,
        'D6': 4,
        'D60': 58,
        'D61': 59,
        'D62': 60,
        'D63': 61,
        'D64': 62,
        'D65': 63,
        'D66': 64,
        'D67': 65,
        'D68': 66,
        'D69': 67,
        'D7': 5,
        'D70': 68,
        'D71': 69,
        'D72': 70,
        'D73': 71,
        'D74': 72,
        'D75': 73,
        'D76': 74,
        'D77': 75,
        'D78': 76,
        'D79': 77,
        'D8': 6,
        'D80': 78,
        'D81': 79,
        'D82': 80,
        'D83': 81,
        'D84': 82,
        'D85': 83,
        'D86': 84,
        'D87': 85,
        'D88': 86,
        'D89': 87,
        'D9': 7,
        'D90': 88,
        'D91': 89,
        'D92': 90,
        'D93': 91,
        'D94': 92,
        'D95': 93,
        'D96': 94,
        'D97': 95,
        'D98': 96,
        'D99': 97,
        'E1': 'q_x (1958 CET ANB MALE)',
        'E10': 0.00145,
        'E100': 0.85261,
        'E101': 1,
        'E102': 0,
        'E103': 0,
        'E104': 0,
        'E105': 0,
        'E106': 0,
        'E107': 0,
        'E108': 0,
        'E109': 0,
        'E11': 0.00144,
        'E110': 0,
        'E111': 0,
        'E112': 0,
        'E113': 0,
        'E114': 0,
        'E115': 0,
        'E116': 0,
        'E117': 0,
        'E118': 0,
        'E119': 0,
        'E12': 0.00143,
        'E120': 0,
        'E121': 0,
        'E122': 0,
        'E13': 0.00144,
        'E14': 0.00147,
        'E15': 0.0015,
        'E16': 0.00155,
        'E17': 0.0016,
        'E18': 0.00165,
        'E19': 0.0017,
        'E2': 0.00376,
        'E20': 0.00173,
        'E21': 0.00177,
        'E22': 0.0018,
        'E23': 0.00182,
        'E24': 0.00184,
        'E25': 0.00186,
        'E26': 0.00189,
        'E27': 0.00191,
        'E28': 0.00194,
        'E29': 0.00197,
        'E3': 0.00162,
        'E30': 0.00201,
        'E31': 0.00205,
        'E32': 0.0021,
        'E33': 0.00215,
        'E34': 0.0022,
        'E35': 0.00225,
        'E36': 0.00233,
        'E37': 0.0024,
        'E38': 0.00251,
        'E39': 0.00264,
        'E4': 0.00156,
        'E40': 0.00279,
        'E41': 0.00297,
        'E42': 0.00317,
        'E43': 0.00343,
        'E44': 0.00373,
        'E45': 0.00402,
        'E46': 0.00432,
        'E47': 0.00463,
        'E48': 0.00494,
        'E49': 0.00527,
        'E5': 0.00154,
        'E50': 0.00563,
        'E51': 0.00602,
        'E52': 0.00645,
        'E53': 0.0069,
        'E54': 0.00741,
        'E55': 0.008,
        'E56': 0.00859,
        'E57': 0.00922,
        'E58': 0.00984,
        'E59': 0.01044,
        'E6': 0.00152,
        'E60': 0.01101,
        'E61': 0.01162,
        'E62': 0.01231,
        'E63': 0.01317,
        'E64': 0.01425,
        'E65': 0.01563,
        'E66': 0.01723,
        'E67': 0.01897,
        'E68': 0.0208,
        'E69': 0.02266,
        'E7': 0.00151,
        'E70': 0.02449,
        'E71': 0.02647,
        'E72': 0.02874,
        'E73': 0.0315,
        'E74': 0.03493,
        'E75': 0.03914,
        'E76': 0.04411,
        'E77': 0.04971,
        'E78': 0.05586,
        'E79': 0.06245,
        'E8': 0.00148,
        'E80': 0.06949,
        'E81': 0.07716,
        'E82': 0.08579,
        'E83': 0.09568,
        'E84': 0.10712,
        'E85': 0.12029,
        'E86': 0.13495,
        'E87': 0.15093,
        'E88': 0.16808,
        'E89': 0.18632,
        'E9': 0.00147,
        'E90': 0.20563,
        'E91': 0.22612,
        'E92': 0.24798,
        'E93': 0.27153,
        'E94': 0.29745,
        'E95': 0.32696,
        'E96': 0.3631,
        'E97': 0.41252,
        'E98': 0.48846,
        'E99': 0.61746,
    },
    'col_templates': {
        'A': [
            {'template': '=+A{R-1}+1', 'ranges': [(3, 101)]},
        ],
        'B': [
            {'template': "=INDEX('CSO Rates'!$A$5:$P$126,MATCH(A{R+0},'CSO Rates'!$A$5:$A$126,0),MATCH($B$1,'CSO Rates'!$A$4:$P$4,0))", 'ranges': [(2, 101)]},
        ],
    },
    'explicit': {
        'B1': '=_xlfn.CONCAT(Inputs!B2,Inputs!B6,Inputs!B3)',
    },
    'col_format': {
        'B': '0.00000_)',
    },
}
FUNCTIONAL_SHEETS['CV_RATES'] = {
    'dims': (1061, 11),
    'statics': {
        'A1': 'Code',
        'B1': 'Issue Age',
        'C1': 'Duration',
        'D1': 'Attained Age',
        'E1': 'qx',
        'F1': 'px',
        'G1': 'Interest Rate',
        'H1': 'Beta (NLP/1000)',
        'I1': 'CV (BOY)',
        'J1': 'CV (EOY)',
        'K1': 'CV FACTOR',
    },
    'col_templates': {
        'B': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",Inputs!$B$7)', 'ranges': [(2, 101)]},
        ],
        'C': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",{R-1})', 'ranges': [(2, 101)]},
        ],
        'D': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",Inputs!$B$7+{R-1}-1)', 'ranges': [(2, 101)]},
        ],
        'E': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",IFERROR(INDEX(MORTALITY!$B:$B,MATCH(Inputs!$B$7+{R-1}-1,MORTALITY!$A:$A,0)),0))', 'ranges': [(2, 101)]},
        ],
        'F': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",IFERROR(1-E{R+0},1))', 'ranges': [(2, 101)]},
        ],
        'G': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",Inputs!$B$20)', 'ranges': [(2, 101)]},
        ],
        'H': [
            {'template': '=IF(A{R+0}="","",IFERROR(1000*INDEX(Commutation!$I:$I,MATCH(Inputs!$B$7,Commutation!$A:$A,0))/(INDEX(Commutation!$H:$H,MATCH(Inputs!$B$7,Commutation!$A:$A,0))-INDEX(Commutation!$H:$H,MATCH(MIN(Inputs!$B$7+Inputs!$B$8,MAX(Commutation!$A:$A)),Commutation!$A:$A,0))),0))', 'ranges': [(2, 101)]},
        ],
        'I': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",IFERROR(J{R-1},0))', 'ranges': [(3, 101)]},
        ],
        'J': [
            {'template': '=IF(A{R+0}="","",IF(E{R+0}=1,1000,IFERROR(((I{R+0}+IF(C{R+0}<=Inputs!$B$8,$H$2,0))*(1+G{R+0})-E{R+0}*1000)/F{R+0},0)))', 'ranges': [(2, 101)]},
        ],
        'K': [
            {'template': '=IF(Inputs!$B$7+{R-1}-1>MAX(MORTALITY!$A:$A),"",IFERROR(J{R+0},0))', 'ranges': [(2, 101)]},
        ],
    },
    'explicit': {
        'A10': '=IF(Inputs!$B$7+9-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A100': '=IF(Inputs!$B$7+99-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A101': '=IF(Inputs!$B$7+100-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A11': '=IF(Inputs!$B$7+10-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A12': '=IF(Inputs!$B$7+11-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A13': '=IF(Inputs!$B$7+12-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A14': '=IF(Inputs!$B$7+13-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A15': '=IF(Inputs!$B$7+14-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A16': '=IF(Inputs!$B$7+15-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A17': '=IF(Inputs!$B$7+16-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A18': '=IF(Inputs!$B$7+17-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A19': '=IF(Inputs!$B$7+18-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A2': '=IF(Inputs!$B$7+1-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A20': '=IF(Inputs!$B$7+19-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A21': '=IF(Inputs!$B$7+20-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A22': '=IF(Inputs!$B$7+21-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A23': '=IF(Inputs!$B$7+22-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A24': '=IF(Inputs!$B$7+23-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A25': '=IF(Inputs!$B$7+24-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A26': '=IF(Inputs!$B$7+25-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A27': '=IF(Inputs!$B$7+26-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A28': '=IF(Inputs!$B$7+27-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A29': '=IF(Inputs!$B$7+28-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A3': '=IF(Inputs!$B$7+2-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A30': '=IF(Inputs!$B$7+29-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A31': '=IF(Inputs!$B$7+30-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A32': '=IF(Inputs!$B$7+31-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A33': '=IF(Inputs!$B$7+32-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A34': '=IF(Inputs!$B$7+33-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A35': '=IF(Inputs!$B$7+34-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A36': '=IF(Inputs!$B$7+35-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A37': '=IF(Inputs!$B$7+36-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A38': '=IF(Inputs!$B$7+37-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A39': '=IF(Inputs!$B$7+38-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A4': '=IF(Inputs!$B$7+3-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A40': '=IF(Inputs!$B$7+39-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A41': '=IF(Inputs!$B$7+40-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A42': '=IF(Inputs!$B$7+41-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A43': '=IF(Inputs!$B$7+42-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A44': '=IF(Inputs!$B$7+43-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A45': '=IF(Inputs!$B$7+44-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A46': '=IF(Inputs!$B$7+45-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A47': '=IF(Inputs!$B$7+46-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A48': '=IF(Inputs!$B$7+47-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A49': '=IF(Inputs!$B$7+48-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A5': '=IF(Inputs!$B$7+4-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A50': '=IF(Inputs!$B$7+49-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A51': '=IF(Inputs!$B$7+50-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A52': '=IF(Inputs!$B$7+51-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A53': '=IF(Inputs!$B$7+52-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A54': '=IF(Inputs!$B$7+53-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A55': '=IF(Inputs!$B$7+54-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A56': '=IF(Inputs!$B$7+55-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A57': '=IF(Inputs!$B$7+56-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A58': '=IF(Inputs!$B$7+57-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A59': '=IF(Inputs!$B$7+58-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A6': '=IF(Inputs!$B$7+5-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A60': '=IF(Inputs!$B$7+59-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A61': '=IF(Inputs!$B$7+60-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A62': '=IF(Inputs!$B$7+61-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A63': '=IF(Inputs!$B$7+62-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A64': '=IF(Inputs!$B$7+63-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A65': '=IF(Inputs!$B$7+64-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A66': '=IF(Inputs!$B$7+65-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A67': '=IF(Inputs!$B$7+66-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A68': '=IF(Inputs!$B$7+67-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A69': '=IF(Inputs!$B$7+68-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A7': '=IF(Inputs!$B$7+6-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A70': '=IF(Inputs!$B$7+69-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A71': '=IF(Inputs!$B$7+70-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A72': '=IF(Inputs!$B$7+71-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A73': '=IF(Inputs!$B$7+72-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A74': '=IF(Inputs!$B$7+73-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A75': '=IF(Inputs!$B$7+74-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A76': '=IF(Inputs!$B$7+75-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A77': '=IF(Inputs!$B$7+76-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A78': '=IF(Inputs!$B$7+77-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A79': '=IF(Inputs!$B$7+78-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A8': '=IF(Inputs!$B$7+7-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A80': '=IF(Inputs!$B$7+79-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A81': '=IF(Inputs!$B$7+80-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A82': '=IF(Inputs!$B$7+81-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A83': '=IF(Inputs!$B$7+82-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A84': '=IF(Inputs!$B$7+83-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A85': '=IF(Inputs!$B$7+84-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A86': '=IF(Inputs!$B$7+85-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A87': '=IF(Inputs!$B$7+86-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A88': '=IF(Inputs!$B$7+87-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A89': '=IF(Inputs!$B$7+88-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A9': '=IF(Inputs!$B$7+8-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A90': '=IF(Inputs!$B$7+89-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A91': '=IF(Inputs!$B$7+90-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A92': '=IF(Inputs!$B$7+91-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A93': '=IF(Inputs!$B$7+92-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A94': '=IF(Inputs!$B$7+93-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A95': '=IF(Inputs!$B$7+94-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A96': '=IF(Inputs!$B$7+95-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A97': '=IF(Inputs!$B$7+96-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A98': '=IF(Inputs!$B$7+97-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'A99': '=IF(Inputs!$B$7+98-1>MAX(MORTALITY!$A:$A),"","WL01")',
        'I2': '=IF(Inputs!$B$7+1-1>MAX(MORTALITY!$A:$A),"",0)',
    },
    'col_format': {
        'D': '0.00',
        'E': '0.00000_)',
        'G': '0%',
    },
}

# ------------------------------------------------------------------------
# PUA_TABLE
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['PUA_TABLE'] = {
    'dims': (103, 10),
    'statics': {
        'A3': 'Issue Age',
        'B3': 'Duration',
        'C3': 'Attained Age',
        'D3': 'PUA SP per $1 SA (Ax)',
        'E3': 'Dx',
        'F3': 'Nx+1',
        'G3': 'Mx',
        'H3': 'Beta (NLP/1000)',
        'I3': 'PUA CVF per 1000 (Fackler)',
    },
    'col_templates': {
        'A': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",Inputs!$B$7)', 'ranges': [(4, 103)]},
        ],
        'B': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",{R-3})', 'ranges': [(4, 103)]},
        ],
        'C': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",Inputs!$B$7+{R-3}-1)', 'ranges': [(4, 103)]},
        ],
        'D': [
            {'template': '=IF(C{R+0}="","",INDEX(Commutation!$K:$K,MATCH(C{R+0},Commutation!$A:$A,0)))', 'ranges': [(4, 103)]},
        ],
        'E': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",IFERROR(INDEX(Commutation!$F:$F,MATCH(C{R+0},Commutation!$A:$A,0)),0))', 'ranges': [(4, 103)]},
        ],
        'F': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",IFERROR(INDEX(Commutation!$H:$H,MATCH(C{R+0}+1,Commutation!$A:$A,0)),0))', 'ranges': [(4, 103)]},
        ],
        'G': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",IFERROR(INDEX(Commutation!$I:$I,MATCH(C{R+0},Commutation!$A:$A,0)),0))', 'ranges': [(4, 103)]},
        ],
        'H': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",IFERROR(1000*INDEX(Commutation!$I:$I,MATCH(Inputs!$B$7,Commutation!$A:$A,0))/INDEX(Commutation!$H:$H,MATCH(Inputs!$B$7,Commutation!$A:$A,0)),0))', 'ranges': [(4, 103)]},
        ],
        'I': [
            {'template': '=IF(Inputs!$B$7+{R-3}-1>MAX(MORTALITY!$A:$A),"",IFERROR((1000*G{R+0}-H{R+0}*F{R+0})/E{R+0},0))', 'ranges': [(4, 103)]},
        ],
    },
    'explicit': {
    },
    'col_format': {
        'C': '0',
        'D': '0.000',
        'E': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'F': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'G': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'H': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
        'I': '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)',
    },
}

# ------------------------------------------------------------------------
# TERM_RATES
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['TERM_RATES'] = {
    'dims': (123, 5),
    'statics': {
        'A1': 'Annual Term Cost per $1 of SA - by Attained Age',
        'A2': 'Age',
        'A3': 0,
        'B2': 'Term Cost (via recursion method)',
        'C2': 'qx',
        'D2': 'v*qx',
        'E2': 'v*q(x+1)*px',
        'E3': 1,
    },
    'col_templates': {
        'A': [
            {'template': '=IF(A{R-1}>=MAX(MORTALITY!$A:$A),"",A{R-1}+1)', 'ranges': [(4, 123)]},
        ],
        'B': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(C{R+0}=1,ETI!$B$7,ETI!$B$7*C{R+0}+ETI!$B$7*(1-C{R+0})*B{R+1}),""))', 'ranges': [(4, 123)]},
        ],
        'C': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(Inputs!$B$6="Male",INDEX(MORTALITY!$B:$B,MATCH(A{R+0},MORTALITY!$A:$A,0)),INDEX(MORTALITY!$E:$E,MATCH(A{R+0},MORTALITY!$D:$D,0))),""))', 'ranges': [(4, 123)]},
        ],
        'D': [
            {'template': '=IF(A{R+0}="","",IFERROR(ETI!$B$7*C{R+0},""))', 'ranges': [(4, 123)]},
        ],
        'E': [
            {'template': '=IF(A{R+0}="","",D{R+0}*(1-C{R-1}))', 'ranges': [(4, 123)]},
        ],
    },
    'explicit': {
        'B3': '=IFERROR(IF(C3=1,ETI!$B$7,ETI!$B$7*C3+ETI!$B$7*(1-C3)*B4),"")',
        'C3': '=IFERROR(IF(Inputs!$B$6="Male",INDEX(MORTALITY!$B:$B,MATCH(A3,MORTALITY!$A:$A,0)),INDEX(MORTALITY!$E:$E,MATCH(A3,MORTALITY!$D:$D,0))),"")',
        'D3': '=IFERROR(ETI!$B$7*C3,"")',
    },
    'col_format': {
        'C': '_(* #,##0.000000_);_(* \\(#,##0.000000\\);_(* \\-??_);_(@_)',
        'E': '0.000000',
    },
}

# ------------------------------------------------------------------------
# DIvidend_Assumption
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['DIvidend_Assumption'] = {
    'dims': (103, 102),
    'statics': {
        'A2': 'Duration',
        'A3': 0,
        'B1': 'Age',
        'B2': 0,
    },
    'col_templates': {
        'A': [
            {'template': '=A{R-1}+1', 'ranges': [(4, 103)]},
        ],
        'AA': [
            {'template': '=Z{R+1}', 'ranges': [(3, 103)]},
        ],
        'AB': [
            {'template': '=AA{R+1}', 'ranges': [(3, 103)]},
        ],
        'AC': [
            {'template': '=AB{R+1}', 'ranges': [(3, 103)]},
        ],
        'AD': [
            {'template': '=AC{R+1}', 'ranges': [(3, 103)]},
        ],
        'AE': [
            {'template': '=AD{R+1}', 'ranges': [(3, 103)]},
        ],
        'AF': [
            {'template': '=AE{R+1}', 'ranges': [(3, 103)]},
        ],
        'AG': [
            {'template': '=AF{R+1}', 'ranges': [(3, 103)]},
        ],
        'AH': [
            {'template': '=AG{R+1}', 'ranges': [(3, 103)]},
        ],
        'AI': [
            {'template': '=AH{R+1}', 'ranges': [(3, 103)]},
        ],
        'AJ': [
            {'template': '=AI{R+1}', 'ranges': [(3, 103)]},
        ],
        'AK': [
            {'template': '=AJ{R+1}', 'ranges': [(3, 103)]},
        ],
        'AL': [
            {'template': '=AK{R+1}', 'ranges': [(3, 103)]},
        ],
        'AM': [
            {'template': '=AL{R+1}', 'ranges': [(3, 103)]},
        ],
        'AN': [
            {'template': '=AM{R+1}', 'ranges': [(3, 103)]},
        ],
        'AO': [
            {'template': '=AN{R+1}', 'ranges': [(3, 103)]},
        ],
        'AP': [
            {'template': '=AO{R+1}', 'ranges': [(3, 103)]},
        ],
        'AQ': [
            {'template': '=AP{R+1}', 'ranges': [(3, 103)]},
        ],
        'AR': [
            {'template': '=AQ{R+1}', 'ranges': [(3, 103)]},
        ],
        'AS': [
            {'template': '=AR{R+1}', 'ranges': [(3, 103)]},
        ],
        'AT': [
            {'template': '=AS{R+1}', 'ranges': [(3, 103)]},
        ],
        'AU': [
            {'template': '=AT{R+1}', 'ranges': [(3, 103)]},
        ],
        'AV': [
            {'template': '=AU{R+1}', 'ranges': [(3, 103)]},
        ],
        'AW': [
            {'template': '=AV{R+1}', 'ranges': [(3, 103)]},
        ],
        'AX': [
            {'template': '=AW{R+1}', 'ranges': [(3, 103)]},
        ],
        'AY': [
            {'template': '=AX{R+1}', 'ranges': [(3, 103)]},
        ],
        'AZ': [
            {'template': '=AY{R+1}', 'ranges': [(3, 103)]},
        ],
        'B': [
            {'template': '=B{R-1}', 'ranges': [(4, 103)]},
        ],
        'BA': [
            {'template': '=AZ{R+1}', 'ranges': [(3, 103)]},
        ],
        'BB': [
            {'template': '=BA{R+1}', 'ranges': [(3, 103)]},
        ],
        'BC': [
            {'template': '=BB{R+1}', 'ranges': [(3, 103)]},
        ],
        'BD': [
            {'template': '=BC{R+1}', 'ranges': [(3, 103)]},
        ],
        'BE': [
            {'template': '=BD{R+1}', 'ranges': [(3, 103)]},
        ],
        'BF': [
            {'template': '=BE{R+1}', 'ranges': [(3, 103)]},
        ],
        'BG': [
            {'template': '=BF{R+1}', 'ranges': [(3, 103)]},
        ],
        'BH': [
            {'template': '=BG{R+1}', 'ranges': [(3, 103)]},
        ],
        'BI': [
            {'template': '=BH{R+1}', 'ranges': [(3, 103)]},
        ],
        'BJ': [
            {'template': '=BI{R+1}', 'ranges': [(3, 103)]},
        ],
        'BK': [
            {'template': '=BJ{R+1}', 'ranges': [(3, 103)]},
        ],
        'BL': [
            {'template': '=BK{R+1}', 'ranges': [(3, 103)]},
        ],
        'BM': [
            {'template': '=BL{R+1}', 'ranges': [(3, 103)]},
        ],
        'BN': [
            {'template': '=BM{R+1}', 'ranges': [(3, 103)]},
        ],
        'BO': [
            {'template': '=BN{R+1}', 'ranges': [(3, 103)]},
        ],
        'BP': [
            {'template': '=BO{R+1}', 'ranges': [(3, 103)]},
        ],
        'BQ': [
            {'template': '=BP{R+1}', 'ranges': [(3, 103)]},
        ],
        'BR': [
            {'template': '=BQ{R+1}', 'ranges': [(3, 103)]},
        ],
        'BS': [
            {'template': '=BR{R+1}', 'ranges': [(3, 103)]},
        ],
        'BT': [
            {'template': '=BS{R+1}', 'ranges': [(3, 103)]},
        ],
        'BU': [
            {'template': '=BT{R+1}', 'ranges': [(3, 103)]},
        ],
        'BV': [
            {'template': '=BU{R+1}', 'ranges': [(3, 103)]},
        ],
        'BW': [
            {'template': '=BV{R+1}', 'ranges': [(3, 103)]},
        ],
        'BX': [
            {'template': '=BW{R+1}', 'ranges': [(3, 103)]},
        ],
        'BY': [
            {'template': '=BX{R+1}', 'ranges': [(3, 103)]},
        ],
        'BZ': [
            {'template': '=BY{R+1}', 'ranges': [(3, 103)]},
        ],
        'C': [
            {'template': '=B{R+1}', 'ranges': [(3, 103)]},
        ],
        'CA': [
            {'template': '=BZ{R+1}', 'ranges': [(3, 103)]},
        ],
        'CB': [
            {'template': '=CA{R+1}', 'ranges': [(3, 103)]},
        ],
        'CC': [
            {'template': '=CB{R+1}', 'ranges': [(3, 103)]},
        ],
        'CD': [
            {'template': '=CC{R+1}', 'ranges': [(3, 103)]},
        ],
        'CE': [
            {'template': '=CD{R+1}', 'ranges': [(3, 103)]},
        ],
        'CF': [
            {'template': '=CE{R+1}', 'ranges': [(3, 103)]},
        ],
        'CG': [
            {'template': '=CF{R+1}', 'ranges': [(3, 103)]},
        ],
        'CH': [
            {'template': '=CG{R+1}', 'ranges': [(3, 103)]},
        ],
        'CI': [
            {'template': '=CH{R+1}', 'ranges': [(3, 103)]},
        ],
        'CJ': [
            {'template': '=CI{R+1}', 'ranges': [(3, 103)]},
        ],
        'CK': [
            {'template': '=CJ{R+1}', 'ranges': [(3, 103)]},
        ],
        'CL': [
            {'template': '=CK{R+1}', 'ranges': [(3, 103)]},
        ],
        'CM': [
            {'template': '=CL{R+1}', 'ranges': [(3, 103)]},
        ],
        'CN': [
            {'template': '=CM{R+1}', 'ranges': [(3, 103)]},
        ],
        'CO': [
            {'template': '=CN{R+1}', 'ranges': [(3, 103)]},
        ],
        'CP': [
            {'template': '=CO{R+1}', 'ranges': [(3, 103)]},
        ],
        'CQ': [
            {'template': '=CP{R+1}', 'ranges': [(3, 103)]},
        ],
        'CR': [
            {'template': '=CQ{R+1}', 'ranges': [(3, 103)]},
        ],
        'CS': [
            {'template': '=CR{R+1}', 'ranges': [(3, 103)]},
        ],
        'CT': [
            {'template': '=CS{R+1}', 'ranges': [(3, 103)]},
        ],
        'CU': [
            {'template': '=CT{R+1}', 'ranges': [(3, 103)]},
        ],
        'CV': [
            {'template': '=CU{R+1}', 'ranges': [(3, 103)]},
        ],
        'CW': [
            {'template': '=CV{R+1}', 'ranges': [(3, 103)]},
        ],
        'CX': [
            {'template': '=CW{R+1}', 'ranges': [(3, 103)]},
        ],
        'D': [
            {'template': '=C{R+1}', 'ranges': [(3, 103)]},
        ],
        'E': [
            {'template': '=D{R+1}', 'ranges': [(3, 103)]},
        ],
        'F': [
            {'template': '=E{R+1}', 'ranges': [(3, 103)]},
        ],
        'G': [
            {'template': '=F{R+1}', 'ranges': [(3, 103)]},
        ],
        'H': [
            {'template': '=G{R+1}', 'ranges': [(3, 103)]},
        ],
        'I': [
            {'template': '=H{R+1}', 'ranges': [(3, 103)]},
        ],
        'J': [
            {'template': '=I{R+1}', 'ranges': [(3, 103)]},
        ],
        'K': [
            {'template': '=J{R+1}', 'ranges': [(3, 103)]},
        ],
        'L': [
            {'template': '=K{R+1}', 'ranges': [(3, 103)]},
        ],
        'M': [
            {'template': '=L{R+1}', 'ranges': [(3, 103)]},
        ],
        'N': [
            {'template': '=M{R+1}', 'ranges': [(3, 103)]},
        ],
        'O': [
            {'template': '=N{R+1}', 'ranges': [(3, 103)]},
        ],
        'P': [
            {'template': '=O{R+1}', 'ranges': [(3, 103)]},
        ],
        'Q': [
            {'template': '=P{R+1}', 'ranges': [(3, 103)]},
        ],
        'R': [
            {'template': '=Q{R+1}', 'ranges': [(3, 103)]},
        ],
        'S': [
            {'template': '=R{R+1}', 'ranges': [(3, 103)]},
        ],
        'T': [
            {'template': '=S{R+1}', 'ranges': [(3, 103)]},
        ],
        'U': [
            {'template': '=T{R+1}', 'ranges': [(3, 103)]},
        ],
        'V': [
            {'template': '=U{R+1}', 'ranges': [(3, 103)]},
        ],
        'W': [
            {'template': '=V{R+1}', 'ranges': [(3, 103)]},
        ],
        'X': [
            {'template': '=W{R+1}', 'ranges': [(3, 103)]},
        ],
        'Y': [
            {'template': '=X{R+1}', 'ranges': [(3, 103)]},
        ],
        'Z': [
            {'template': '=Y{R+1}', 'ranges': [(3, 103)]},
        ],
    },
    'explicit': {
        'AA2': '=+Z2+1',
        'AB2': '=+AA2+1',
        'AC2': '=+AB2+1',
        'AD2': '=+AC2+1',
        'AE2': '=+AD2+1',
        'AF2': '=+AE2+1',
        'AG2': '=+AF2+1',
        'AH2': '=+AG2+1',
        'AI2': '=+AH2+1',
        'AJ2': '=+AI2+1',
        'AK2': '=+AJ2+1',
        'AL2': '=+AK2+1',
        'AM2': '=+AL2+1',
        'AN2': '=+AM2+1',
        'AO2': '=+AN2+1',
        'AP2': '=+AO2+1',
        'AQ2': '=+AP2+1',
        'AR2': '=+AQ2+1',
        'AS2': '=+AR2+1',
        'AT2': '=+AS2+1',
        'AU2': '=+AT2+1',
        'AV2': '=+AU2+1',
        'AW2': '=+AV2+1',
        'AX2': '=+AW2+1',
        'AY2': '=+AX2+1',
        'AZ2': '=+AY2+1',
        'B3': '=Inputs!$B$18',
        'BA2': '=+AZ2+1',
        'BB2': '=+BA2+1',
        'BC2': '=+BB2+1',
        'BD2': '=+BC2+1',
        'BE2': '=+BD2+1',
        'BF2': '=+BE2+1',
        'BG2': '=+BF2+1',
        'BH2': '=+BG2+1',
        'BI2': '=+BH2+1',
        'BJ2': '=+BI2+1',
        'BK2': '=+BJ2+1',
        'BL2': '=+BK2+1',
        'BM2': '=+BL2+1',
        'BN2': '=+BM2+1',
        'BO2': '=+BN2+1',
        'BP2': '=+BO2+1',
        'BQ2': '=+BP2+1',
        'BR2': '=+BQ2+1',
        'BS2': '=+BR2+1',
        'BT2': '=+BS2+1',
        'BU2': '=+BT2+1',
        'BV2': '=+BU2+1',
        'BW2': '=+BV2+1',
        'BX2': '=+BW2+1',
        'BY2': '=+BX2+1',
        'BZ2': '=+BY2+1',
        'C2': '=+B2+1',
        'CA2': '=+BZ2+1',
        'CB2': '=+CA2+1',
        'CC2': '=+CB2+1',
        'CD2': '=+CC2+1',
        'CE2': '=+CD2+1',
        'CF2': '=+CE2+1',
        'CG2': '=+CF2+1',
        'CH2': '=+CG2+1',
        'CI2': '=+CH2+1',
        'CJ2': '=+CI2+1',
        'CK2': '=+CJ2+1',
        'CL2': '=+CK2+1',
        'CM2': '=+CL2+1',
        'CN2': '=+CM2+1',
        'CO2': '=+CN2+1',
        'CP2': '=+CO2+1',
        'CQ2': '=+CP2+1',
        'CR2': '=+CQ2+1',
        'CS2': '=+CR2+1',
        'CT2': '=+CS2+1',
        'CU2': '=+CT2+1',
        'CV2': '=+CU2+1',
        'CW2': '=+CV2+1',
        'CX2': '=+CW2+1',
        'D2': '=+C2+1',
        'E2': '=+D2+1',
        'F2': '=+E2+1',
        'G2': '=+F2+1',
        'H2': '=+G2+1',
        'I2': '=+H2+1',
        'J2': '=+I2+1',
        'K2': '=+J2+1',
        'L2': '=+K2+1',
        'M2': '=+L2+1',
        'N2': '=+M2+1',
        'O2': '=+N2+1',
        'P2': '=+O2+1',
        'Q2': '=+P2+1',
        'R2': '=+Q2+1',
        'S2': '=+R2+1',
        'T2': '=+S2+1',
        'U2': '=+T2+1',
        'V2': '=+U2+1',
        'W2': '=+V2+1',
        'X2': '=+W2+1',
        'Y2': '=+X2+1',
        'Z2': '=+Y2+1',
    },
    'col_format': {
        'AA': '0.0%',
        'AB': '0.0%',
        'AC': '0.0%',
        'AD': '0.0%',
        'AE': '0.0%',
        'AF': '0.0%',
        'AG': '0.0%',
        'AH': '0.0%',
        'AI': '0.0%',
        'AJ': '0.0%',
        'AK': '0.0%',
        'AL': '0.0%',
        'AM': '0.0%',
        'AN': '0.0%',
        'AO': '0.0%',
        'AP': '0.0%',
        'AQ': '0.0%',
        'AR': '0.0%',
        'AS': '0.0%',
        'AT': '0.0%',
        'AU': '0.0%',
        'AV': '0.0%',
        'AW': '0.0%',
        'AX': '0.0%',
        'AY': '0.0%',
        'AZ': '0.0%',
        'B': '0.0%',
        'BA': '0.0%',
        'BB': '0.0%',
        'BC': '0.0%',
        'BD': '0.0%',
        'BE': '0.0%',
        'BF': '0.0%',
        'BG': '0.0%',
        'BH': '0.0%',
        'BI': '0.0%',
        'BJ': '0.0%',
        'BK': '0.0%',
        'BL': '0.0%',
        'BM': '0.0%',
        'BN': '0.0%',
        'BO': '0.0%',
        'BP': '0.0%',
        'BQ': '0.0%',
        'BR': '0.0%',
        'BS': '0.0%',
        'BT': '0.0%',
        'BU': '0.0%',
        'BV': '0.0%',
        'BW': '0.0%',
        'BX': '0.0%',
        'BY': '0.0%',
        'BZ': '0.0%',
        'C': '0.0%',
        'CA': '0.0%',
        'CB': '0.0%',
        'CC': '0.0%',
        'CD': '0.0%',
        'CE': '0.0%',
        'CF': '0.0%',
        'CG': '0.0%',
        'CH': '0.0%',
        'CI': '0.0%',
        'CJ': '0.0%',
        'CK': '0.0%',
        'CL': '0.0%',
        'CM': '0.0%',
        'CN': '0.0%',
        'CO': '0.0%',
        'CP': '0.0%',
        'CQ': '0.0%',
        'CR': '0.0%',
        'CS': '0.0%',
        'CT': '0.0%',
        'CU': '0.0%',
        'CV': '0.0%',
        'CW': '0.0%',
        'CX': '0.0%',
        'D': '0.0%',
        'E': '0.0%',
        'F': '0.0%',
        'G': '0.0%',
        'H': '0.0%',
        'I': '0.0%',
        'J': '0.0%',
        'K': '0.0%',
        'L': '0.0%',
        'M': '0.0%',
        'N': '0.0%',
        'O': '0.0%',
        'P': '0.0%',
        'Q': '0.0%',
        'R': '0.0%',
        'S': '0.0%',
        'T': '0.0%',
        'U': '0.0%',
        'V': '0.0%',
        'W': '0.0%',
        'X': '0.0%',
        'Y': '0.0%',
        'Z': '0.0%',
    },
}

# ------------------------------------------------------------------------
# Projection_Monthly
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['Projection_Monthly'] = {
    'dims': (988, 28),
    'statics': {
        'A1': 'Plan',
        'AA1': 'CV Total (Incl PUA)',
        'AB1': 'Premium Offset Applied (Div + Accum Div growth)',
        'B1': 'Date',
        'C1': 'Month Index',
        'D1': 'Duration',
        'E1': 'Attained Age',
        'F1': 'Last Anniv',
        'G1': 'Next Anniv',
        'H1': 'Frac within Year',
        'I1': 'CV Factor_n',
        'J1': 'CV Factor_n+1',
        'K1': 'CV per 1000',
        'L1': 'CV Total (Base)',
        'M1': 'Pre-Dividend PUA SA (cum)',
        'N1': 'Dividend Base SA',
        'O1': 'Monthly Div Rate',
        'P1': 'Dividend Amount',
        'Q1': 'Accum Div ',
        'R1': 'Accum Div Prev End',
        'S1': 'Accum Div Prev Growth',
        'T1': 'Out-of-Pocket Premium',
        'U1': 'Cash Dividend Paid',
        'V1': 'Dividend to PUA',
        'W1': 'PUA SP per $1 SA',
        'X1': 'PUA SA Add (this month)',
        'Y1': 'PUA SA (cum)',
        'Z1': 'CV from PUA',
    },
    'col_templates': {
        'A': [
            {'template': '=IF(A{R-1}="","",IF(INT((C{R-1}+1)/12)+Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4))', 'ranges': [(3, 985)]},
        ],
        'AA': [
            {'template': '=IF(A{R+0}="","",IFERROR(L{R+0}+Z{R+0},0))', 'ranges': [(2, 985)]},
        ],
        'AB': [
            {'template': '=IF(A{R+0}="","",IF(Inputs!$B$19="Premium Offset",MIN(Inputs!$B$22,P{R+0}+S{R+0}),0))', 'ranges': [(2, 985)]},
        ],
        'B': [
            {'template': '=IF(A{R+0}="","",EOMONTH(B{R-1},1))', 'ranges': [(3, 985)]},
        ],
        'C': [
            {'template': '=IF(A{R+0}="","",C{R-1}+1)', 'ranges': [(3, 985)]},
        ],
        'D': [
            {'template': '=IF(A{R+0}="","",INT(C{R+0}/12)+1)', 'ranges': [(2, 985)]},
        ],
        'E': [
            {'template': '=IF(A{R+0}="","",D{R+0}+Inputs!$B$7-1)', 'ranges': [(2, 985)]},
        ],
        'F': [
            {'template': '=IF(A{R+0}="","",DATE(YEAR(B$2)+D{R+0}-1,MONTH(B$2),DAY(B$2)))', 'ranges': [(2, 985)]},
        ],
        'G': [
            {'template': '=IF(A{R+0}="","",DATE(YEAR(F{R+0})+1,MONTH(F{R+0}),DAY(F{R+0})))', 'ranges': [(2, 985)]},
        ],
        'H': [
            {'template': '=IF(A{R+0}="","",IF(G{R+0}>F{R+0},(B{R+0}-F{R+0})/(G{R+0}-F{R+0}),0))', 'ranges': [(2, 985)]},
        ],
        'I': [
            {'template': '=IF(A{R+0}="","",IFERROR(INDEX(CV_RATES!$I:$I,MATCH(D{R+0},CV_RATES!$C:$C,0)),0))', 'ranges': [(2, 985)]},
        ],
        'J': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(E{R+0}=MAX(MORTALITY!$A:$A),1000,INDEX(CV_RATES!$J:$J,MATCH(D{R+0},CV_RATES!$C:$C,0))),0))', 'ranges': [(2, 985)]},
        ],
        'K': [
            {'template': '=IF(A{R+0}="","",IFERROR(I{R+0}+H{R+0}*(J{R+0}-I{R+0}),0))', 'ranges': [(2, 985)]},
        ],
        'L': [
            {'template': '=IF(A{R+0}="","",IFERROR(K{R+0}*Inputs!$B$9,0))', 'ranges': [(2, 985)]},
        ],
        'M': [
            {'template': '=IF(A{R+0}="","",Y{R-1})', 'ranges': [(3, 985)]},
        ],
        'N': [
            {'template': '=IF(A{R+0}="","",Inputs!$B$17+M{R+0})', 'ranges': [(2, 985)]},
        ],
        'O': [
            {'template': '=IF(A{R+0}="","",VLOOKUP(D{R+0},DIvidend_Assumption!$A$3:$CX$103,($E$2+2),0)/12)', 'ranges': [(2, 985)]},
        ],
        'P': [
            {'template': '=IF(A{R+0}="","",IFERROR(AA{R-1}*O{R+0},""))', 'ranges': [(3, 985)]},
        ],
        'Q': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(Inputs!$B$19="Accumulate",S{R+0}+P{R+0},IF(Inputs!$B$19="Premium Offset",MAX(0,S{R+0}+P{R+0}-Inputs!$B$22),S{R+0})),""))', 'ranges': [(2, 985)]},
        ],
        'R': [
            {'template': '=IF(A{R+0}="","",Q{R-1})', 'ranges': [(3, 985)]},
        ],
        'S': [
            {'template': '=IF(A{R+0}="","",IFERROR(R{R+0}*(1+Inputs!$B$20/12),""))', 'ranges': [(2, 985)]},
        ],
        'T': [
            {'template': '=IF(A{R+0}="","",IF(D{R+0}<=Inputs!$B$8,IF(Inputs!$B$19="Premium Offset",MAX(0,Inputs!$B$22-(P{R+0}+S{R+0})),Inputs!$B$22),0))', 'ranges': [(2, 985)]},
        ],
        'U': [
            {'template': '=IF(A{R+0}="","",IF(Inputs!$B$19="Cash",P{R+0},0))', 'ranges': [(2, 985)]},
        ],
        'V': [
            {'template': '=IF(A{R+0}="","",IF(Inputs!$B$19="PUA",P{R+0},0))', 'ranges': [(2, 985)]},
        ],
        'W': [
            {'template': '=IF(A{R+0}="","",IFERROR(INDEX(PUA_TABLE!$D$4:$D$105,MATCH(E{R+0},PUA_TABLE!$C$4:$C$105,1)),0))', 'ranges': [(2, 985)]},
        ],
        'X': [
            {'template': '=IF(A{R+0}="","",IF(V{R+0}>0,IF(W{R+0}>0,V{R+0}/W{R+0},0),0))', 'ranges': [(2, 985)]},
        ],
        'Y': [
            {'template': '=IF(A{R+0}="","",M{R+0}+X{R+0})', 'ranges': [(2, 985)]},
        ],
        'Z': [
            {'template': '=IF(A{R+0}="","",IFERROR(Y{R+0}*INDEX(PUA_TABLE!$I$4:$I$105,MATCH(E{R+0},PUA_TABLE!$C$4:$C$105,1))/1000,0))', 'ranges': [(3, 985)]},
        ],
    },
    'explicit': {
        'A2': '=IF(Inputs!$B$7>MAX(MORTALITY!$A:$A),"",Inputs!$B$4)',
        'B2': '=IF(A2="","",Inputs!$B$11)',
        'C2': '=IF(A2="","",0)',
        'M2': '=IF(A2="","",0)',
        'P2': '=IF(A2="","",0)',
        'R2': '=IF(A2="","",0)',
        'Z2': '=IF(A2="","",IFERROR(V2,0))',
    },
    'col_format': {
        'AA': '0.00',
        'AB': '0.00',
        'B': 'mm-dd-yy',
        'F': 'mm-dd-yy',
        'G': 'mm-dd-yy',
        'H': '0.00',
        'I': '0.00',
        'J': '0.00',
        'K': '0.00',
        'L': '0.00',
        'M': '0.00',
        'N': '0',
        'O': '0.00%',
        'P': '0.00',
        'Q': '0.00',
        'R': '0.00',
        'S': '0.00',
        'T': '0',
        'U': '0.00',
        'V': '0.00',
        'W': '0.000',
        'X': '0.00',
        'Y': '0.00',
        'Z': '0.00',
    },
}

# ------------------------------------------------------------------------
# RPU
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['RPU'] = {
    'dims': (24, 2),
    'statics': {
        'A1': 'Reduced Paid-Up (RPU) - Actuarial Calculation',
        'A10': 'RPU SA (Uncapped) = CV / Ax',
        'A12': '═══ Method 2: Reserve Ratio (Actuarial Formula) ═══',
        'A13': 'Reserve per $1000 at Duration t',
        'A14': 'RPU SA = SA × (tVx/1000) / Ax',
        'A16': '═══════════════════════════════════════════════',
        'A17': 'FINAL RPU Sum Assured',
        'A19': 'Comparison:',
        'A20': 'Is Capping Applied?',
        'A22': 'Note: When CV/Ax > SA, capping prevents over-insurance.',
        'A23': 'Both methods should yield the same result when using',
        'A24': 'consistent actuarial assumptions.',
        'A3': 'Status',
        'A4': 'Attained Age',
        'A5': 'Original Sum Assured',
        'A7': '═══ Method 1: Cash Value / Net Single Premium ═══',
        'A8': 'Gross Cash Value (tVx)',
        'A9': 'Net Single Premium at Attained Age (Ax)',
    },
    'col_templates': {
    },
    'explicit': {
        'B10': '=IF(Inputs!$B$27=FALSE,"",IFERROR(B8/B9,""))',
        'B13': '=IF(Inputs!$B$27=FALSE,"",INDEX(CV_RATES!$J:$J,MATCH(Inputs!$B$15,CV_RATES!$C:$C,0)))',
        'B14': '=IF(Inputs!$B$27=FALSE,"",IFERROR(B5*(B13/1000)/B9,""))',
        'B17': '=IF(Inputs!$B$27=FALSE,"",MIN(B10,B5))',
        'B20': '=IF(Inputs!$B$27=FALSE,"",IF(B10>B5,"YES - Calculated RPU exceeds Original SA","NO - Within limits"))',
        'B3': '=IF(Inputs!$B$27=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")',
        'B4': '=IF(Inputs!$B$27=FALSE,"",Inputs!$B$14)',
        'B5': '=IF(Inputs!$B$27=FALSE,"",Inputs!$B$17)',
        'B8': '=IF(Inputs!$B$27=FALSE,"",INDEX(CV_RATES!$J:$J,MATCH(Inputs!$B$15,CV_RATES!$C:$C,0))/1000*Inputs!$B$17)',
        'B9': '=IF(Inputs!$B$27=FALSE,"",INDEX(Commutation!$K:$K,MATCH(Inputs!$B$14,Commutation!$A:$A,0)))',
    },
    'col_format': {
    },
}

# ------------------------------------------------------------------------
# ETI
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['ETI'] = {
    'dims': (51, 6),
    'statics': {
        'A1': 'Extended Term Insurance (ETI) - Calculations',
        'A3': 'Discount Rate',
        'A4': 'Face Amount',
        'A5': 'Gross Cash Value',
        'A6': 'Attained Age at Val Date',
        'A7': 'v (1/(1+i))',
        'A8': 'Policy Year of ETI',
        'B3': 0.05,
        'B8': 'Age',
        'C8': 'Term Cost per $1 SA for 1 Year',
        'D2': 'Status',
        'D3': 'ETI Years (full)',
        'D4': 'Partial Year (fraction)',
        'D5': 'ETI Duration (years)',
        'D6': 'ETI Expiry Date',
        'D8': 'Cash Value required for Base face Amt for 1 Year',
        'E8': 'Remaining CV after cumulative ETI cost',
        'F8': 'Indicator',
    },
    'col_templates': {
        'A': [
            {'template': '=IF(OR(A{R-1}="",B{R-1}>=MAX(MORTALITY!$A:$A)),"",A{R-1}+1)', 'ranges': [(10, 51)]},
        ],
        'B': [
            {'template': '=IF(OR(B{R-1}="",B{R-1}>=MAX(MORTALITY!$A:$A)),"",B{R-1}+1)', 'ranges': [(10, 51)]},
        ],
        'C': [
            {'template': '=IF(A{R+0}="","",IFERROR(INDEX(TERM_RATES!$E:$E,MATCH(B{R+0},TERM_RATES!$A:$A,0)),""))', 'ranges': [(9, 51)]},
        ],
        'D': [
            {'template': '=IF(A{R+0}="","",IFERROR(C{R+0}*$B$4,""))', 'ranges': [(9, 51)]},
        ],
        'E': [
            {'template': '=IF(A{R+0}="","",IFERROR($B$5-SUM($D$9:D{R+0}),""))', 'ranges': [(10, 51)]},
        ],
        'F': [
            {'template': '=IF(A{R+0}="","",IF(AND(E{R+1}<0,E{R+0}>0),1,0))', 'ranges': [(9, 50)]},
        ],
    },
    'explicit': {
        'A9': '=IF(Inputs!$B$27=FALSE,"",IF($B$6>MAX(MORTALITY!$A:$A),"",1))',
        'B4': '=IF(Inputs!$B$27=FALSE,"",Inputs!$B$17)',
        'B5': '=IF(Inputs!$B$27=FALSE,"",APL!$B$9)',
        'B6': '=IF(Inputs!$B$27=FALSE,"",Inputs!B14)',
        'B7': '=1/(1+B3)',
        'B9': '=B6',
        'E2': '=IF(Inputs!$B$27=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")',
        'E3': '=IF(Inputs!$B$27=FALSE,"",IFERROR(MATCH(TRUE,E9:E51<0,0)-1,SUMPRODUCT((E9:E51<>"")*1)))',
        'E4': '=IF(Inputs!$B$27=FALSE,"",IFERROR(IF(E3="","",IF(INDEX(E9:E51,E3+1)<0,INDEX(E9:E51,E3)/(INDEX(E9:E51,E3)-INDEX(E9:E51,E3+1)),0)),""))',
        'E5': '=IF(Inputs!$B$27=FALSE,"",IFERROR(IF(OR(E3="",E4=""),"",E3+E4),""))',
        'E6': '=IF(Inputs!$B$27=FALSE,"",IFERROR(IF(E5="","",Inputs!B12+365*E5),""))',
        'E9': '=IF(A9="","",IFERROR($B$5-SUM($D$9),""))',
        'F51': '=IF(A51="","",0)',
    },
    'col_format': {
        'D': '0.00',
        'E': '0.00',
    },
}

# ------------------------------------------------------------------------
# APL
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['APL'] = {
    'dims': (199, 9),
    'statics': {
        'A1': 'Automatic Premium Loan (APL)',
        'A10': 'Loanable % of CV',
        'A11': 'Premium Paying Term (PPT)',
        'A13': 'Period #',
        'A3': 'APL Enabled (TRUE/FALSE)',
        'A4': 'Policy Loan Interest Rate (p.a.)',
        'A5': 'Payments per Year (mode)',
        'A6': 'Valuation Date',
        'A7': 'Annual Premium',
        'A8': 'Modal Premium',
        'A9': 'CV at Valuation Date',
        'B10': 0.8,
        'B13': 'Due Date (approx)',
        'B4': 0.08,
        'C13': 'Beginning Balance',
        'D13': 'Interest',
        'D3': 'Legend',
        'E13': 'APL Drawn',
        'F13': 'Payment Covered by APL',
        'F3': 'Link with Data/Input',
        'F4': 'Formulae',
        'F5': 'Dropdown Selection',
        'G13': 'Ending Balance',
        'G3': 'Status',
        'G4': 'PPT End Date',
        'H13': 'Net CV (CV - Loan)',
        'I13': 'Policy Status',
    },
    'col_templates': {
        'A': [
            {'template': '=IF(OR(A{R-1}="",H{R-1}<=0),"",A{R-1}+1)', 'ranges': [(15, 199)]},
        ],
        'B': [
            {'template': '=IF(A{R+0}="","",EOMONTH(B{R-1},12/$B$5))', 'ranges': [(15, 199)]},
        ],
        'C': [
            {'template': '=IF(A{R+0}="","",G{R-1})', 'ranges': [(15, 199)]},
        ],
        'D': [
            {'template': '=IF(A{R+0}="","",C{R+0}*$B$4/$B$5)', 'ranges': [(15, 199)]},
        ],
        'E': [
            {'template': '=IF(A{R+0}="","",IF(AND(B{R+0}<=$H$4,C{R+0}+D{R+0}+$B$8<VLOOKUP(B{R+0},Projection_Monthly!$B:$AA,26,TRUE)*$B$10),$B$8,0))', 'ranges': [(15, 199)]},
        ],
        'F': [
            {'template': '=IF(A{R+0}="","",MIN($B$8,E{R+0}))', 'ranges': [(15, 199)]},
        ],
        'G': [
            {'template': '=IF(A{R+0}="","",C{R+0}+D{R+0}+E{R+0})', 'ranges': [(15, 199)]},
        ],
        'H': [
            {'template': '=IF(A{R+0}="","",VLOOKUP(B{R+0},Projection_Monthly!$B:$AA,26,TRUE)-G{R+0})', 'ranges': [(15, 199)]},
        ],
        'I': [
            {'template': '=IF(A{R+0}="","",IF(H{R+0}>0,"Active","LAPSED"))', 'ranges': [(15, 199)]},
        ],
    },
    'explicit': {
        'A14': '=IF(Inputs!$B$27=FALSE,"",1)',
        'B11': '=Inputs!$B$8',
        'B14': '=IF(Inputs!$B$27=FALSE,"",B6)',
        'B3': '=TRUE()',
        'B5': '=VLOOKUP(Inputs!B23,Inputs!D20:E23,2,0)',
        'B6': '=Inputs!$B$12',
        'B7': '=IF(Inputs!$B$27=FALSE,"",IFERROR(INDEX(Inputs!B:B,MATCH("Annual Base Premium",Inputs!A:A,0)),0))',
        'B8': '=IF(Inputs!$B$27=FALSE,"",IFERROR(INDEX(Inputs!B:B,MATCH("Modal Premium",Inputs!A:A,0)),IF(B7>0,B7/B5,0)))',
        'B9': '=VLOOKUP(B6,Projection_Monthly!$B$2:$AA$985,26,0)',
        'C14': '=IF(Inputs!$B$27=FALSE,"",0)',
        'D14': '=IF(Inputs!$B$27=FALSE,"",C14*$B$4/$B$5)',
        'E14': '=IF(Inputs!$B$27=FALSE,"",IF(AND(B14<=$H$4,$B$8+C14+D14<$B$9*$B$10),$B$8,0))',
        'F14': '=IF(Inputs!$B$27=FALSE,"",MIN($B$8,E14))',
        'G14': '=IF(Inputs!$B$27=FALSE,"",C14+D14+E14)',
        'H14': '=IF(Inputs!$B$27=FALSE,"",VLOOKUP(B14,Projection_Monthly!$B:$AA,26,TRUE)-G14)',
        'H3': '=IF(Inputs!$B$27=FALSE,"N/A - Policy is already Paid-Up (beyond PPT)","Applicable")',
        'H4': '=EDATE(Inputs!$B$11,Inputs!$B$8*12)',
        'I14': '=IF(Inputs!$B$27=FALSE,"",IF(H14>0,"Active","LAPSED"))',
    },
    'col_format': {
        'B': 'mm-dd-yy',
        'C': '0',
        'D': '0',
        'E': '0',
        'G': '0',
        'H': '0',
    },
}

# ------------------------------------------------------------------------
# Commutation
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['Commutation'] = {
    'dims': (102, 13),
    'statics': {
        'A1': 'Age',
        'B1': 'qx',
        'C1': 'lx',
        'C2': 10000000,
        'D1': 'dx',
        'E1': 'v^x',
        'F1': 'Dx',
        'G1': 'Cx',
        'H1': 'Nx',
        'I1': 'Mx',
        'J1': 'ax_due',
        'K1': 'Ax',
        'L1': 'NLP_per_1000',
        'M1': 'LPP_NLP_per_1000',
    },
    'col_templates': {
        'A': [
            {'template': '=IF(OR(A{R-1}="",A{R-1}>=MAX(MORTALITY!$A:$A)),"",A{R-1}+1)', 'ranges': [(3, 102)]},
        ],
        'B': [
            {'template': '=IF(A{R+0}="","",IFERROR(INDEX(MORTALITY!$B:$B,MATCH(A{R+0},MORTALITY!$A:$A,0)),0))', 'ranges': [(2, 102)]},
        ],
        'C': [
            {'template': '=IF(A{R+0}="","",IFERROR(C{R-1}*(1-B{R-1}),0))', 'ranges': [(3, 102)]},
        ],
        'D': [
            {'template': '=IF(A{R+0}="","",IFERROR(C{R+0}*B{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'E': [
            {'template': '=IF(A{R+0}="","",IFERROR((1/(1+Inputs!$B$20))^A{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'F': [
            {'template': '=IF(A{R+0}="","",IFERROR(C{R+0}*E{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'G': [
            {'template': '=IF(A{R+0}="","",IFERROR(D{R+0}*(1/(1+Inputs!$B$20))^(A{R+0}+1),0))', 'ranges': [(2, 102)]},
        ],
        'H': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(B{R+0}=1,F{R+0},F{R+0}+H{R+1}),0))', 'ranges': [(2, 102)]},
        ],
        'I': [
            {'template': '=IF(A{R+0}="","",IFERROR(IF(B{R+0}=1,G{R+0},G{R+0}+I{R+1}),0))', 'ranges': [(2, 102)]},
        ],
        'J': [
            {'template': '=IF(A{R+0}="","",IFERROR(H{R+0}/F{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'K': [
            {'template': '=IF(A{R+0}="","",IFERROR(I{R+0}/F{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'L': [
            {'template': '=IF(A{R+0}="","",IFERROR(1000*I{R+0}/H{R+0},0))', 'ranges': [(2, 102)]},
        ],
        'M': [
            {'template': '=IF(A{R+0}="","",IFERROR(1000*I{R+0}/(H{R+0}-INDEX(H:H,MATCH(A{R+0}+Inputs!$B$8,$A:$A,0))),0))', 'ranges': [(2, 102)]},
        ],
    },
    'explicit': {
        'A2': '=Inputs!$B$7',
    },
    'col_format': {
        'B': '0.00000_)',
    },
}

# ------------------------------------------------------------------------
# Loan
# ------------------------------------------------------------------------
FUNCTIONAL_SHEETS['Loan'] = {
    'dims': (48, 26),
    'statics': {
        'A1': 'LOAN MODULE - Multiple Loans & Repayments',
        'A10': 'Premium Paying Term (PPT)',
        'A12': 'LOAN TRANSACTIONS (Enter up to 10 loans)',
        'A13': 'Loan #',
        'A14': 1,
        'A15': 2,
        'A16': 3,
        'A17': 4,
        'A18': 5,
        'A19': 6,
        'A20': 7,
        'A21': 8,
        'A22': 9,
        'A23': 10,
        'A26': 'ADVANCE METHOD (Compound Interest - Interest charged at beginning of period)',
        'A27': 'Event #',
        'A3': 'Input Parameters',
        'A4': 'Loan Interest Rate',
        'A41': 'NOTES:',
        'A42': '1. ADVANCE METHOD: Interest is charged at the beginning of each period (prepaid interest).',
        'A43': '2. ARREAR METHOD: Interest is charged at the end of each period (accrued interest).',
        'A44': '3. Both methods use the formula: Interest = Principal × ((1 + Rate)^(Days/365) - 1)',
        'A45': '4. Outstanding balance shown is the amount owed after each transaction/event.',
        'A46': '5. Loan interest is charged daily and compounded at each anniversary or event.',
        'A47': "6. Maximum loan amount typically cannot exceed the policy's cash value.",
        'A48': '7. Anniversaries are automatically calculated from Issue Date for periods between first loan and Valuation Date.',
        'A5': 'Guaranteed Interest Rate',
        'A6': 'Issue Age',
        'A7': 'Issue Date',
        'A8': 'Valuation Date',
        'A9': 'Face Amount (SA)',
        'B13': 'Loan Date',
        'B14': datetime.datetime(2023, 10, 4),
        'B15': datetime.datetime(2024, 10, 4),
        'B16': datetime.datetime(2025, 10, 4),
        'B17': datetime.datetime(2026, 10, 4),
        'B27': 'Effective Date',
        'B4': 0.05,
        'C13': 'Loan Amount',
        'C14': 5765.88,
        'C27': 'Type',
        'D27': 'Amount',
        'E12': 'REPAYMENT TRANSACTIONS (Enter up to 10 repayments)',
        'E13': 'Repay #',
        'E14': 1,
        'E15': 2,
        'E16': 3,
        'E17': 4,
        'E18': 5,
        'E19': 6,
        'E20': 7,
        'E21': 8,
        'E22': 9,
        'E23': 10,
        'E27': 'Days from Prior',
        'E28': 0,
        'F13': 'Repayment Date',
        'F14': datetime.datetime(2026, 10, 4),
        'F27': 'Interest (Advance)',
        'G13': 'Repayment Amount',
        'G14': 319.44,
        'G27': 'Outstanding After',
        'H27': 'Is Anniversary?',
        'J26': 'ARREAR METHOD (Compound Interest - Interest charged at end of period)',
        'J27': 'Event #',
        'K27': 'Date',
        'L27': 'Type',
        'M27': 'Amount',
        'N27': 'Days from Prior',
        'O27': 'Interest Accrued',
        'P27': 'Outstanding After',
        'Q27': 'Is Anniversary?',
        'S1': 'HELPER COLUMNS (can be hidden)',
        'S2': 'Date',
        'S4': datetime.datetime(2024, 10, 4),
        'S5': datetime.datetime(2025, 10, 4),
        'S6': datetime.datetime(2026, 10, 4),
        'T2': 'Type',
        'T4': 'Anni',
        'T5': 'Anni',
        'T6': 'Anni',
        'U2': 'Amount',
        'V2': 'Sort Key',
        'W1': 'SORTED EVENTS',
        'W10': 8,
        'W11': 9,
        'W12': 10,
        'W13': 11,
        'W14': 12,
        'W15': 13,
        'W16': 14,
        'W17': 15,
        'W18': 16,
        'W19': 17,
        'W2': '#',
        'W20': 18,
        'W21': 19,
        'W22': 20,
        'W23': 21,
        'W24': 22,
        'W25': 23,
        'W26': 24,
        'W27': 25,
        'W28': 26,
        'W29': 27,
        'W3': 1,
        'W30': 28,
        'W31': 29,
        'W32': 30,
        'W33': 31,
        'W34': 32,
        'W35': 33,
        'W36': 34,
        'W37': 35,
        'W38': 36,
        'W39': 37,
        'W4': 2,
        'W5': 3,
        'W6': 4,
        'W7': 5,
        'W8': 6,
        'W9': 7,
        'X2': 'Date',
        'Y2': 'Type',
        'Z2': 'Amount',
    },
    'col_templates': {
        'D': [
            {'template': '=G{R-1}', 'ranges': [(29, 32)]},
        ],
        'E': [
            {'template': '=IF(B{R+0}-B{R-1}=366,365,B{R+0}-B{R-1})', 'ranges': [(29, 32)]},
        ],
        'G': [
            {'template': '=G{R-1}+F{R+0}', 'ranges': [(29, 32)]},
        ],
        'H': [
            {'template': '=IF(B{R+0}="","",IF(C{R+0}="ANNIV","Yes",""))', 'ranges': [(28, 40)]},
        ],
        'N': [
            {'template': '=IF(K{R+0}="","",IF(K{R-1}="",0,K{R+0}-K{R-1}))', 'ranges': [(29, 40)]},
        ],
        'O': [
            {'template': '=IF(K{R+0}="","",IFERROR(P{R-1}*((1+$B$4)^(N{R+0}/365)-1),0))', 'ranges': [(29, 40)]},
        ],
        'P': [
            {'template': '=IF(K{R+0}="","",P{R-1}+O{R+0}+IF(L{R+0}="LOAN",M{R+0},IF(L{R+0}="REPAY",-M{R+0},0)))', 'ranges': [(29, 40)]},
        ],
        'Q': [
            {'template': '=IF(K{R+0}="","",IF(L{R+0}="ANNIV","Yes",""))', 'ranges': [(28, 40)]},
        ],
        'S': [
            {'template': '=IF(F{R+0}="","",F{R+0})', 'ranges': [(14, 23)]},
            {'template': '=IFERROR(IF(AND(DATE(YEAR($B$7)+{R-23},MONTH($B$7),DAY($B$7))>_xlfn.MINIFS($B$14:$B$23,$B$14:$B$23,">0"),DATE(YEAR($B$7)+{R-23},MONTH($B$7),DAY($B$7))<=$B$8),DATE(YEAR($B$7)+{R-23},MONTH($B$7),DAY($B$7)),""),"")', 'ranges': [(24, 40)]},
        ],
        'T': [
            {'template': '=IF(F{R+0}="","","REPAY")', 'ranges': [(14, 23)]},
            {'template': '=IF(S{R+0}="","","ANNIV")', 'ranges': [(24, 40)]},
        ],
        'U': [
            {'template': '=IF(G{R+0}="","",G{R+0})', 'ranges': [(14, 23)]},
            {'template': '=IF(S{R+0}="","",0)', 'ranges': [(24, 40)]},
        ],
        'V': [
            {'template': '=IF(F{R+0}="","",F{R+0}+0.2)', 'ranges': [(14, 23)]},
            {'template': '=IF(S{R+0}="","",S{R+0})', 'ranges': [(24, 40)]},
        ],
        'Y': [
            {'template': '=IF(X{R+0}="","",INDEX($T$3:$T$64,MATCH(X{R+0},$S$3:$S$64,0)))', 'ranges': [(3, 9), (11, 40)]},
        ],
        'Z': [
            {'template': '=IF(X{R+0}="","",INDEX($U$3:$U$64,MATCH(X{R+0},$S$3:$S$64,0)))', 'ranges': [(3, 9), (11, 40)]},
        ],
    },
    'explicit': {
        'A28': '=IF(X3="","",W3)',
        'A29': '=IF(X4="","",W4)',
        'A30': '=IF(X5="","",W5)',
        'A31': '=IF(X6="","",W6)',
        'A32': '=IF(X7="","",W7)',
        'A33': '=IF(X8="","",W8)',
        'A34': '=IF(X9="","",W9)',
        'A35': '=IF(X11="","",W10)',
        'A36': '=IF(X12="","",W11)',
        'A37': '=IF(X13="","",W12)',
        'A38': '=IF(X14="","",W13)',
        'A39': '=IF(X15="","",W14)',
        'A40': '=IF(X16="","",W15)',
        'B10': '=Inputs!$B$8',
        'B28': '=IF(X3="","",X3)',
        'B29': '=IF(X4="","",X4)',
        'B30': '=IF(X5="","",X5)',
        'B31': '=IF(X6="","",X6)',
        'B32': '=IF(X7="","",X7)',
        'B33': '=IF(X8="","",X8)',
        'B34': '=IF(X9="","",X9)',
        'B35': '=IF(X11="","",X11)',
        'B36': '=IF(X12="","",X12)',
        'B37': '=IF(X13="","",X13)',
        'B38': '=IF(X14="","",X14)',
        'B39': '=IF(X15="","",X15)',
        'B40': '=IF(X16="","",X16)',
        'B5': '=Inputs!$B$20',
        'B6': '=Inputs!$B$7',
        'B7': '=Inputs!$B$11',
        'B8': '=Inputs!$B$12',
        'B9': '=Inputs!$B$17',
        'C28': '=IF(X3="","",Y3)',
        'C29': '=IF(X4="","",Y4)',
        'C30': '=IF(X5="","",Y5)',
        'C31': '=IF(X6="","",Y6)',
        'C32': '=IF(X7="","",Y7)',
        'C33': '=IF(X8="","",Y8)',
        'C34': '=IF(X9="","",Y9)',
        'C35': '=IF(X11="","",Y11)',
        'C36': '=IF(X12="","",Y12)',
        'C37': '=IF(X13="","",Y13)',
        'C38': '=IF(X14="","",Y14)',
        'C39': '=IF(X15="","",Y15)',
        'C40': '=IF(X16="","",Y16)',
        'D28': '=IF(X3="","",Z3)',
        'D33': '=IF(X8="","",Z8)',
        'D34': '=IF(X9="","",Z9)',
        'D35': '=IF(X11="","",Z11)',
        'D36': '=IF(X12="","",Z12)',
        'D37': '=IF(X13="","",Z13)',
        'D38': '=IF(X14="","",Z14)',
        'D39': '=IF(X15="","",Z15)',
        'D40': '=IF(X16="","",Z16)',
        'F28': '=D28*((1+$B$4)^(E28/365)-1)',
        'F29': '=D29*((1+$B$4)^(E30/365)-1)',
        'F30': '=D30*((1+$B$4)^(E31/365)-1)',
        'F31': '=D31*((1+$B$4)^(E32/365)-1)',
        'F32': '=D32*((1+$B$4)^(E33/365)-1)',
        'G28': '=D28+F28',
        'J28': '=IF(X3="","",W3)',
        'J29': '=IF(X4="","",W4)',
        'J30': '=IF(X5="","",W5)',
        'J31': '=IF(X6="","",W6)',
        'J32': '=IF(X7="","",W7)',
        'J33': '=IF(X8="","",W8)',
        'J34': '=IF(X9="","",W9)',
        'J35': '=IF(X11="","",W10)',
        'J36': '=IF(X12="","",W11)',
        'J37': '=IF(X13="","",W12)',
        'J38': '=IF(X14="","",W13)',
        'J39': '=IF(X15="","",W14)',
        'J40': '=IF(X16="","",W15)',
        'K28': '=IF(X3="","",X3)',
        'K29': '=IF(X4="","",X4)',
        'K30': '=IF(X5="","",X5)',
        'K31': '=IF(X6="","",X6)',
        'K32': '=IF(X7="","",X7)',
        'K33': '=IF(X8="","",X8)',
        'K34': '=IF(X9="","",X9)',
        'K35': '=IF(X11="","",X11)',
        'K36': '=IF(X12="","",X12)',
        'K37': '=IF(X13="","",X13)',
        'K38': '=IF(X14="","",X14)',
        'K39': '=IF(X15="","",X15)',
        'K40': '=IF(X16="","",X16)',
        'L28': '=IF(X3="","",Y3)',
        'L29': '=IF(X4="","",Y4)',
        'L30': '=IF(X5="","",Y5)',
        'L31': '=IF(X6="","",Y6)',
        'L32': '=IF(X7="","",Y7)',
        'L33': '=IF(X8="","",Y8)',
        'L34': '=IF(X9="","",Y9)',
        'L35': '=IF(X11="","",Y11)',
        'L36': '=IF(X12="","",Y12)',
        'L37': '=IF(X13="","",Y13)',
        'L38': '=IF(X14="","",Y14)',
        'L39': '=IF(X15="","",Y15)',
        'L40': '=IF(X16="","",Y16)',
        'M28': '=IF(X3="","",Z3)',
        'M29': '=IF(X4="","",Z4)',
        'M30': '=IF(X5="","",Z5)',
        'M31': '=IF(X6="","",Z6)',
        'M32': '=IF(X7="","",Z7)',
        'M33': '=IF(X8="","",Z8)',
        'M34': '=IF(X9="","",Z9)',
        'M35': '=IF(X11="","",Z11)',
        'M36': '=IF(X12="","",Z12)',
        'M37': '=IF(X13="","",Z13)',
        'M38': '=IF(X14="","",Z14)',
        'M39': '=IF(X15="","",Z15)',
        'M40': '=IF(X16="","",Z16)',
        'N28': '=IF(K28="","",0)',
        'O28': '=IF(K28="","",0)',
        'P28': '=IF(K28="","",IF(L28="LOAN",M28,IF(L28="REPAY",-M28,0)))',
        'S11': '=IF(B21="","",B21)',
        'S12': '=IF(B22="","",B22)',
        'S13': '=IF(B23="","",B23)',
        'S3': '=IF(B14="","",B14)',
        'S8': '=IF(B19="","",B19)',
        'S9': '=IF(B20="","",B20)',
        'T11': '=IF(B21="","","LOAN")',
        'T12': '=IF(B22="","","LOAN")',
        'T13': '=IF(B23="","","LOAN")',
        'T3': '=IF(B14="","","LOAN")',
        'T9': '=IF(B20="","","LOAN")',
        'U11': '=IF(C21="","",C21)',
        'U12': '=IF(C22="","",C22)',
        'U13': '=IF(C23="","",C23)',
        'U3': '=IF(C14="","",C14)',
        'U4': '=IF(C15="","",C15)',
        'U5': '=IF(C16="","",C16)',
        'U6': '=IF(C17="","",C17)',
        'U9': '=IF(C20="","",C20)',
        'V11': '=IF(B21="","",B21+0.1)',
        'V12': '=IF(B22="","",B22+0.1)',
        'V13': '=IF(B23="","",B23+0.1)',
        'V3': '=IF(B14="","",B14+0.1)',
        'V4': '=IF(B15="","",B15+0.1)',
        'V5': '=IF(B16="","",B16+0.1)',
        'V6': '=IF(B17="","",B17+0.1)',
        'V9': '=IF(B20="","",B20+0.1)',
        'X11': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),8),"")',
        'X12': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),9),"")',
        'X13': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),10),"")',
        'X14': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),11),"")',
        'X15': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),12),"")',
        'X16': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),13),"")',
        'X17': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),14),"")',
        'X18': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),15),"")',
        'X19': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),16),"")',
        'X20': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),17),"")',
        'X21': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),18),"")',
        'X22': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),19),"")',
        'X23': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),20),"")',
        'X24': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),21),"")',
        'X25': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),22),"")',
        'X26': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),23),"")',
        'X27': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),24),"")',
        'X28': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),25),"")',
        'X29': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),26),"")',
        'X3': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),1),"")',
        'X30': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),27),"")',
        'X31': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),28),"")',
        'X32': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),29),"")',
        'X33': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),30),"")',
        'X34': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),31),"")',
        'X35': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),32),"")',
        'X36': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),33),"")',
        'X37': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),34),"")',
        'X38': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),35),"")',
        'X39': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),36),"")',
        'X4': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),2),"")',
        'X40': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),37),"")',
        'X5': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),3),"")',
        'X6': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),4),"")',
        'X7': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),5),"")',
        'X8': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),6),"")',
        'X9': '=IFERROR(SMALL(IF($S$3:$S$64<>"",$S$3:$S$64),7),"")',
    },
    'col_format': {
        'B': 'mm/dd/yyyy',
        'D': '#,##0.00',
        'E': '#,##0.00',
        'F': '#,##0.00',
        'G': '#,##0.00',
        'K': 'mm/dd/yyyy',
        'M': '#,##0.00',
        'N': '#,##0.00',
        'O': '#,##0.00',
        'P': '#,##0.00',
        'S': 'mm/dd/yyyy',
        'U': '#,##0.00',
        'X': 'mm/dd/yyyy',
        'Z': '#,##0.00',
    },
}


# ========================================================================
# SHARED HELPERS (formula-template rendering, sheet builders)
# ========================================================================
ROW_PLACEHOLDER_RE = re.compile(r"\{R([+-]\d+)\}")


def render_template(template: str, row: int) -> str:
    """Substitute {R+n}/{R-n} placeholders with the actual row number."""
    return ROW_PLACEHOLDER_RE.sub(lambda m: str(row + int(m.group(1))), template)


def to_number_if_possible(s):
    if not isinstance(s, str):
        return s
    s = s.strip()
    if s == "":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except ValueError:
        return s


def load_csv_grid(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    width = max(len(r) for r in rows)
    return [r + [""] * (width - len(r)) for r in rows]


def build_cso_sheet(ws, grid):
    for r, row in enumerate(grid):
        excel_row = r + 1
        for c, raw in enumerate(row):
            excel_col = c + 1
            val = raw.strip() if isinstance(raw, str) else raw
            if val == "":
                continue
            if r <= 3:
                value = val  # header rows / concatenated-code row: text
            else:
                value = to_number_if_possible(val)
            ws.cell(row=excel_row, column=excel_col, value=value)


def build_functional_sheet(ws, sheet_def):
    # 1. static label/header cells
    for coord, value in sheet_def["statics"].items():
        ws[coord] = value

    # 2. template-generated formulas -- track exactly which rows were
    #    written per column so number formats (step 4) only touch those.
    formula_rows_by_col = {}
    for col_letter, groups in sheet_def["col_templates"].items():
        rows_written = formula_rows_by_col.setdefault(col_letter, set())
        for g in groups:
            template = g["template"]
            for a, b in g["ranges"]:
                for row in range(a, b + 1):
                    ws[f"{col_letter}{row}"] = render_template(template, row)
                    rows_written.add(row)

    # 3. explicit one-off formulas
    for coord, formula in sheet_def["explicit"].items():
        ws[coord] = formula
        m = re.match(r"^([A-Z]+)(\d+)$", coord)
        if m:
            formula_rows_by_col.setdefault(m.group(1), set()).add(int(m.group(2)))

    # 4. number formats -- applied ONLY to the rows that actually got a
    #    formula from that column, since some sheets (e.g. Loan) reuse
    #    the same column letter for structurally different data blocks
    #    at different row ranges with their own static header cells.
    for col_letter, fmt in sheet_def.get("col_format", {}).items():
        if not fmt or fmt == "General":
            continue
        for row in formula_rows_by_col.get(col_letter, ()):
            ws[f"{col_letter}{row}"].number_format = fmt


# ========================================================================
# 'Inputs' sheet builder (label-matched, same as calc_engine_34.py)
# ========================================================================
LABEL_TO_ROW = {
    "Mortality Table": 2,
    "Underwriting Class": 3,
    "Plan Code": 4,
    "Policy Number": 5,
    "Gender": 6,
    "Premium Paying Term (PPT)": 8,
    "Units": 9,
    "PUA Units": 10,
    "Issue Date": 11,
    "Valuation Date": 12,
    "DOB": 13,
    "Declared Dividend Rate (p.a.)": 18,
    "Dividend Option": 19,
    "Accumulation Interest Rate (p.a.)": 20,
    "Annual Base Premium": 21,
    "Premium Mode": 23,
    "Contract Status": 25,
}
DATE_LABELS = {"Issue Date", "Valuation Date", "DOB"}
PERCENT_LABELS = {"Declared Dividend Rate (p.a.)", "Accumulation Interest Rate (p.a.)"}
TEXT_LABELS = {"Policy Number", "Underwriting Class", "Gender", "Dividend Option", "Premium Mode", "Contract Status"}

_DATE_FORMATS = ["%d-%m-%Y", "%d-%m-%y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"]


def _parse_date(raw: str):
    raw = raw.strip()
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(raw, fmt)
        except ValueError:
            continue
    raise ValueError(f"Could not parse date: {raw!r} (expected one of {_DATE_FORMATS})")


def _parse_percent(raw: str):
    raw = raw.strip().replace("%", "")
    val = float(raw)
    # Inputs_v2.csv may already store this as a fraction (0.025) or as a
    # percent-looking number (2.5). Reference workbook stores a fraction.
    return val / 100.0 if val > 1 else val


def _parse_number_or_text(raw: str):
    raw = raw.strip()
    try:
        if "." in raw:
            return float(raw)
        return int(raw)
    except ValueError:
        return raw


def load_label_value_csv(path: str):
    """Reads a 2-column label,value CSV (main_v2.build_inputs_csv()'s
    output format) into a {label: raw_string} dict. Ignores any extra
    columns and a header row if the first cell isn't a known label."""
    values = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if len(row) < 2:
                continue
            label, value = row[0].strip(), row[1]
            if value is None:
                continue
            value = str(value).strip()
            if label == "" or value == "":
                continue
            values[label] = value
    return values


# ----------------------------------------------------------------------
# Build the 'Inputs' sheet from the label/value dict + INPUTS_FORMULAS.
# ----------------------------------------------------------------------
def build_inputs_sheet_from_labels(ws, label_values: dict):
    for label, row in LABEL_TO_ROW.items():
        if label not in label_values:
            continue  # not supplied for this policy -- leave blank
        raw = label_values[label]
        coord = f"B{row}"
        if coord in INPUTS_FORMULAS:
            continue  # safety net, shouldn't happen given LABEL_TO_ROW above

        if label in DATE_LABELS:
            value = _parse_date(raw)
        elif label in PERCENT_LABELS:
            value = _parse_percent(raw)
        elif label in TEXT_LABELS:
            value = raw  # keep as text -- preserves leading zeros on policy number
        else:
            value = _parse_number_or_text(raw)
        ws.cell(row=row, column=2, value=value)

    for coord, spec in INPUTS_FORMULAS.items():
        if spec[0] == "formula":
            ws[coord] = spec[1]
        else:
            ws[coord] = ArrayFormula(spec[2], spec[1])

    for coord, fmt in INPUTS_FORMATS.items():
        ws[coord].number_format = fmt

    # --- D-column labels for the native summary block (rows 2-6) -------
    # These are static text in the reference workbook's 'Inputs' sheet,
    # paired with the formula cells in INPUTS_FORMULAS above (E2:E6).
    # main_v2.extract_summary_from_inputs_de() matches on this label
    # text, so it must be present for Cash Value/PUA Cash Value/ETI/
    # RPU/APL to show up in the summary.
    #
    # Only D1:E6 are written here -- confirmed against the reference
    # workbook's actual cell fill colors: D1:E6 is filled green
    # (FFB3FBB6, the "Output" block), while everything else on this
    # sheet outside the raw input rows (dividend lookup table, matched
    # policies table, legend/tag columns) has no place here since it's
    # neither a CSV input nor part of the green-marked output block.
    #
    # NOTE: 'Loan Outstanding Advance' / 'Loan Outstanding Arrear' rows
    # (previously added at D28:E29) have been removed -- they are not
    # part of the original reference workbook and are not green-marked.
    # If your main_v2.py pipeline's SUMMARY_LABELS still expects those
    # two fields, you'll need to either add them back here or source
    # them from elsewhere; main_v2.extract_summary_from_inputs_de()
    # will simply return None for any label it doesn't find.
    ws["D1"] = "Output"
    ws["D2"] = "Cash Value"
    ws["D3"] = "PUA Cash Value"
    ws["D4"] = "ETI"
    ws["D5"] = "RPU"
    ws["D6"] = "APL"


# ========================================================================
# ALIGN DRIVER: open v9's output, strip it back to a shell, rebuild every
# sheet from the verified data above.
# ========================================================================
REQUIRED_SHEETS = [
    "Inputs", "CSO Rates", "MORTALITY", "CV_RATES",
    "PUA_TABLE", "TERM_RATES", "DIvidend_Assumption", "Projection_Monthly",
    "RPU", "ETI", "APL", "Commutation", "Loan",
]


def strip_and_rebuild_sheets(wb):
    """Delete every sheet name v9 might have created and recreate only the
    14 that make up the true module 34 structure, in the correct order."""
    for name in list(wb.sheetnames):
        del wb[name]
    for name in REQUIRED_SHEETS:
        wb.create_sheet(name)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", required=True, help="xlsx produced by calc_engine_v9.py")
    ap.add_argument("--inputs", required=True, help="REAL (unpadded) module 34 Inputs_v2.csv")
    ap.add_argument("--cso", required=True, help="REAL module 34 cso_rates.csv")
    ap.add_argument("--output", required=True, help="Path to write the aligned .xlsx")
    args = ap.parse_args()

    if not os.path.exists(args.source):
        raise FileNotFoundError(
            f"Source workbook not found: {args.source}\n"
            f"(this should be whatever calc_engine_v9.py just wrote, e.g. WL_model_34.xlsx)"
        )

    print(f"[1/5] Opening v9's output: {args.source}")
    wb = load_workbook(args.source, data_only=False)
    original_sheets = wb.sheetnames
    print(f"       Found sheets: {original_sheets}")

    print("[2/5] Stripping back to a clean shell (removing all v9 content/sheets) ...")
    strip_and_rebuild_sheets(wb)

    print(f"[3/5] Parsing {args.inputs} / {args.cso} and rebuilding 'Inputs' + 'CSO Rates' ...")
    label_values = load_label_value_csv(args.inputs)
    build_inputs_sheet_from_labels(wb["Inputs"], label_values)
    build_cso_sheet(wb["CSO Rates"], load_csv_grid(args.cso))

    print("[4/5] Rebuilding the 11 formula-driven calculation sheets ...")
    for sheet_name, sheet_def in FUNCTIONAL_SHEETS.items():
        build_functional_sheet(wb[sheet_name], sheet_def)

    print(f"[5/5] Saving aligned workbook: {args.output}")
    wb.save(args.output)
    print("Done.")


if __name__ == "__main__":
    main()
